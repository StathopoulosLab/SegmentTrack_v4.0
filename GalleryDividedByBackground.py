"""This function show gallery of plots of intensity traces divided by bkg.


It takes as input the .xls file of spots intensity divided by bkg spatially
selected and crea   tes the gallery.
"""


import numpy as np
import time
import pyqtgraph as pg
from openpyxl import load_workbook

from PyQt5 import QtWidgets


class GalleryDividedByBackground:
    def __init__(self, foldername):

        wb    =  load_workbook(foldername + '/DividedByBackground_SpatialSelected.xlsx')
        s_wb  =  wb[wb.sheetnames[0]]

        col_idx_in  =  1
        while str(s_wb.cell(column=col_idx_in, row=1).value)[:5] != 'SptId':
            col_idx_in  +=  1

        for col_idx_fin in range(col_idx_in, s_wb.max_column + 1):
            if str(s_wb.cell(column=col_idx_fin, row=1).value)[:5] != 'SptId':
                col_idx_fin  +=  1
                break

        row_idx_fin  =  2
        while row_idx_fin < s_wb.max_row and s_wb.cell(column=col_idx_in, row=row_idx_fin).value != '':
            row_idx_fin  +=  1

        spts_id  =  []
        for kk in range(col_idx_in, col_idx_fin + 1):
            spts_id.append(str(s_wb.cell(column=kk, row=1).value)[6:])

        pbar  =  ProgressBar(total1=len(spts_id))
        pbar.update_progressbar(0)
        pbar.show()

        spts_vals  =  np.zeros((len(spts_id), row_idx_fin - 1))
        for ii in range(spts_vals.shape[0]):
            pbar.update_progressbar(ii)
            for t in range(spts_vals.shape[1]):
                spts_vals[ii, t]  =  s_wb.cell(column=ii + col_idx_in, row=t + 2).value

        pbar.close()
        y_sup  =  spts_vals.max()

        n_rows   =  6
        n_cols   =  7
        num_win  =  len(spts_id) // (n_cols * n_rows) + 1
        for win_idxs in range(num_win):
            time.sleep(3)
            str_win1  =  "win" + str(win_idxs) + "  =  pg.GraphicsWindow()"
            str_win2  =  "win" + str(win_idxs) + ".setWindowTitle('Transcriptional Traces " + str(win_idxs + 1) + "')"
            str_win3  =  "win" + str(win_idxs) + ".showMaximized()"
            exec(str_win1)
            exec(str_win2)
            exec(str_win3)

            idx_name = 0
            for idx_r in range(n_rows):
                for idx_c in range(n_cols):
                    str_addplot  =  "p" + str(idx_name) +  "= win" + str(win_idxs) + ".addPlot(row=" + str(idx_r) + ", col=" + str(idx_c) + ")"
                    exec(str_addplot)
                    idx_name  +=  1

            for k in range(n_cols * n_rows):
                if k + win_idxs * n_cols * n_rows <= len(spts_id) - 1:
                    str_cmnd1  =  "p" + str(k) + ".plot(spts_vals[k  + win_idxs * n_cols * n_rows, :], pen='r', symbol='o', symbolSize=2)"
                    str_cmnd2  =  "p" + str(k) + ".setYRange(0, y_sup)"
                    str_cmnd3  =  "tag_text" + str(k) + " = pg.TextItem('tag = ' + spts_id[k + win_idxs * n_cols * n_rows], color='g')"
                    str_cmnd4  =  "tag_text" + str(k) + ".setPos(1, y_sup)"
                    str_cmnd5  =  "p" + str(k) + ".addItem(tag_text" + str(k) + ")"
                    exec(str_cmnd1)
                    exec(str_cmnd2)
                    exec(str_cmnd3)
                    exec(str_cmnd4)
                    exec(str_cmnd5)
                else:
                    break
        print(StrangePatch)


class ProgressBar(QtWidgets.QWidget):
    """Simple Progress bar widget"""
    def __init__(self, parent=None, total1=20):
        super().__init__(parent)
        self.name_line1  =  QtWidgets.QLineEdit()

        self.progressbar  =  QtWidgets.QProgressBar()
        self.progressbar.setMinimum(1)
        self.progressbar.setMaximum(total1)

        main_layout  =  QtWidgets.QGridLayout()
        main_layout.addWidget(self.progressbar, 0, 0)

        self.setLayout(main_layout)
        self.setWindowTitle("Progress")
        self.setGeometry(500, 300, 300, 50)

    def update_progressbar(self, val1):
        """Progress bar updater"""
        self.progressbar.setValue(val1)
        QtWidgets.qApp.processEvents()
