"""This is the main window of the software to exytract and analyze bursty time series.

This is version 4.0. Since October 2021.
"""

import os
import sys
import glob
import time
import traceback
from importlib import reload
from functools import reduce
import numpy as np
import pyqtgraph as pg
import tifffile
import skimage.morphology as skmr
from PyQt5.QtCore import Qt
from PyQt5 import QtGui, QtWidgets, QtCore
from openpyxl import load_workbook

import AnalysisConverter
import NucleiDetectLog
import NucleiDetect
import NucleiSegmentStackMultiCore
import NucleiConnectMultiCore
import NucleiSpotsConnection
import MultiLoadLsmOrTif5D
import MultiLoadCzi5D
import MultiLoadLif5D
import SpotsConnection
import LabelsModify
import AnalysisSaver
import SpotsDetectionChopper
import SpotsDetection3D
import ParametersExtraction
import FalseColored3Ch
import BurstStatisticWriter
import FromTile2GlobCoordinate
import FromJournal2Fitting
import RemoveBadNuclei
import AnalysisLoader
import MultiTracePlotting_v2
import VisualNucSpot_v2
import ComprehensiveBurstAnalysisWriter
import ComprehensiveActivationWriter
import RemoveShortTraces2
import MergeXlsFiles
import WriteSteadySpotsBursting
import MultiColorIntensityGenerator
import WriteSptsIntsMinusBkg
import WriteSptsIntsDividedByBkg
import AvSpotsTime
import TimeavSptsIntensity
import SpatialDividedByBkg
import PopUpTool
import GalleryDividedByBackground
import CalibrationSpatial
import RescueFunctions
import SpotTracker
import VisualTracked
import TracesImage


class MainWindow(QtWidgets.QMainWindow):
    """Main windows: coordinates all the actions, algorithms, visualization tools and analysis tools"""

    def __init__(self, parent=None):

        QtWidgets.QMainWindow.__init__(self, parent)

        widget  =  QtWidgets.QWidget(self)
        self.setCentralWidget(widget)

        exit_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/exit.png'), '&Exit', self)
        exit_action.setShortcut('Ctrl+Q')
        exit_action.setStatusTip("Exit application")
        exit_action.triggered.connect(self.close)

        multiload_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/load-hi.png'), "&Load data file", self)
        multiload_action.setShortcut('Ctrl+L')
        multiload_action.setStatusTip("Load .tif, .lsm, .czi or .lif files: if they are more than one, they will be concatenate")
        multiload_action.triggered.connect(self.load_several_files)

        load_analysis_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/load-hi.png'), "&Load analysis", self)
        load_analysis_action.setShortcut('Ctrl+A')
        load_analysis_action.setStatusTip("Load analysis")
        load_analysis_action.triggered.connect(self.load_analysis)

        save_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/save-md.png'), "&Save analysis", self)
        save_action.setShortcut('Ctrl+S')
        save_action.setStatusTip("Save analysis")
        save_action.triggered.connect(self.find_zero_in_mtss)

        rescue_analysis_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/float-blisters.png'), "&Rescue analysis", self)
        rescue_analysis_action.setShortcut('Ctrl+K')
        rescue_analysis_action.setStatusTip("Rescue analysis")
        rescue_analysis_action.triggered.connect(self.rescue_analysis)

        popup_nuclei_raw_action  =  QtWidgets.QAction("&Pop-up Raw Nuclei", self)
        popup_nuclei_raw_action.setStatusTip("Generate figures with the raw nuclei data")
        popup_nuclei_raw_action.triggered.connect(self.popup_nuclei_raw)

        popup_nuclei_detected_action  =  QtWidgets.QAction("&Pop-up Detected Nuclei", self)
        popup_nuclei_detected_action.setStatusTip("Generate figures with the detected nuclei")
        popup_nuclei_detected_action.triggered.connect(self.popup_nuclei_detected)

        popup_nuclei_segmented_action  =  QtWidgets.QAction("&Pop-up Segmented Nuclei", self)
        popup_nuclei_segmented_action.setStatusTip("Generate figures with the detected nuclei")
        popup_nuclei_segmented_action.triggered.connect(self.popup_nuclei_segmented)

        popup_nuclei_tracked_action  =  QtWidgets.QAction("&Pop-up Tracked Nuclei", self)
        popup_nuclei_tracked_action.setStatusTip("Generate figures with the tracked nuclei")
        popup_nuclei_tracked_action.triggered.connect(self.popup_nuclei_trackeded)

        popup_spots_raw_action  =  QtWidgets.QAction("&Pop-up Raw Spots", self)
        popup_spots_raw_action.setStatusTip("Generate figures with the raw spots")
        popup_spots_raw_action.triggered.connect(self.popup_spots_raw)

        popup_spots_segm_action  =  QtWidgets.QAction("&Pop-up Segmented Spots", self)
        popup_spots_segm_action.setStatusTip("Generate figures with the segmented spots")
        popup_spots_segm_action.triggered.connect(self.popup_spots_segm)

        popup_nucactive_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/popup.png'), "&Pop-up Active Nuclei", self)
        popup_nucactive_action.setStatusTip("Generate a figure with the active nuclei map")
        popup_nucactive_action.triggered.connect(self.popup_nucactive)

        del_frame_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/delete_frame.png'), "&Delete Current Frame", self)
        del_frame_action.setStatusTip("Delete the current frame from raw data")
        del_frame_action.triggered.connect(self.del_frame)

        roi_crop_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/crop.png'), "&Crop Stack", self)
        roi_crop_action.setStatusTip("Define a ROI to crop data stack")
        roi_crop_action.triggered.connect(self.roi_crop)

        activ_dynamic_fitting_action  =  QtWidgets.QAction("&Activation Dynamic", self)
        activ_dynamic_fitting_action.setStatusTip("Perform an analysis on the dynamic of nuclei activation")
        activ_dynamic_fitting_action.triggered.connect(self.act_dynamic_study)

        test_spots_detection_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/sccpre.png'), "&Test Spots Detection", self)
        test_spots_detection_action.setStatusTip("Test the spots detection on a sample of your data")
        test_spots_detection_action.triggered.connect(self.test_spots_detection)

        test_nuclei_detection_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/sccpre.png'), "&Test Nuclei Detection", self)
        test_nuclei_detection_action.setStatusTip("Test the nuclei detection on a sample of your data")
        test_nuclei_detection_action.triggered.connect(self.test_nuclei_detection)
        test_nuclei_detection_action.setShortcut("Ctrl+N")

        spatial_comprehensive_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/geo_pin.jpg'), "&Spatial Comprehensive", self)
        spatial_comprehensive_action.setStatusTip('Perform a comprehensive spatial analysis for bursting')
        spatial_comprehensive_action.triggered.connect(self.spatial_comprehensive)
        spatial_comprehensive_action.setShortcut('Ctrl+H')

        single_nucleus_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/plot_single_trace.png'), "&Single Nucleus Inspection", self)
        single_nucleus_action.setStatusTip('Perform a comprehensive spatial analysis for bursting')
        single_nucleus_action.triggered.connect(self.single_nucleus)
        single_nucleus_action.setShortcut('Ctrl+B')

        multi_plot_show_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/plot_multi_trace.png'), "&Show Multi Plot", self)
        multi_plot_show_action.setStatusTip("Inspection of All Spots Traces")
        multi_plot_show_action.triggered.connect(self.multi_plot_show)
        multi_plot_show_action.setShortcut('Ctrl+M')

        set_color_channel_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/set_channels.png'), "&Set Channels Numbers", self)
        set_color_channel_action.setStatusTip("&Set Channels Numbers")
        set_color_channel_action.triggered.connect(self.set_color_channel)

        multi_color_intensity_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/multi_color_icon.png'), "&Multi Color Intensity", self)
        multi_color_intensity_action.setStatusTip("Visualize spots intensity as color on nuclei")
        multi_color_intensity_action.triggered.connect(self.multi_color_intensity)
        multi_color_intensity_action.setShortcut('Ctrl+Y')

        spts_intensity_minus_bkg_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/background_remove.jpg'), "&Spots Remove Bkg", self)
        spts_intensity_minus_bkg_action.setStatusTip("Write an .xls file with the intensity of the spots minus the background")
        spts_intensity_minus_bkg_action.triggered.connect(self.spts_intensity_minus_bkg)

        timeav_spts_intensity_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/clessidra.svg'), "&Time Average Spots by Bkg", self)
        timeav_spts_intensity_action.setStatusTip("Write an .xls file with the average time intensity of the spots normalized by the background")
        timeav_spts_intensity_action.triggered.connect(self.timeav_spts_intensity)
        timeav_spts_intensity_action.setShortcut('Ctrl+E')

        check_saturation_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/check_saturation.png'), "&Check Saturation", self)
        check_saturation_action.setStatusTip('Inspection of All Spots Traces')
        check_saturation_action.triggered.connect(self.check_saturation)

        calibration_spatial_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/caliber.png'), '&Traces Calibration', self)
        calibration_spatial_action.setStatusTip("Shows the gallery of the spots spatially selected divided by the background")
        calibration_spatial_action.triggered.connect(self.calibration_spatial)

        traces_image_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/foot_traces.png'), '&Traces Image', self)
        traces_image_action.setStatusTip("Plot the traces as a image, even pooling togheter several analyses")
        traces_image_action.triggered.connect(self.traces_image)

        rmv_mitoticalTS_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/eraser.png'), '&Remove Mitotical TS', self)
        rmv_mitoticalTS_action.setStatusTip("Pop up tool to remove mitotical spots")
        rmv_mitoticalTS_action.triggered.connect(self.rmv_mitoticalTS)
        rmv_mitoticalTS_action.setShortcut('Ctrl+U')

#         sisters_split_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/sisters_split.png'), '&Allele Split', self)
#         sisters_split_action.setStatusTip("Split the sister chromatine tracking")
#         sisters_split_action.triggered.connect(self.sisters_split)
#         sisters_split_action.setShortcut('Ctrl+G')

        remove_nucsdust_action  =  QtWidgets.QAction(QtGui.QIcon('Icons/duster.png'), '&Remove Nuclear Dust', self)
        remove_nucsdust_action.setStatusTip("Remove small objects detected in the nuclei channel")
        remove_nucsdust_action.triggered.connect(self.remove_nucsdust)
        remove_nucsdust_action.setShortcut('Ctrl+O')

        analysis_conversion_action  =  QtWidgets.QAction('&Analysis Converter', self)
        analysis_conversion_action.setStatusTip("Convert analysis 3.3 -> 4.0")
        analysis_conversion_action.triggered.connect(self.analysis_conversion)

        self.statusBar()

        menubar   =  self.menuBar()

        fileMenu  =  menubar.addMenu('&File')
        fileMenu.addAction(multiload_action)
        fileMenu.addAction(save_action)
        fileMenu.addAction(load_analysis_action)
        fileMenu.addAction(rescue_analysis_action)
        fileMenu.addAction(exit_action)

        modifyMenu  =  menubar.addMenu('&Modify')
        modifyMenu.addAction(roi_crop_action)
        modifyMenu.addAction(del_frame_action)
        modifyMenu.addAction(set_color_channel_action)
        modifyMenu.addAction(rmv_mitoticalTS_action)
        modifyMenu.addAction(remove_nucsdust_action)

        testMenu  =  menubar.addMenu('&Test Analysis')
        testMenu.addAction(test_nuclei_detection_action)
        testMenu.addAction(test_spots_detection_action)

        checkMenu  =  menubar.addMenu('&Check Data')
        checkMenu.addAction(check_saturation_action)

        posterioriMenu  =  menubar.addMenu('&Post Processing')
        posterioriMenu.addAction(spatial_comprehensive_action)
        posterioriMenu.addAction(single_nucleus_action)
        posterioriMenu.addAction(multi_plot_show_action)
        posterioriMenu.addAction(multi_color_intensity_action)
        posterioriMenu.addAction(spts_intensity_minus_bkg_action)
        posterioriMenu.addAction(timeav_spts_intensity_action)
        posterioriMenu.addAction(calibration_spatial_action)
        posterioriMenu.addAction(traces_image_action)
#         posterioriMenu.addAction(sisters_split_action)

        popupMenu     =  menubar.addMenu('&PopUp')
        popup_nuclei  =  popupMenu.addMenu(QtGui.QIcon('Icons/popup.png'), "PopUp Nuclei")
        popup_nuclei.addAction(popup_nuclei_raw_action)
        popup_nuclei.addAction(popup_nuclei_detected_action)
        popup_nuclei.addAction(popup_nuclei_segmented_action)
        popup_nuclei.addAction(popup_nuclei_tracked_action)

        popup_spots  =  popupMenu.addMenu(QtGui.QIcon('Icons/popup.png'), "PopUp Spots")
        popup_spots.addAction(popup_spots_raw_action)
        popup_spots.addAction(popup_spots_segm_action)
        popupMenu.addAction(popup_nucactive_action)

        momentaryMenu  =  menubar.addMenu("&Temporary")
        momentaryMenu.addAction(activ_dynamic_fitting_action)
        momentaryMenu.addAction(analysis_conversion_action)

        # helpMenu  =  menubar.addMenu('&Help')

        frame1  =  pg.ImageView(self)
        frame1.ui.roiBtn.hide()
        frame1.ui.menuBtn.hide()

        frame2   =  pg.ImageView(self)
        frame2.ui.roiBtn.hide()
        frame2.ui.menuBtn.hide()
        frame2.ui.histogram.hide()

        frame3   =  pg.ImageView(self)
        frame3.ui.roiBtn.hide()
        frame3.ui.menuBtn.hide()

        frame4   =  pg.ImageView(self)
        frame4.ui.roiBtn.hide()
        frame4.ui.menuBtn.hide()
        frame4.ui.histogram.hide()

        fname_edt = QtWidgets.QLineEdit(self)
        fname_edt.setToolTip('Name of the file you are working on')

        sld1  =  QtWidgets.QSlider(QtCore.Qt.Horizontal, self)
        sld1.valueChanged.connect(self.sld1_update)

        start_cut_btn  =  QtWidgets.QPushButton("Start", self)
        start_cut_btn.clicked.connect(self.start_cut)
        start_cut_btn.setToolTip('Select the first frame to consider')
        start_cut_btn.setFixedSize(50, 25)

        end_cut_btn  =  QtWidgets.QPushButton("End", self)
        end_cut_btn.clicked.connect(self.end_cut)
        end_cut_btn.setToolTip('Select the last frame to consider')
        end_cut_btn.setFixedSize(50, 25)

        reload_cut_btn  =  QtWidgets.QPushButton("Reload", self)
        reload_cut_btn.clicked.connect(self.reload_files)
        reload_cut_btn.setToolTip('Reload selected files')
        reload_cut_btn.setFixedSize(110, 25)

        auto_run_btn  =  QtWidgets.QPushButton("Auto Run", self)
        auto_run_btn.clicked.connect(self.auto_run)
        auto_run_btn.setToolTip('Run all the unsupervised analysis in a row')
        auto_run_btn.setFixedSize(110, 25)
        auto_run_btn.setEnabled(False)

        auto_run_tggl = QtWidgets.QCheckBox('Enable Auto', self)
        auto_run_tggl.stateChanged.connect(self.auto_run_enable)
        auto_run_tggl.setFixedSize(110, 25)

        nuc_detect_btn  =  QtWidgets.QPushButton("N-Detect", self)
        nuc_detect_btn.clicked.connect(self.nuclei_detection)
        nuc_detect_btn.setToolTip('Nuclei detection')
        nuc_detect_btn.setFixedSize(110, 25)

        param_detect_lbl  =  QtWidgets.QLabel("Gauss Size", self)
        param_detect_lbl.setFixedSize(60, 25)

        param_detect_edt  =  QtWidgets.QLineEdit(self)
        param_detect_edt.textChanged[str].connect(self.param_detect_var)
        param_detect_edt.setToolTip('Sets the size of the Gaussian Filter for the pre-smoothing (1.5) or for the logharitmic detection (0.995 - 1.005)')
        param_detect_edt.setFixedSize(35, 25)

        param_detect_box  =  QtWidgets.QHBoxLayout()
        param_detect_box.addWidget(param_detect_lbl)
        param_detect_box.addWidget(param_detect_edt)

        gaus_log_detect_combo = QtWidgets.QComboBox(self)
        gaus_log_detect_combo.addItem("Gauss Flt")
        gaus_log_detect_combo.addItem("Log Flt")
        gaus_log_detect_combo.setCurrentIndex(0)
        gaus_log_detect_combo.setToolTip('Switch between a linear nuclei detection and a logaritmic nuclei detection')
        gaus_log_detect_combo.activated[str].connect(self.gaus_log_detect)

        nuc_segment_btn  =  QtWidgets.QPushButton("N-Segment", self)
        nuc_segment_btn.clicked.connect(self.nuclei_segmentation)
        nuc_segment_btn.setToolTip('Nuclei Segmentation')
        nuc_segment_btn.setFixedSize(110, 25)

        gfilt_water_lbl  =  QtWidgets.QLabel('W-Shed', self)
        gfilt_water_lbl.setFixedSize(60, 25)

        gfilt_water_edt  =  QtWidgets.QLineEdit(self)
        gfilt_water_edt.textChanged[str].connect(self.gfilt_water_var)
        gfilt_water_edt.setToolTip('Sets the size parameter for the Water Shed algorithm, suggested value 7')
        gfilt_water_edt.setFixedSize(35, 25)

        circ_thr_lbl  =  QtWidgets.QLabel('Circ Thr', self)
        circ_thr_lbl.setFixedSize(60, 25)

        circ_thr_edt  =  QtWidgets.QLineEdit(self)
        circ_thr_edt.textChanged[str].connect(self.circ_thr_var)
        circ_thr_edt.setToolTip('Circularity Threshold of the detected nuclei, suggested value is 0.65')
        circ_thr_edt.setFixedSize(35, 25)

        modify_segm_btn  =  QtWidgets.QPushButton("Modify Segm", self)
        modify_segm_btn.clicked.connect(self.modify_tool)
        modify_segm_btn.setToolTip('Modify Nuclei Segmentation')
        modify_segm_btn.setFixedSize(110, 25)

        nuc_track_btn  =  QtWidgets.QPushButton("N-Track", self)
        nuc_track_btn.clicked.connect(self.nuclei_tracking)
        nuc_track_btn.setToolTip('Nuclei Tracking')
        nuc_track_btn.setFixedSize(110, 25)

        dist_thr_lbl  =  QtWidgets.QLabel('Dist Thr', self)
        dist_thr_lbl.setFixedSize(60, 25)

        dist_thr_edt  =  QtWidgets.QLineEdit(self)
        dist_thr_edt.textChanged[str].connect(self.dist_thr_var)
        dist_thr_edt.setToolTip('Distance threshold to track nuclei; suggested value 10')
        dist_thr_edt.setFixedSize(35, 25)

        spots_thr_lbl  =  QtWidgets.QLabel('Spots Thr', self)
        spots_thr_lbl.setFixedSize(60, 25)

        spots_thr_edt  =  QtWidgets.QLineEdit(self)
        spots_thr_edt.textChanged[str].connect(self.spots_thr_var)
        spots_thr_edt.setToolTip('Intensity threshold to segment spots: it is expressed in terms of standard deviation, suggested value 7')
        spots_thr_edt.setFixedSize(35, 25)

        volume_thr_lbl  =  QtWidgets.QLabel('Vol Thr', self)
        volume_thr_lbl.setFixedSize(60, 25)

        volume_thr_edt  =  QtWidgets.QLineEdit(self)
        volume_thr_edt.textChanged[str].connect(self.volume_thr_var)
        volume_thr_edt.setToolTip('Threshold volume on spot detection: suggested value 5')
        volume_thr_edt.setFixedSize(35, 25)

        spots_detect_btn  =  QtWidgets.QPushButton("S-Detect", self)
        spots_detect_btn.clicked.connect(self.spots_detect)
        spots_detect_btn.setToolTip('Spots Detection')
        spots_detect_btn.setFixedSize(110, 25)

        spots_visual_btn  =  QtWidgets.QPushButton("Visualize Spots", self)
        spots_visual_btn.clicked.connect(self.spots_visual)
        spots_visual_btn.setToolTip('Updates Spots Detection')
        spots_visual_btn.setFixedSize(110, 25)

        nuc_spots_conn_btn = QtWidgets.QPushButton("S-N Connect", self)
        nuc_spots_conn_btn.clicked.connect(self.nuc_spots_conn)
        nuc_spots_conn_btn.setToolTip('Connect Spots to Nuclei')
        nuc_spots_conn_btn.setFixedSize(110, 25)

        fake_coloured_time_btn = QtWidgets.QPushButton("False Clrs Time", self)
        fake_coloured_time_btn.clicked.connect(self.fake_coloured_time)
        fake_coloured_time_btn.setToolTip('Generate false coloured movie with time information')
        fake_coloured_time_btn.setFixedSize(110, 25)

        time_step_lbl  =  QtWidgets.QLabel('T Step', self)
        time_step_lbl.setFixedSize(50, 25)

        time_step_edt  =  QtWidgets.QLineEdit(self)
        time_step_edt.textChanged[str].connect(self.time_step_var)
        time_step_edt.setToolTip('Duration in seconds of the time step')
        time_step_edt.setFixedSize(55, 25)

        busy_lbl  =  QtWidgets.QLabel("Ready")
        busy_lbl.setStyleSheet('color: green')

        busy_box  =  QtWidgets.QHBoxLayout()
        busy_box.addWidget(busy_lbl)
        busy_box.addStretch()

        time_lbl = QtWidgets.QLabel("time     " + '0', self)
        time_lbl.setFixedSize(110, 13)

        frame_numb_lbl = QtWidgets.QLabel("frame  " + '0', self)
        frame_numb_lbl.setFixedSize(110, 13)

        hor_line_one  =  QtWidgets.QFrame()
        hor_line_one.setFrameStyle(QtWidgets.QFrame.HLine)

        hor_line_two  =  QtWidgets.QFrame()
        hor_line_two.setFrameStyle(QtWidgets.QFrame.HLine)

        cut_box_h  =  QtWidgets.QHBoxLayout()
        cut_box_h.addWidget(start_cut_btn)
        cut_box_h.addWidget(end_cut_btn)

        cut_box  =  QtWidgets.QVBoxLayout()
        cut_box.addLayout(cut_box_h)
        cut_box.addWidget(reload_cut_btn)

        h1box  =  QtWidgets.QHBoxLayout()
        h1box.addWidget(frame1)
        h1box.addWidget(frame2)

        h2box  = QtWidgets.QHBoxLayout()
        h2box.addWidget(frame3)
        h2box.addWidget(frame4)

        v2box  =  QtWidgets.QVBoxLayout()
        v2box.addWidget(fname_edt)
        v2box.addLayout(h1box)
        v2box.addLayout(h2box)
        v2box.addWidget(sld1)
        v2box.addLayout(busy_box)

        key_ver1  =  QtWidgets.QVBoxLayout()
        key_ver1.addWidget(hor_line_one)
        key_ver1.addWidget(auto_run_tggl)
        key_ver1.addWidget(auto_run_btn)
        key_ver1.addWidget(hor_line_two)
        key_ver1.addLayout(param_detect_box)
        key_ver1.addWidget(gaus_log_detect_combo)
        key_ver1.addWidget(nuc_detect_btn)
        key_ver1.addStretch()

        gseg_hor  =  QtWidgets.QHBoxLayout()
        gseg_hor.addWidget(gfilt_water_lbl)
        gseg_hor.addWidget(gfilt_water_edt)

        circ_thr_hor  =  QtWidgets.QHBoxLayout()
        circ_thr_hor.addWidget(circ_thr_lbl)
        circ_thr_hor.addWidget(circ_thr_edt)

        key_ver2  =  QtWidgets.QVBoxLayout()
        key_ver2.addLayout(gseg_hor)
        key_ver2.addLayout(circ_thr_hor)
        key_ver2.addWidget(nuc_segment_btn)
        key_ver2.addStretch()

        dist_thr_hor  =  QtWidgets.QHBoxLayout()
        dist_thr_hor.addWidget(dist_thr_lbl)
        dist_thr_hor.addWidget(dist_thr_edt)

        key_modifing  =  QtWidgets.QVBoxLayout()
        key_modifing.addWidget(modify_segm_btn)

        key_ver3  =  QtWidgets.QVBoxLayout()
        key_ver3.addLayout(dist_thr_hor)
        key_ver3.addWidget(nuc_track_btn)
        key_ver3.addStretch()

        spots_thr_hor  =  QtWidgets.QHBoxLayout()
        spots_thr_hor.addWidget(spots_thr_lbl)
        spots_thr_hor.addWidget(spots_thr_edt)

        volume_thr_hor  =  QtWidgets.QHBoxLayout()
        volume_thr_hor.addWidget(volume_thr_lbl)
        volume_thr_hor.addWidget(volume_thr_edt)

        key_ver4  =  QtWidgets.QVBoxLayout()
        key_ver4.addLayout(spots_thr_hor)
        key_ver4.addLayout(volume_thr_hor)
        key_ver4.addWidget(spots_detect_btn)
        key_ver4.addStretch()
        key_ver4.addWidget(nuc_spots_conn_btn)
        key_ver4.addWidget(spots_visual_btn)
        key_ver4.addWidget(fake_coloured_time_btn)

        key_time_step  =  QtWidgets.QHBoxLayout()
        key_time_step.addWidget(time_step_lbl)
        key_time_step.addWidget(time_step_edt)

        key_tot  =  QtWidgets.QVBoxLayout()
        key_tot.addLayout(cut_box)
        key_tot.addStretch()
        key_tot.addLayout(key_ver1)
        key_tot.addLayout(key_ver2)
        key_tot.addLayout(key_modifing)
        key_tot.addLayout(key_ver3)
        key_tot.addStretch()
        key_tot.addLayout(key_ver4)
        key_tot.addStretch()
        key_tot.addLayout(key_time_step)
        key_tot.addWidget(time_lbl)
        key_tot.addWidget(frame_numb_lbl)

        layout   =  QtWidgets.QHBoxLayout(widget)
        layout.addLayout(v2box)
        layout.addLayout(key_tot)

        mycmap  =  np.fromfile("mycmap.bin", "uint16").reshape((10000, 3))   # / 255.0
        self.colors4map  =  []
        for k in range(mycmap.shape[0]):
            self.colors4map.append(mycmap[k, :])
        self.colors4map[0]  =  np.array([0, 0, 0])

        self.frame1  =  frame1
        self.frame2  =  frame2
        self.frame3  =  frame3
        self.frame4  =  frame4

        self.fname_edt              =  fname_edt
        self.sld1                   =  sld1
        self.nuc_segment_btn        =  nuc_segment_btn
        self.auto_run_btn           =  auto_run_btn
        self.nuc_detect_btn         =  nuc_detect_btn
        self.gfilt_water_lbl        =  gfilt_water_lbl
        self.gfilt_water_edt        =  gfilt_water_edt
        self.circ_thr_lbl           =  circ_thr_lbl
        self.circ_thr_edt           =  circ_thr_edt
        self.modify_segm_btn        =  modify_segm_btn
        self.nuc_track_btn          =  nuc_track_btn
        self.dist_thr_lbl           =  dist_thr_lbl
        self.dist_thr_edt           =  dist_thr_edt
        self.spots_thr_edt          =  spots_thr_edt
        self.volume_thr_edt         =  volume_thr_edt
        self.volume_thr_lbl         =  volume_thr_lbl
        self.spots_detect_btn       =  spots_detect_btn
        self.param_detect_lbl       =  param_detect_lbl
        self.param_detect_edt       =  param_detect_edt
        self.gaus_log_detect_combo  =  gaus_log_detect_combo
        self.time_step_edt          =  time_step_edt
        self.gaus_log_detect_value  =  "Gauss Flt"

        self.data_flag             =  0
        self.labbs_flag            =  0
        self.nuclei_flag           =  0
        self.nuclei_t_visual_flag  =  0
        self.spots_segm_flag       =  0
        self.spots_trk_flag        =  0
        self.start_cut_value       =  0
        self.end_cut_value         =  0
        self.bm_dm_am              =  0
        self.log_flag              =  0
        self.nucs_spts_ch          =  np.array([1, 0])
        self.busy_lbl              =  busy_lbl
        self.t_track_end_value     =  0

        self.time_lbl          =  time_lbl
        self.frame_numb_lbl    =  frame_numb_lbl
        self.software_version  =  "SegmentTrackSingleCycleGUI_v4.0"

        self.setGeometry(100, 100, 1200, 800)
        self.setWindowTitle(self.software_version)
        self.setWindowIcon(QtGui.QIcon('Icons/MLL_Logo2.png'))
        self.show()

    def closeEvent(self, event):
        "Close the GUI, asking confirmation"
        quit_msg  =  "Are you sure you want to exit the program?"
        reply     =  QtWidgets.QMessageBox.question(self, 'Message', quit_msg, QtWidgets.QMessageBox.Yes, QtWidgets.QMessageBox.No)

        if reply == QtWidgets.QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()

    def busy_indicator(self):
        """Write a red text (BUSY) as a label on the GUI (bottom left)"""
        self.busy_lbl.setText("Busy")
        self.busy_lbl.setStyleSheet('color: red')

    def ready_indicator(self):
        """Write a green text (READY) as a label on the GUI (bottom left)"""
        self.busy_lbl.setText("Ready")
        self.busy_lbl.setStyleSheet('color: green')

    def auto_run_enable(self, state):
        """Enable the possibility to run all the  unsupervised part in a row"""
        if state == QtCore.Qt.Checked:
            self.auto_run_btn.setEnabled(True)
            self.nuc_detect_btn.setEnabled(False)
            self.nuc_segment_btn.setEnabled(False)
            self.spots_detect_btn.setEnabled(False)
        else:
            self.auto_run_btn.setEnabled(False)
            self.nuc_detect_btn.setEnabled(True)
            self.nuc_segment_btn.setEnabled(True)
            self.spots_detect_btn.setEnabled(True)

    def roi_crop(self):
        """Call CroppingTool Tool"""
        self.mpp8  =  CroppingTool(self.filedata.imarray_red, self.filedata.imarray_green)
        self.mpp8.show()
        self.mpp8.procStart.connect(self.crop_tool_sgnl)

    def load_several_files(self):
        """Load, concatenate and visualize raw data to start the analysis"""
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:
            self.fnames  =  QtWidgets.QFileDialog.getOpenFileNames(None, "Select czi (or lsm) data files to concatenate...", filter="*.lsm *.czi *.tif *.lif")[0]

            if str(self.fnames[0])[-3:] == 'lsm' or str(self.fnames[0])[-3:] == 'tif':
                self.filedata         =  MultiLoadLsmOrTif5D.MultiLoadLsmOrTif5D(self.fnames, self.nucs_spts_ch)
                self.time_step_edt.setText(str(self.filedata.time_step_value))

            if str(self.fnames[0])[-3:] == 'czi':
                self.filedata         =  MultiLoadCzi5D.MultiProcLoadCzi5D(self.fnames, self.nucs_spts_ch)
                self.time_step_edt.setText(str(self.filedata.time_step_value))

            if str(self.fnames[0])[-3:] == 'lif':
                self.filedata         =  MultiLoadLif5D.MultiLoadLif5D(self.fnames, self.nucs_spts_ch)
                self.time_step_edt.setText(str(self.filedata.time_step_value))

            self.frame1.setImage(self.filedata.imarray_red[0, :, :])
            self.frame3.setImage(self.filedata.imarray_green[0, :, :])
            self.data_flag  =  1
            self.sld1.setMaximum(self.filedata.imarray_red.shape[0] - 1)
            self.sld1.setValue(0)

            joined_fnames  = ' '
            for s in range(len(self.fnames)):
                joined_fnames  +=  str(self.fnames[s]) +  ' ----- '

            self.fname_edt.setText(joined_fnames)

        except Exception:
            traceback.print_exc()

        self.ready_indicator()

    def load_analysis(self):
        """Load an already done analysis"""
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:
            foldername     =  str(QtWidgets.QFileDialog.getExistingDirectory(None, "Select the folder with the analyzed data"))
            self.fname_edt.setText(foldername)
            app.processEvents()
            app.processEvents()

            self.filedata     =  AnalysisLoader.RawData(foldername)
            self.spots_3D     =  AnalysisLoader.SpotsIntsVol(foldername)
            self.features_3D  =  AnalysisLoader.Features(foldername)

            self.nucs_spts_ch    =  np.fromfile(foldername + '/nucs_spts_ch.bin', "uint16")
            self.nuclei_tracked  =  np.load(foldername + '/nuclei_tracked.npy')
            self.labbs           =  np.sign(self.nuclei_tracked)

            # wb    =  open_workbook(foldername + '/journal.xls')
            wb    =  load_workbook(foldername + '/journal.xlsx')
            s_wb  =  wb.worksheets[0]

            # if s_wb.col(1)[1].value == "Log Flt":
            if s_wb["B2"].value == "Log Flt":
                self.gaus_log_detect_combo.setCurrentIndex(0)
                self.param_detect_lbl.setText("Gauss Size")
            else:
                self.gaus_log_detect_combo.setCurrentIndex(1)
                self.param_detect_lbl.setText("Thr Coeff")

            self.param_detect_edt.setText(str(s_wb["B3"].value))

            self.gfilt_water_edt.setText(str(int(s_wb["B4"].value)))
            self.circ_thr_edt.setText(str(s_wb["B5"].value))
            self.dist_thr_edt.setText(str(int(s_wb["B6"].value)))
            self.spots_thr_edt.setText(str(s_wb["B7"].value))
            self.volume_thr_edt.setText(str(int(s_wb["B8"].value)))
            self.time_step_edt.setText(str(float(s_wb["B12"].value)))

            self.gaus_log_detect_value  =  s_wb["B2"].value
            self.gfilt_water_value      =  int(s_wb["B4"].value)
            self.circ_thr_value         =  s_wb["B5"].value
            self.dist_thr_value         =  int(s_wb["B6"].value)
            self.spots_thr_value        =  s_wb["B7"].value
            self.volume_thr_value       =  int(s_wb["B8"].value)
            self.max_dist               =  int(s_wb["B11"].value)
            self.time_step_value        =  float(s_wb["B12"].value)
            self.px_brd                 =  int(s_wb["B13"].value)

            self.nuclei_seg  =  np.zeros(self.filedata.imarray_red.shape, dtype=np.uint32)
            for t in range(self.nuclei_tracked.shape[0]):
                self.nuclei_seg[t, :, :]  =  skmr.label(self.nuclei_tracked[t, :, :], connectivity=1)

            self.sld1.setMaximum(self.filedata.imarray_red.shape[0] - 1)

            self.data_flag             =  1
            self.nuclei_t_visual_flag  =  1
            self.spots_segm_flag       =  1

            self.frame1.setImage(self.filedata.imarray_red[0, :, :])
            self.frame3.setImage(self.filedata.imarray_green[0, :, :])
            self.frame2.setImage(self.nuclei_tracked[0, :, :], levels=(0, self.nuclei_tracked.max()))
            self.mycmap  =  pg.ColorMap(np.linspace(0, 1, self.nuclei_tracked.max()), color=self.colors4map)
            self.frame2.setColorMap(self.mycmap)
            self.fnames  =  foldername
            self.nuc_spots_conn4load_analysis()

        except Exception:
            traceback.print_exc()

        self.ready_indicator()

    def start_cut(self):
        """Select the current frame as first frame to analyze"""
        self.filedata.imarray_red    =  self.filedata.imarray_red[self.sld1.value():, :, :]
        self.filedata.imarray_green  =  self.filedata.imarray_green[self.sld1.value():, :, :]
        self.filedata.green4D        =  self.filedata.green4D[self.sld1.value():, :, :, :]
        self.start_cut_value         =  self.sld1.value()
        self.sld1.setMaximum(self.filedata.imarray_red[:, 0, 0].size - 1)
        self.sld1.setValue(0)

    def end_cut(self):
        """Select the current frame as last frame to analyze"""
        self.filedata.imarray_red    =  self.filedata.imarray_red[:self.sld1.value() + 1, :, :]
        self.filedata.imarray_green  =  self.filedata.imarray_green[:self.sld1.value() + 1, :, :]
        self.filedata.green4D        =  self.filedata.green4D[:self.sld1.value() + 1, :, :, :]
        self.end_cut_value           =  self.sld1.value()
        self.sld1.setMaximum(self.filedata.imarray_red[:, 0, 0].size - 1)
        self.sld1.setValue(self.filedata.imarray_red[:, 0, 0].size - 1)

    def reload_files(self):
        """reload( and concatenate the files already selected"""
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:

            if str(self.fnames[0])[-3:] == 'lsm' or str(self.fnames[0])[-3:] == 'tif':
                self.filedata  =  MultiLoadLsmOrTif5D.MultiLoadLsmOrTif5D(self.fnames, self.nucs_spts_ch)
            if str(self.fnames[0])[-3:] == 'czi':
                self.filedata  =  MultiLoadCzi5D.MultiProcLoadCzi5D(self.fnames, self.nucs_spts_ch)

            self.frame1.setImage(self.filedata.imarray_red[0, :, :])
            self.frame3.setImage(self.filedata.imarray_green[0, :, :])
            self.data_flag = 1
            self.sld1.setMaximum(self.filedata.imarray_red[:, 0, 0].size - 1)
            self.sld1.setValue(0)

        except Exception:
            traceback.print_exc()

        self.ready_indicator()

    def sld1_update(self):
        self.time_lbl.setText("time  " + time.strftime("%M:%S", time.gmtime(self.sld1.value() * self.time_step_value)))
        self.frame_numb_lbl.setText("frame  "  +  str(self.sld1.value()))

        if self.data_flag  ==  1:
            self.frame1.setImage(self.filedata.imarray_red[self.sld1.value(), :, :])
            self.frame3.setImage(self.filedata.imarray_green[self.sld1.value(), :, :])
        if self.labbs_flag == 1:
            self.frame2.setImage(np.sign(self.labbs[self.sld1.value(), :, :]))
        if self.nuclei_flag == 1:
            self.frame2.setImage(self.nuclei_seg[self.sld1.value(), :, :])
        if self.nuclei_t_visual_flag == 1:
            self.frame2.setImage(self.nuclei_tracked[self.sld1.value(), :, :], levels=(0, self.nuclei_tracked.max()))
        if self.spots_segm_flag == 1:
            self.frame4.setImage(np.sign(self.spots_3D.spots_ints[self.sld1.value(), :, :]))
        if self.spots_trk_flag == 1:
            self.frame4.setImage(np.sign(self.spots_tracked_3D[self.sld1.value(), :, :]))

    def find_zero_in_mtss(self):
        """Launch the tool to set the time zero on the mitosis"""
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:
            self.mpp17  =  CheckSelectedRawData(self.filedata.imarray_red[0], self.nucs_spts_ch)
            self.mpp17.show()
        except Exception:
            traceback.print_exc()

        self.ready_indicator()

        self.mpp17.procStart.connect(self.loop_for_find_mtss)

    def loop_for_find_mtss(self):
        """Manage the possible situation of the tool to check the zero"""
        if self.mpp17.flag_yes_no == "yes":
            self.time_zero  =  self.mpp17.pz[0] - self.mpp17.mtss_frame
            print(self.time_zero)
            self.save_analysis()
            self.mpp17.close()

        if self.mpp17.flag_yes_no == "no":
            print("no")
            self.mpp17.close()
            self.find_zero_in_mtss()

    def save_analysis(self):
        """Save the analysis"""
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:
            fwritename  =  QtWidgets.QFileDialog.getSaveFileName()[0]
            reload(AnalysisSaver)
            reload(WriteSptsIntsDividedByBkg)

            AnalysisSaver.AnalysisSaver(fwritename, self.fnames, self.nuc_active.nuclei_active3c, self.nuclei_tracked, self.spots_tracked_3D, self.features_3D, self.nuc_active.n_active_vector,
                                        self.filedata.imarray_red, self.filedata.imarray_green, self.spots_3D, self.gfilt_water_value, self.circ_thr_value, self.dist_thr_value, self.spots_thr_value,
                                        self.volume_thr_value, self.start_cut_value, self.end_cut_value, self.nuc_active.popt, self.nuc_active.perr, self.max_dist, self.gaus_log_detect_value,
                                        self.param_detect_value, self.time_step_value, self.time_zero, self.px_brd, self.filedata.pix_size, self.filedata.pix_size_Z, self.nucs_spts_ch, self.t_track_end_value, self.software_version)
            BurstStatisticWriter.BurstStatisticWriter(fwritename, self.features_3D)
            WriteSptsIntsDividedByBkg.WriteSptsIntsDividedByBkg(fwritename, self.filedata.green4D, self.spots_3D, self.spots_tracked_3D)
            os.remove('rescue_fnames.txt')
            rescue_del  =  glob.glob("*.npy")
            for k in rescue_del:
                os.remove(k)

        except Exception:
            traceback.print_exc()

        self.ready_indicator()

    def gaus_log_detect(self, text):
        """Change label for nucs detection parameters"""
        self.gaus_log_detect_value  =  text
        if text == "Gauss Flt":
            self.param_detect_lbl.setText("Gauss Size")
        else:
            self.param_detect_lbl.setText("Thr Coeff")

    def param_detect_var(self, text):
        """Set nuclei detection parameter"""
        self.param_detect_value  =  float(text)

    def gfilt_water_var(self, text):
        """Set gaussian filter kernel size for water shed pre smoothing"""
        self.gfilt_water_value  =  int(text)

    def circ_thr_var(self, text):
        """Set circularity parameter"""
        self.circ_thr_value  =  float(text)

    def dist_thr_var(self, text):
        self.dist_thr_value  =  float(text)

    def spots_thr_var(self, text):
        self.spots_thr_value  =  float(text)

    def volume_thr_var(self, text):
        self.volume_thr_value  =  int(text)

    def time_step_var(self, text):
        self.time_step_value  =  float(text)

    def popup_nuclei_raw(self):
        """Popup raw nuclei"""
        PopUpTool.PopUpTool(self.filedata.imarray_red, 'Nuclei Raw Data')

    def popup_nuclei_detected(self):
        """Popup detected nuclei"""
        PopUpTool.PopUpTool(np.sign(self.labbs), 'Detected Nuclei')

    def popup_nuclei_segmented(self):
        """Popup segmented nuclei"""
        PopUpTool.PopUpToolWithMap(self.nuclei_seg, 'Segmented Nuclei', self.mycmap)

    def popup_nuclei_trackeded(self):
        """"Popup tracked nuclei"""
        PopUpTool.PopUpToolWithMap(self.nuclei_tracked, 'Tracked Nuclei', self.mycmap)

    def popup_spots_raw(self):
        """Popup green raw data"""
        PopUpTool.PopUpTool(self.filedata.imarray_green, 'Spots Raw Data')

    def popup_spots_segm(self):
        """Popup tracked spots"""
        PopUpTool.PopUpTool(np.sign(self.spots_tracked_3D), 'Segmented Spots')

    def popup_nucactive(self):
        """Popup false colored movie"""
        pg.image(self.nuc_active.nuclei_active3c, title="Active Nuclei")
        pg.plot(self.nuc_active.n_active_vector, pen='r', symbol='x')

    def nuclei_detection(self):
        """Detect raw data nuclei"""
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:

            if self.gaus_log_detect_value == "Log Flt":
                self.labbs  =  NucleiDetectLog.NucleiDetectLog(self.filedata.imarray_red, self.param_detect_value).labbs
            else:
                self.labbs  =  NucleiDetect.NucleiDetect(self.filedata.imarray_red, self.param_detect_value).labbs

            self.labbs_flag            =  1
            self.nuclei_flag           =  0
            self.nuclei_t_visual_flag  =  0
            self.frame2.setImage(np.sign(self.labbs[self.sld1.value(), :, :]))
            np.save('rescue_labbs', self.labbs)
            np.save('rescue_rawnucs', self.filedata.imarray_red)
            if self.gaus_log_detect_value == "Log Flt":
                np.save('rescue_labbs_info', np.array([self.param_detect_value, 1]))
            else:
                np.save('rescue_labbs_info', np.array([self.param_detect_value, 0]))

            file  =  open('rescue_fnames.txt', "w")
            file.write(self.fnames[0])
            for k in range(1, len(self.fnames)):
                file.write('\n' + self.fnames[k])
            file.close()

        except Exception:
            traceback.print_exc()

        self.ready_indicator()

    def nuclei_segmentation(self):
        """Segments detected nuclei"""
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:
            self.nuclei_seg            =  NucleiSegmentStackMultiCore.NucleiSegmentStackMultiCore(self.labbs, self.circ_thr_value, self.gfilt_water_value).nuclei_labels
            self.labbs_flag            =  0
            self.nuclei_flag           =  1
            self.nuclei_t_visual_flag  =  0
            self.frame2.setImage(self.nuclei_seg[self.sld1.value(), :, :])
            self.mycmap  =  pg.ColorMap(np.linspace(0, 1, self.nuclei_seg.max()), color=self.colors4map)
            self.frame2.setColorMap(self.mycmap)
            np.save('rescue_nuclei_seg', self.nuclei_seg)
            np.save('rescue_nuclei_seg_info', np.array([self.circ_thr_value, self.gfilt_water_value]))

        except Exception:
            traceback.print_exc()

        self.ready_indicator()

    def nuclei_tracking(self):
        """Track segmented nuclei"""
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:
            # self.px_brd                =  BorderPixelRemove.getNumb()
            self.px_brd                =  3
            nuclei_tracked             =  NucleiConnectMultiCore.NucleiConnectMultiCore(self.nuclei_seg, self.dist_thr_value).nuclei_tracked
            self.nuclei_tracked        =  RemoveBadNuclei.RemoveBorderNuclei(nuclei_tracked, self.px_brd).nuclei_tracked
            self.labbs_flag            =  0
            self.nuclei_flag           =  0
            self.nuclei_t_visual_flag  =  1
            self.frame2.setImage(self.nuclei_tracked[self.sld1.value(), :, :], levels=(0, self.nuclei_tracked.max()))
            self.mycmap  =  pg.ColorMap(np.linspace(0, 1, self.nuclei_tracked.max()), color=self.colors4map)
            self.frame2.setColorMap(self.mycmap)
            np.save('rescue_nuclei_tracked', self.nuclei_tracked)
            np.save('rescue_nuclei_tracked_info', np.array([self.dist_thr_value, self.px_brd]))

        except Exception:
            traceback.print_exc()

        self.ready_indicator()

    def spots_detect(self):
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:
            self.spots_3D         =  SpotsDetectionChopper.SpotsDetectionChopper(self.filedata.green4D, self.spots_thr_value, self.volume_thr_value)
            np.save('rescue_spts_ints', self.spots_3D.spots_ints)
            np.save('rescue_spts_vol', self.spots_3D.spots_vol)
            np.save('rescue_spts_tzxy', self.spots_3D.spots_tzxy)
            np.save('rescue_spts_info', np.array([self.spots_thr_value, self.volume_thr_value]))
            self.spots_segm_flag  =  1
            self.frame4.setImage(np.sign(self.spots_3D.spots_ints[self.sld1.value(), :, :]))

        except Exception:
            traceback.print_exc()

        self.ready_indicator()

    def nuc_spots_conn(self):
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:
            self.max_dist          =  SpotNcuDistanceThr.getNumb()
            self.spots_tracked_3D  =  SpotsConnection.SpotsConnection(self.nuclei_tracked, np.sign(self.spots_3D.spots_vol), self.max_dist).spots_tracked
            self.frame4.setImage(np.sign(self.spots_tracked_3D[self.sld1.value(), :, :]))

            self.tab3_flag         =  1
            self.spots_segm_flag   =  0
            self.spots_trk_flag    =  1

            ipp_3D                 =  self.spots_3D.spots_ints.reshape(self.spots_3D.spots_ints.size)
            i                      =  np.where(ipp_3D == 0)[0]
            ipp_3D                 =  np.delete(ipp_3D, i, axis=0)
            self.ipp_3D_av         =  ipp_3D.sum() / float(self.spots_3D.spots_vol.sum())                                                               # ipp is defined just to calculate the average intensity value of the spots, ipp_av
            self.features_3D       =  ParametersExtraction.ParametersExtraction(self.spots_3D.spots_ints, self.spots_tracked_3D, self.spots_3D.spots_vol)   # spots_3D.spots_vol * np.sign(self.spots_tracked_3D))

            self.nuc_active  =  NucleiSpotsConnection.NucleiSpotsConnection(self.spots_tracked_3D, self.nuclei_tracked)
            pg.image(self.nuc_active.nuclei_active3c)
            pg.plot(self.nuc_active.n_active_vector, symbol='x', pen='r')

        except Exception:
            traceback.print_exc()

        self.ready_indicator()

    def nuc_spots_conn4load_analysis(self):                                       # nuclei spot connection for loading analysis: does not require 'max_dist' because takes it from the journal file
        self.spots_tracked_3D  =  SpotsConnection.SpotsConnection(self.nuclei_tracked, np.sign(self.spots_3D.spots_vol), self.max_dist).spots_tracked

        self.tab3_flag        =  1
        self.spots_segm_flag  =  0
        self.spots_trk_flag   =  1
        self.frame4.setImage(np.sign(self.spots_tracked_3D[self.sld1.value(), :, :]))

        ipp_3D            =  self.spots_3D.spots_ints.reshape(self.spots_3D.spots_ints.size)
        i                 =  np.where(ipp_3D == 0)[0]
        ipp_3D            =  np.delete(ipp_3D, i, axis=0)
        self.ipp_3D_av    =  ipp_3D.sum() / float(self.spots_3D.spots_vol.sum())                                                               # ipp is defined just to calculate the average intensity value of the spots, ipp_av
        self.features_3D  =  ParametersExtraction.ParametersExtraction(self.spots_3D.spots_ints, self.spots_tracked_3D, self.spots_3D.spots_vol)   # spots_3D.spots_vol * np.sign(self.spots_tracked_3D))

        self.nuc_active  =  NucleiSpotsConnection.NucleiSpotsConnection(self.spots_tracked_3D, self.nuclei_tracked)
        pg.image(self.nuc_active.nuclei_active3c)
        pg.plot(self.nuc_active.n_active_vector, symbol='x', pen='r')

    def modify_tool(self):
        self.mpp1  =  ModifierCycleTool(self.filedata.imarray_red, np.copy(self.nuclei_seg), self.sld1.value())
        self.mpp1.show()
        self.mpp1.procStart.connect(self.sgnl_update_cycle)

    def sgnl_update_cycle(self, message):
        self.nuclei_seg  =  self.mpp1.nuclei_seg
        np.save('rescue_nuclei_seg', self.nuclei_seg)
        self.mpp1.close()

        self.time_lbl.setText("time  " + time.strftime("%M:%S", time.gmtime(self.sld1.value() * self.time_step_value)))

        self.sld1.setSliderPosition(message)

        self.frame1.setImage(self.filedata.imarray_red[message, :, :])
        self.frame3.setImage(self.filedata.imarray_green[message, :, :])
        self.frame2.setImage(self.nuclei_seg[message, :, :])
        if self.spots_segm_flag == 1:
            self.frame4.setImage(np.sign(self.spots_3D.spots_ints[message, :, :]))
        self.nuclei_flag           =  1
        self.nuclei_t_visual_flag  =  1

    def crop_tool_sgnl(self, message):
        """Crop raw data stack following the prescription of the Cropping Tool"""
        pts  =  self.mpp8.roi.parentBounds()
        x0   =  np.round(np.max([0, pts.x()])).astype(int)
        y0   =  np.round(np.max([0, pts.y()])).astype(int)
        x1   =  np.round(np.min([pts.x() + pts.width(), self.filedata.imarray_red.shape[1]])).astype(int)
        y1   =  np.round(np.min([pts.y() + pts.height(), self.filedata.imarray_red.shape[2]])).astype(int)

        self.filedata.imarray_red    =  self.filedata.imarray_red[:, x0:x1, y0:y1]
        self.filedata.imarray_green  =  self.filedata.imarray_green[:, x0:x1, y0:y1]
        self.filedata.green4D        =  self.filedata.green4D[:, :, x0:x1, y0:y1]
        self.frame1.setImage(self.filedata.imarray_red[message, :, :])
        self.frame3.setImage(self.filedata.imarray_green[message, :, :])
        self.mpp8.close()

    def spots_visual(self):
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:
            ipp_3D     =  self.spots_3D.spots_ints.reshape(self.spots_3D.spots_ints.size)
            i          =  np.where(ipp_3D == 0)[0]
            ipp_3D     =  np.delete(ipp_3D, i, axis=0)
            ipp_3D_av  =  ipp_3D.sum() / float(self.spots_3D.spots_vol.sum())                                             # ipp is defined just to calculate the average intensity value of the spots, ipp_av

            self.spsp  =  SpotsAnalyser(self.filedata.imarray_green, self.filedata.imarray_red, self.spots_tracked_3D, self.features_3D, ipp_3D_av, self.spots_3D.spots_ints, self.spots_3D.spots_vol)
            self.spsp.show()

        except Exception:
            traceback.print_exc()

        self.ready_indicator()

    def del_frame(self):
        self.filedata.imarray_red    =  np.delete(self.filedata.imarray_red, self.sld1.value(), 0)
        self.filedata.imarray_green  =  np.delete(self.filedata.imarray_green, self.sld1.value(), 0)
        self.filedata.green4D        =  np.delete(self.filedata.green4D, self.sld1.value(), 0)
        self.sld1.setMaximum(self.imarray_red[:, 0, 0].size - 1)
        self.sld1_update()

    def fake_coloured_time(self):
        self.mpp3  =  FalseColouredTime(self.nuclei_tracked, self.spots_tracked_3D, self.nuc_active.nuclei_active, self.time_step_value, self.fnames)
        self.mpp3.show()

    def act_dynamic_study(self):
        self.mpp5  =  ActivationDynamicStudy()
        self.mpp5.show()

    def tile_coordinates(self):
        tile_fname  =  str(QtWidgets.QFileDialog.getOpenFileName(None, "Select czi (or lsm) tile data files (maximum intensity projected)", filter="*.tif *.lsm"))
        xy_coord    =  FromTile2GlobCoordinate.FromTile2GlobCoordinate(tile_fname)
        txt_path    =  QtWidgets.QFileDialog.getOpenFileName(None, "Define a .txt file in which write")

        filetxt  =  open(txt_path, "w")
        filetxt.write(tile_fname  + "\n")
        filetxt.write("Start coordinate X:  " + str(xy_coord.x_coord_cntr_img[0])  + "\n")
        filetxt.write("End coordinate X:    " + str(xy_coord.x_coord_cntr_img[-1]) + "\n")
        filetxt.write("Start coordinate Y:  " + str(xy_coord.y_coord_cntr_img[0])  + "\n")
        filetxt.write("End coordinate Y:    " + str(xy_coord.y_coord_cntr_img[-1]) + "\n")
        filetxt.close()

    def auto_run(self):
        """Auto run the inspectionless part of the analysis"""
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:
            self.nuclei_detection()
            self.nuclei_segmentation()
            self.spots_detect()

        except Exception:
            traceback.print_exc()

        self.ready_indicator()

    def test_spots_detection(self):
        """Launch test spots detection tool"""
        self.mpp7  =  TestSpotsDetectionSetting(self.filedata.imarray_red, self.filedata.imarray_green, self.filedata.green4D, self.time_step_value)
        self.mpp7.show()
        self.mpp7.procStart.connect(self.insert_close_spots_test)

    def insert_close_spots_test(self):
        """Insert analysis parameters from the test spots analysis tool"""
        self.spots_thr_edt.setText(str(self.mpp7.spots_thr_value))
        self.volume_thr_edt.setText(str(self.mpp7.volume_thr_value))
        self.mpp7.close()

    def test_nuclei_detection(self):
        """Launch test nuclei detection tool"""
        self.mpp13  =  TestNucleiDetectionSetting(self.filedata.imarray_red, self.time_step_value)
        self.mpp13.show()
        self.mpp13.procStart.connect(self.insert_close_nucs_test)

    def insert_close_nucs_test(self):
        """Insert analysis parameters from the test nuclei analysis tool"""
        if self.mpp13.gaus_log_detect_value == "Gauss Flt":
            self.gaus_log_detect_combo.setCurrentIndex(0)
            self.param_detect_lbl.setText("Gauss Size")
        else:
            self.gaus_log_detect_combo.setCurrentIndex(1)
            self.param_detect_lbl.setText("Thr Coeff")

        self.gaus_log_detect_value  =  self.mpp13.gaus_log_detect_value
        self.param_detect_edt.setText(str(self.mpp13.param_detect_value))
        self.gfilt_water_edt.setText(str(int(self.mpp13.gfilt_water_value)))
        self.circ_thr_edt.setText(str(self.mpp13.circ_thr_value))
        self.mpp13.close()

    def single_nucleus(self):
        foldername  =  str(QtWidgets.QFileDialog.getExistingDirectory(None, "Select the folder with the analyzed data"))
        self.mpp9   =  SingleNucleusInspections(foldername)
        self.mpp9.show()

    def multi_plot_show(self):
        """Launch multi plot tool"""
        self.mpp11  =  MultiPlotShowing()
        self.mpp11.show()

    def spatial_comprehensive(self):
        """Launch the spatial comprehensive analysis tool"""
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:
            foldername  =  str(QtWidgets.QFileDialog.getExistingDirectory(None, "Select the folder with the analyzed data"))
            self.mpp10  =  SpatialComprehensive(foldername, self.nucs_spts_ch)
            self.mpp10.show()

        except Exception:
            traceback.print_exc()

        self.ready_indicator()

    def set_color_channel(self):
        """Launch popup tool to set color channels"""
        self.mpp6  =  SetColorChannel(self.nucs_spts_ch)
        self.mpp6.show()
        self.mpp6.procStart.connect(self.color_channels_vars)

    def color_channels_vars(self):
        """Set color channels values"""
        self.nucs_spts_ch  =  self.mpp6.channels_values - 1
        self.mpp6.close()

    def multi_color_intensity(self):
        """Launch the multicolor intensity tool"""
        self.mpp14  =  MultiColorIntensity()
        self.mpp14.show()

    def spts_intensity_minus_bkg(self):
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:
            foldername  =  str(QtWidgets.QFileDialog.getExistingDirectory(None, "Select the folder with the analyzed data"))
            self.fname_edt.setText(foldername)
            app.processEvents()
            app.processEvents()
            WriteSptsIntsMinusBkg.WriteSptsIntsMinusBkg(foldername)

        except Exception:
            traceback.print_exc()

        self.ready_indicator()

    def timeav_spts_intensity(self):
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:
            foldername      =  str(QtWidgets.QFileDialog.getExistingDirectory(None, "Select the folder with the analyzed data"))
            selected_spots  =  TimeavSptsIntensity.TimeavSptsIntensity(foldername)

            self.mpp15  =  SpotsTimeAverage(foldername, selected_spots.spots_vals, selected_spots.t_step, selected_spots.time_zero)
            self.mpp15.show()

        except Exception:
            traceback.print_exc()

        self.ready_indicator()

    def check_saturation(self):
        """Check SaturationInfo in the spots channel"""
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:

            self.mpp16  =  SaturationInfo(self.filedata.green4D)
            self.mpp16.show()

        except Exception:
            traceback.print_exc()

        self.ready_indicator()

    def calibration_spatial(self):
        foldername    =  str(QtWidgets.QFileDialog.getExistingDirectory(None, "Select the folder with the analyzed data"))
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:
            self.fname_edt.setText(foldername)
            app.processEvents()
            app.processEvents()

            calib_fact  =  CalibrationFactor.getCalibFact()
            CalibrationSpatial.CalibrationSpatial(foldername, calib_fact)

        except Exception:
            traceback.print_exc()
            pass

        self.ready_indicator()

    def rescue_analysis(self):
        """In case of unaspected software collapse, this function restores the analyis as it was before the collapse"""
        self.busy_indicator()
        app.processEvents()
        app.processEvents()

        try:
            # self.fnames  =  QtWidgets.QFileDialog.getOpenFileNames(None, "Select czi (or lsm) data files to concatenate...", filter="*.lsm *.czi *.tif")[0]
            file         =  open('rescue_fnames.txt', "r")
            a            =  file.readlines()
            self.fnames  =  []
            for k in range(len(a)):
                if a[k][-1:] == '\n':
                    self.fnames.append(a[k][:-1])
                else:
                    self.fnames.append(a[k])

            joined_fnames  = ' '
            for s in range(len(self.fnames)):
                joined_fnames  +=  str(self.fnames[s]) +  ' ----- '

            self.fname_edt.setText(joined_fnames)

            if str(self.fnames[0])[-3:] == 'lsm' or str(self.fnames[0])[-3:] == 'tif':
                self.filedata         =  MultiLoadLsmOrTif5D.MultiLoadLsmOrTif5D(self.fnames, self.nucs_spts_ch)
                self.time_step_edt.setText(str(self.filedata.time_step_value))

            if str(self.fnames[0])[-3:] == 'czi':
                self.filedata  =  MultiLoadCzi5D.MultiProcLoadCzi5D(self.fnames, self.nucs_spts_ch)
                self.time_step_edt.setText(str(self.filedata.time_step_value))

            if os.path.isfile('rescue_rawnucs.npy'):
                raw_red_rescued  =  np.load('rescue_rawnucs.npy')
                self.labbs       =  np.load('rescue_labbs.npy')

                rescue_raw            =  RescueFunctions.RescueRaw(raw_red_rescued, self.filedata)
                self.end_cut_value    =  rescue_raw.end_cut_value
                self.start_cut_value  =  rescue_raw.start_cut_value

                self.frame1.setImage(self.filedata.imarray_red[0, :, :])
                self.frame3.setImage(self.filedata.imarray_green[0, :, :])
                self.data_flag             =  1
                self.labbs_flag            =  1
                self.nuclei_flag           =  0
                self.nuclei_t_visual_flag  =  0
                self.frame2.setImage(np.sign(self.labbs[self.sld1.value(), :, :]))
                self.sld1.setMaximum(self.filedata.imarray_red.shape[0] - 1)
                self.sld1.setValue(0)

                bff_rescue  =  np.load('rescue_labbs_info.npy')
                self.param_detect_edt.setText(str(bff_rescue[0]))
                if bff_rescue[1] == 0:
                    self.gaus_log_detect_combo.setCurrentIndex(0)
                    self.param_detect_lbl.setText("Gauss Size")
                else:
                    self.gaus_log_detect_combo.setCurrentIndex(1)
                    self.param_detect_lbl.setText("Thr Coeff")

            if os.path.isfile('rescue_nuclei_seg.npy'):
                self.nuclei_seg            =  np.load('rescue_nuclei_seg.npy')
                self.labbs_flag            =  0
                self.nuclei_flag           =  1
                self.nuclei_t_visual_flag  =  0
                self.frame2.setImage(self.nuclei_seg[self.sld1.value(), :, :])
                self.mycmap  =  pg.ColorMap(np.linspace(0, 1, self.nuclei_seg.max()), color=self.colors4map)
                self.frame2.setColorMap(self.mycmap)
                bff_seg_info  =  np.load('rescue_nuclei_seg_info.npy')
                self.gfilt_water_edt.setText(str(int(bff_seg_info[1])))
                self.circ_thr_edt.setText(str(bff_seg_info[0]))

            if os.path.isfile('rescue_nuclei_tracked.npy'):
                self.nuclei_tracked        =  np.load('rescue_nuclei_tracked.npy')
                self.labbs_flag            =  0
                self.nuclei_flag           =  0
                self.nuclei_t_visual_flag  =  1
                self.frame2.setImage(self.nuclei_tracked[self.sld1.value(), :, :], levels=(0, self.nuclei_tracked.max()))
                self.mycmap  =  pg.ColorMap(np.linspace(0, 1, self.nuclei_tracked.max()), color=self.colors4map)
                self.frame2.setColorMap(self.mycmap)
                bff_nucs_trck  =  np.load('rescue_nuclei_tracked_info.npy')
                self.px_brd    =  int(bff_nucs_trck[1])
                self.dist_thr_edt.setText(str(int(bff_nucs_trck[0])))

            if os.path.isfile('rescue_spts_ints.npy'):
                self.spots_3D  =  AnalysisLoader.SpotsIntsVolRescue()
                bff_spts_info  =  np.load('rescue_spts_info.npy')
                self.spots_thr_edt.setText(str(bff_spts_info[0]))
                self.volume_thr_edt.setText(str(int(bff_spts_info[1])))
                self.spots_segm_flag  =  1
                self.frame4.setImage(np.sign(self.spots_3D.spots_ints[self.sld1.value(), :, :]))

        except Exception:
            traceback.print_exc()
            pass

        self.ready_indicator()

    def rmv_mitoticalTS(self):
        """Call PopUpTool up tool to remove mitotical spots"""
        self.spots_3D          =  None
        self.spots_tracked_3D  =  None
        self.features_3D       =  None
        fnames                 =  QtWidgets.QFileDialog.getOpenFileNames(None, "Select czi (or lsm) data files to check mitotical TS", filter="*.lsm *.czi *.tif *.lif")[0]
        self.mpp20             =  RemoveMitoticalSpots(self.filedata.imarray_red, self.filedata.imarray_green, self.filedata.green4D, fnames, self.nucs_spts_ch)
        self.mpp20.show()
        self.mpp20.procStart.connect(self.update_mip_spots_sgnl)

    def update_mip_spots_sgnl(self):
        """Update the work of RemoveMitoticalSpots in the main GUI"""
        self.spots_segm_flag         =  1
        self.spots_3D                =  self.mpp20.spots_3D
        self.t_track_end_value       =  self.mpp20.t_track_end_value
        self.frame4.setImage(np.sign(self.spots_3D.spots_ints[self.sld1.value(), :, :]))
        self.spots_thr_edt.setText(str(self.mpp20.spts_thr_value))
        self.volume_thr_edt.setText(str(int(self.mpp20.vol_thr_value)))
        self.frame3.updateImage()

        self.mpp20.close()

    def traces_image(self):
        self.mpp21  =  TracesImageTool()

#     def sisters_split(self):
#         analysis_folder  =  str(QtWidgets.QFileDialog.getExistingDirectory(None, "Select the folder with the analyzed data"))
#         # analysis_folder  =  '/home/atrullo/Desktop/Louise2Spots/Kr_E3_09242021/TestAnalysis3_short'
#         # self.mpp22       =  SistersTool(analysis_folder, self.software_version, raw_data)
#         self.mpp22       =  SistersTool(analysis_folder, self.software_version)
#         # self.mpp22.show()

    def remove_nucsdust(self):
        """Activate the popup tool to remove small detected objects"""
        self.mpp19  =  RemoveNucleiDust(self.filedata.imarray_red, self.nuclei_seg)
        self.mpp19.show()
        self.mpp19.procStart.connect(self.update_nuclei_seg)

    def update_nuclei_seg(self):
        """Update result of the thresholding"""
        self.nuclei_seg  =  np.copy(self.mpp19.nucs_thr)
        self.frame3.updateImage()
        self.mpp19.close()

    def analysis_conversion(self):
        analysis_folder  =  str(QtWidgets.QFileDialog.getExistingDirectory(None, "Select the folder with the analyzed data"))
        AnalysisConverter.AnalysisConverter(analysis_folder)


class ModifierCycleTool(QtWidgets.QWidget):
    """Activate the tool to manually correct the segmentation"""
    procStart  =  QtCore.pyqtSignal(int)

    def __init__(self, imarray_red, nuclei_seg, cif_start):
        QtWidgets.QWidget.__init__(self)

        frameShortcut  =  QtWidgets.QShortcut(QtGui.QKeySequence(QtCore.Qt.ShiftModifier + QtCore.Qt.Key_End), self)
        frameShortcut.activated.connect(self.shuffle_clrs)

        self.imarray_red  =  imarray_red
        self.nuclei_seg   =  nuclei_seg
        self.cif_start    =  cif_start

        tabs  =  QtWidgets.QTabWidget()
        tab1  =  QtWidgets.QWidget()
        tab2  =  QtWidgets.QWidget()

        mycmap  =  np.fromfile("mycmap.bin", "uint16").reshape((10000, 3))    # / 255.0
        self.colors4map  =  []
        for k in range(mycmap.shape[0]):
            self.colors4map.append(mycmap[k, :])
        self.colors4map[0]  =  np.array([0, 0, 0])

        framepp1  =  pg.ImageView(self, name='Frame1')
        framepp1.getImageItem().mouseClickEvent  =  self.click
        framepp1.ui.roiBtn.hide()
        framepp1.ui.menuBtn.hide()
        framepp1.setImage(self.nuclei_seg)
        mycmap  =  pg.ColorMap(np.linspace(0, 1, self.nuclei_seg.max()), color=self.colors4map)
        framepp1.setColorMap(mycmap)
        framepp1.timeLine.sigPositionChanged.connect(self.update_frame2)

        framepp2  =  pg.ImageView(self)
        framepp2.ui.roiBtn.hide()
        framepp2.ui.menuBtn.hide()
        framepp2.setImage(self.imarray_red)
        framepp2.timeLine.sigPositionChanged.connect(self.update_frame1)
        framepp2.view.setXLink('Frame1')
        framepp2.view.setYLink('Frame1')

        shuffle_clrs_btn  =  QtWidgets.QPushButton("Shuffle Colors", self)
        shuffle_clrs_btn.setFixedSize(120, 25)
        shuffle_clrs_btn.clicked.connect(self.shuffle_clrs)
        shuffle_clrs_btn.setToolTip('Shuffle colors')

        modify_btn  =  QtWidgets.QPushButton("Modify", self)
        modify_btn.setFixedSize(120, 25)
        modify_btn.clicked.connect(self.modify_lbls)
        modify_btn.setToolTip('Modify selection accordin to the segment (Ctrl+Suppr)')

        update_mainwindows_btn  =  QtWidgets.QPushButton("Update Nuclei", self)
        update_mainwindows_btn.setFixedSize(120, 25)
        update_mainwindows_btn.clicked.connect(self.update_mainwindows)

        frame_numb_lbl = QtWidgets.QLabel("frame  " + '0', self)
        frame_numb_lbl.setFixedSize(110, 13)

        end_pts  =  np.zeros((2, 2, 100))
        ar_reg   =  np.zeros((1000, 100))

        frame1_box  =  QtWidgets.QHBoxLayout()
        frame1_box.addWidget(framepp1)

        frame2_box  =  QtWidgets.QHBoxLayout()
        frame2_box.addWidget(framepp2)

        btn_box  =  QtWidgets.QHBoxLayout()
        btn_box.addWidget(shuffle_clrs_btn)
        btn_box.addStretch()
        btn_box.addWidget(frame_numb_lbl)
        btn_box.addWidget(modify_btn)
        btn_box.addWidget(update_mainwindows_btn)

        tab1.setLayout(frame1_box)
        tab2.setLayout(frame2_box)

        tabs.addTab(tab1, "Segmented")
        tabs.addTab(tab2, "RAW")

        layout  =  QtWidgets.QVBoxLayout()
        layout.addWidget(tabs)                                             # tabs is a Widget not a Layout!!!!!
        layout.addLayout(btn_box)

        self.end_pts         =  end_pts
        self.ar_reg          =  ar_reg
        self.framepp1        =  framepp1
        self.framepp2        =  framepp2
        self.frame_numb_lbl  =  frame_numb_lbl
        self.c_count         =  0

        self.setLayout(layout)
        self.setGeometry(300, 300, 600, 400)
        self.setWindowTitle("Modifier Tool")

        self.framepp1.setCurrentIndex(self.cif_start)

    def keyPressEvent(self, event):
        if event.key() == (QtCore.Qt.ControlModifier and Qt.Key_Z):

            cif                                =  self.framepp1.currentIndex
            self.nuclei_seg[cif, :, :]         =  self.bufframe
            self.framepp1.updateImage()
            self.framepp1.setCurrentIndex(cif)

        if event.key() == (QtCore.Qt.ControlModifier and Qt.Key_Delete):
            self.modify_lbls()

    def click(self, event):
        event.accept()
        pos        =  event.pos()
        modifiers  =  QtWidgets.QApplication.keyboardModifiers()

        if modifiers  ==  QtCore.Qt.ShiftModifier:
            if self.c_count - 2 * (self.c_count // 2) == 0:
                self.pos1  =  pos
            else:
                try:
                    self.framepp1.removeItem(self.roi)
                except AttributeError:
                    pass

                self.roi      =  pg.LineSegmentROI([self.pos1, pos], pen='r')
                self.framepp1.addItem(self.roi)

            self.c_count  +=  1

    def modify_lbls(self):

        cif      =  self.framepp1.currentIndex
        pp       =  self.roi.getHandles()
        pp       =  [self.roi.mapToItem(self.framepp1.imageItem, p.pos()) for p in pp]
        end_pts  =  np.array([[int(pp[0].x()), int(pp[0].y())], [int(pp[1].x()), int(pp[1].y())]])
        bufframe                           =  np.copy(self.nuclei_seg[cif, :, :])
        self.nuclei_seg[cif, :, :]         =  LabelsModify.LabelsModify(self.nuclei_seg[cif, :, :], end_pts).labels_fin
        self.framepp1.updateImage()
        self.bufframe  =  bufframe

    def update_frame2(self):
        self.framepp2.setCurrentIndex(self.framepp1.currentIndex)
        self.frame_numb_lbl.setText("frame  "  +  str(self.framepp1.currentIndex))

    def update_frame1(self):
        self.framepp1.setCurrentIndex(self.framepp2.currentIndex)
        self.frame_numb_lbl.setText("frame  "  +  str(self.framepp2.currentIndex))

    def shuffle_clrs(self):
        colors_bff  =  self.colors4map[1:]
        np.random.shuffle(colors_bff)
        self.colors4map[1:]  =  colors_bff
        mycmap  =  pg.ColorMap(np.linspace(0, 1, self.nuclei_seg.max()), color=self.colors4map)
        self.framepp1.setColorMap(mycmap)
        self.framepp1.updateImage()

    @QtCore.pyqtSlot()
    def update_mainwindows(self):
        val  =  self.framepp1.currentIndex

        self.procStart.emit(val)


class CroppingTool(QtWidgets.QWidget):
    """Popup tool to crop raw data"""
    procStart  =  QtCore.pyqtSignal(int)

    def __init__(self, imarray_red, imarray_green):
        QtWidgets.QWidget.__init__(self)

        imarray_tot  =  np.zeros(np.append(imarray_red.shape, 3))
        imarray_tot[:, :, :, 0]  =  imarray_red
        imarray_tot[:, :, :, 1]  =  imarray_green

        framepp1  =  pg.ImageView(self)
        framepp1.ui.roiBtn.hide()
        framepp1.ui.menuBtn.hide()
        framepp1.setImage(imarray_tot)

        roi  =  pg.RectROI([20, 20], [20, 20], pen='r')
        framepp1.addItem(roi)

        send_crop_btn  =  QtWidgets.QPushButton("Crop", self)
        send_crop_btn.setFixedSize(120, 25)
        send_crop_btn.clicked.connect(self.crop_to_mainwindows)

        keys  =  QtWidgets.QHBoxLayout()
        keys.addStretch()
        keys.addWidget(send_crop_btn)

        layout  =  QtWidgets.QVBoxLayout()
        layout.addWidget(framepp1)
        layout.addLayout(keys)

        self.framepp1  =  framepp1
        self.roi       =  roi

        self.setLayout(layout)
        self.setGeometry(300, 300, 600, 400)
        self.setWindowTitle("Modifier Tool")

    @QtCore.pyqtSlot()
    def crop_to_mainwindows(self):
        """Send modification to the main gui"""
        val  =  self.framepp1.currentIndex
        self.procStart.emit(val)


class SpotsAnalyser(QtWidgets.QWidget):
    """Pop up tool to check spots detection"""
    def __init__(self, imarray_green, imarray_red, spots_tracked, features_3D, ipp_3D_av, spots_3D_ints, spots_vol):
        QtWidgets.QWidget.__init__(self)

        self.spots_tracked_3D  =  spots_tracked
        self.features_3D       =  features_3D
        self.ipp_3D_av         =  ipp_3D_av
        self.spots_3D_ints     =  spots_3D_ints
        self.spots_vol         =  spots_vol

        imarray_green3c              =  np.zeros(np.append(imarray_green.shape, 3))
        imarray_green3c[:, :, :, 1]  =  imarray_green

        imarray_red3c              =  np.zeros(np.append(imarray_red.shape, 3))
        imarray_red3c[:, :, :, 0]  =  imarray_red

        frame1  =  pg.ImageView(self, name='Frame1')
        frame1.getImageItem().mouseClickEvent  =  self.click
        frame1.setImage(imarray_green3c)
        frame1.ui.roiBtn.hide()
        frame1.ui.menuBtn.hide()
        frame1.timeLine.sigPositionChanged.connect(self.update_frame2)

        frame2  =  pg.ImageView(self, name='Frame2')
        frame2.getImageItem().mouseClickEvent  =  self.click
        frame2.setImage(np.sign(self.spots_tracked_3D))
        frame2.timeLine.sigPositionChanged.connect(self.update_frame1)
        frame2.ui.roiBtn.hide()
        frame2.ui.menuBtn.hide()
        frame2.view.setXLink('Frame1')
        frame2.view.setYLink('Frame1')

        frame3  =  pg.ImageView(self, name='Frame3')
        frame3.getImageItem().mouseClickEvent  =  self.click
        frame3.setImage(imarray_red3c)
        frame3.timeLine.sigPositionChanged.connect(self.update_frame3)
        frame3.ui.menuBtn.hide()
        frame3.ui.roiBtn.hide()
        frame3.view.setXLink('Frame2')
        frame3.view.setYLink('Frame2')

        tabs  =  QtWidgets.QTabWidget()
        tab1  =  QtWidgets.QWidget()
        tab2  =  QtWidgets.QWidget()
        tab3  =  QtWidgets.QWidget()

        frame1_box  =  QtWidgets.QHBoxLayout()
        frame1_box.addWidget(frame1)

        frame2_box  =  QtWidgets.QHBoxLayout()
        frame2_box.addWidget(frame2)

        frame3_box  =  QtWidgets.QHBoxLayout()
        frame3_box.addWidget(frame3)

        tab1.setLayout(frame1_box)
        tab2.setLayout(frame2_box)
        tab3.setLayout(frame3_box)

        tabs.addTab(tab1, "Raw Spots")
        tabs.addTab(tab2, "Detected Spots")
        tabs.addTab(tab3, "Raw Nuclei")

        tabsld  =  QtWidgets.QVBoxLayout()
        tabsld.addWidget(tabs)

        show_spt_bytag_lbl  =  QtWidgets.QLabel("Spot by Tag", self)
        # show_spt_bytag_btn.clicked.connect(self.show_spt_trace)
        show_spt_bytag_lbl.setToolTip('Tag of the spot you want to see the trace')
        show_spt_bytag_lbl.setFixedSize(110, 25)

        spt_tag_call_edt  =  QtWidgets.QLineEdit(self)
        spt_tag_call_edt.setToolTip("Tag of the spot you want to trace")
        spt_tag_call_edt.setFixedSize(30, 22)
        spt_tag_call_edt.textChanged[str].connect(self.spt_tag_call_var)
        spt_tag_call_edt.returnPressed.connect(self.show_spt_trace)

        frame_numb_lbl = QtWidgets.QLabel("frame  " + '0', self)
        frame_numb_lbl.setFixedSize(110, 13)

        spot_tag_box  =  QtWidgets.QHBoxLayout()
        spot_tag_box.addWidget(show_spt_bytag_lbl)
        spot_tag_box.addWidget(spt_tag_call_edt)

        key_box  =  QtWidgets.QHBoxLayout()
        key_box.addWidget(frame_numb_lbl)
        key_box.addStretch()
        key_box.addLayout(spot_tag_box)

        layout   =  QtWidgets.QVBoxLayout()
        layout.addLayout(tabsld)
        layout.addLayout(key_box)

        lookUT_colors  =  [(0, 0, 0), (255, 0, 0), (255, 255, 255)]
        self.lookUT  =  pg.ColorMap(np.array([.3, .6, .9]), color=lookUT_colors)

        self.frame1            =  frame1
        self.frame2            =  frame2
        self.frame3            =  frame3
        self.frame_numb_lbl    =  frame_numb_lbl
        self.spt_tag_call_edt  =  spt_tag_call_edt

        self.tab3_flag  =  0

        self.setLayout(layout)
        self.setGeometry(300, 300, 600, 400)
        self.setWindowTitle('Spots Analyser')

    def spt_tag_call_var(self, text):
        self.spt_tag_call_value  =  float(text)

    def update_frame2(self):
        self.frame2.setCurrentIndex(self.frame1.currentIndex)
        self.frame3.setCurrentIndex(self.frame1.currentIndex)
        self.frame_numb_lbl.setText("frame  "  +  str(self.frame1.currentIndex))

    def update_frame1(self):
        self.frame1.setCurrentIndex(self.frame2.currentIndex)
        self.frame3.setCurrentIndex(self.frame2.currentIndex)
        self.frame_numb_lbl.setText("frame  "  +  str(self.frame2.currentIndex))

    def update_frame3(self):
        self.frame1.setCurrentIndex(self.frame3.currentIndex)
        self.frame2.setCurrentIndex(self.frame3.currentIndex)
        self.frame_numb_lbl.setText("frame  "  +  str(self.frame3.currentIndex))

    def click(self, event):
        event.accept()
        pos        =  event.pos()
        modifiers  =  QtWidgets.QApplication.keyboardModifiers()
        flag_3D    =  0

        if modifiers  ==  QtCore.Qt.ShiftModifier:

            ref_point_3D  =  int(self.spots_tracked_3D[self.frame1.currentIndex, np.round(pos[0]).astype(int), np.round(pos[1]).astype(int)])

            if ref_point_3D != 0:
                self.file2show_3D  =  (self.spots_tracked_3D == ref_point_3D).astype(int)
                spots_vol_ref      =  self.spots_vol * (self.spots_tracked_3D == ref_point_3D)
                flag_3D            =  1
                kidx_3D            =  np.where(self.features_3D.statistics_info[:, 0] == ref_point_3D)[0][0]
                the_title          =  'Spot ID ' + str(ref_point_3D)
                self.spt_tag_call_edt.setText(str(ref_point_3D))
            else:
                self.file2show_3D  =  np.sign(self.spots_tracked_3D)
                spots_vol_ref      =  self.spots_vol * np.sign(self.spots_tracked_3D)
                the_title          =  'Spot ID ' + str(0)

            cif  =  self.frame1.currentIndex
            self.frame2.setImage(self.file2show_3D + np.sign(self.spots_tracked_3D))
            self.frame2.setCurrentIndex(cif)
            self.frame2.setColorMap(self.lookUT)
            w  =  pg.plot(title=the_title)
            w.addLegend()
            w.plot(np.nan_to_num((self.file2show_3D * self.spots_3D_ints).sum(2).sum(1) / spots_vol_ref.sum(2).sum(1)), symbol='x', pen='r', name='average on the spot')
            w.plot(self.ipp_3D_av * np.ones(self.file2show_3D.shape[0]))

            ww  =  pg.plot(title=the_title)
            ww.addLegend()
            ww.plot((self.file2show_3D * self.spots_3D_ints).sum(2).sum(1), symbol='o', pen='b', name='spot int sum')
            # ww.plot((self.file2show_3D * self.spots_3D_ints).sum(2).sum(1) - 3540 * spots_vol_ref.sum(2).sum(1), symbol='x', pen='r', name='no background')
            pg.plot(spots_vol_ref.sum(2).sum(1), symbol='x', pen='r', title='volume')

            if flag_3D == 1:
                step_info  =  int((self.features_3D.statistics_info.shape[1] - 3) / 5.0)
                w_feats  =  pg.plot(10 * np.ones(10), pen=[0, 0, 0], title='Parameters')
                w_feats.plot(np.ones(10), pen=[0, 0, 0])
                text1  =  pg.TextItem("numb = "               + str(self.features_3D.statistics_info[kidx_3D, 1]), anchor=(0, 0), fill=(0, 0, 0))
                text2  =  pg.TextItem("duration = "           + str(self.features_3D.statistics_info[kidx_3D, 2:2 + step_info]), anchor=(0, 0), fill=(0, 0, 0))
                text3  =  pg.TextItem("amplitude average = "  + str(self.features_3D.statistics_info[kidx_3D, 2 + 3 * step_info:2 + 4 * step_info]), anchor=(0, 0), fill=(0, 0, 0))
                text4  =  pg.TextItem("amplitude maximum = "  + str(self.features_3D.statistics_info[kidx_3D, 2 + step_info:2 + 2 * step_info]), anchor=(0, 0), fill=(0, 0, 0))
                text5  =  pg.TextItem("amplitude integral = " + str(self.features_3D.statistics_info[kidx_3D, 2 + 2 * step_info:2 + 3 * step_info]), anchor=(0, 0), fill=(0, 0, 0))

                text1.setPos(1, 10)
                text2.setPos(1, 8)
                text3.setPos(1, 6)
                text4.setPos(1, 4)
                text5.setPos(1, 2)

                w_feats.addItem(text1)
                w_feats.addItem(text2)
                w_feats.addItem(text3)
                w_feats.addItem(text4)
                w_feats.addItem(text5)

                w_feats.showAxis('bottom', False)
                w_feats.showAxis('left', False)

    def show_spt_trace(self):
        flag_3D    =  0

        if self.spt_tag_call_value != 0 and (self.spots_tracked_3D == self.spt_tag_call_value).sum() != 0:
            self.file2show_3D  =  (self.spots_tracked_3D == self.spt_tag_call_value).astype(int)
            spots_vol_ref      =  self.spots_vol * (self.spots_tracked_3D == self.spt_tag_call_value)
            flag_3D            =  1
            kidx_3D            =  np.where(self.features_3D.statistics_info[:, 0] == self.spt_tag_call_value)[0][0]
            the_title          =  'Spot ID ' + str(self.spt_tag_call_value)
        else:
            self.file2show_3D  =  np.sign(self.spots_tracked_3D)
            spots_vol_ref      =  self.spots_vol * np.sign(self.spots_tracked_3D)
            the_title          =  'Spot ID ' + str(0)


        cif  =  self.frame1.currentIndex
        self.frame2.setImage(self.file2show_3D + np.sign(self.spots_tracked_3D))
        self.frame2.setCurrentIndex(cif)
        self.frame2.setColorMap(self.lookUT)
        w  =  pg.plot(title=the_title)
        w.addLegend()
        w.plot(np.nan_to_num((self.file2show_3D * self.spots_3D_ints).sum(2).sum(1) / spots_vol_ref.sum(2).sum(1)), symbol='x', pen='r', name='average on the spot')
        w.plot(self.ipp_3D_av * np.ones(self.file2show_3D.shape[0]))

        ww = pg.plot()
        ww.addLegend()
        ww.plot((self.file2show_3D * self.spots_3D_ints).sum(2).sum(1), symbol='o', pen='b', name='spot int sum')
        pg.plot(spots_vol_ref.sum(2).sum(1), symbol='x', pen='r', title='volume')


        if flag_3D == 1:
            step_info  =  int((self.features_3D.statistics_info.shape[1] - 3) / 5.0)
            w_feats    =  pg.plot(10 * np.ones(10), pen=[0, 0, 0], title='Parameters')
            w_feats.plot(np.ones(10), pen=[0, 0, 0])
            text1  =  pg.TextItem("numb = "               + str(self.features_3D.statistics_info[kidx_3D, 1]), anchor=(0, 0), fill=(0, 0, 0))
            text2  =  pg.TextItem("duration = "           + str(self.features_3D.statistics_info[kidx_3D, 2:2 + step_info]), anchor=(0, 0), fill=(0, 0, 0))
            text3  =  pg.TextItem("amplitude average = "  + str(self.features_3D.statistics_info[kidx_3D, 2 + 3 * step_info:2 + 4 * step_info]), anchor=(0, 0), fill=(0, 0, 0))
            text4  =  pg.TextItem("amplitude maximum = "  + str(self.features_3D.statistics_info[kidx_3D, 2 + step_info:2 + 2 * step_info]), anchor=(0, 0), fill=(0, 0, 0))
            text5  =  pg.TextItem("amplitude integral = " + str(self.features_3D.statistics_info[kidx_3D, 2 + 2 * step_info:2 + 3 * step_info]), anchor=(0, 0), fill=(0, 0, 0))

            text1.setPos(1, 10)
            text2.setPos(1, 8)
            text3.setPos(1, 6)
            text4.setPos(1, 4)
            text5.setPos(1, 2)

            w_feats.addItem(text1)
            w_feats.addItem(text2)
            w_feats.addItem(text3)
            w_feats.addItem(text4)
            w_feats.addItem(text5)

            w_feats.showAxis('bottom', False)
            w_feats.showAxis('left', False)


class FalseColouredTime(QtWidgets.QWidget):
    """Pop up tool to geenrate false multicoloured movies"""
    def __init__(self, nuclei_tracked, spots_tracked, nuclei_active, time_step_value, fnames):
        QtWidgets.QWidget.__init__(self)
        reload(FalseColored3Ch)

        self.nuclei_tracked   =  nuclei_tracked
        self.spots_tracked    =  spots_tracked
        self.nuclei_active    =  nuclei_active
        self.time_step_value  =  time_step_value
        self.fnames           =  fnames

        framepp1  =  pg.ImageView(self)
        framepp1.ui.roiBtn.hide()
        framepp1.ui.menuBtn.hide()

        run_btn  =  QtWidgets.QPushButton("Run", self)
        run_btn.clicked.connect(self.run)
        run_btn.setToolTip('Generate the false colored movie with time selection')
        run_btn.setFixedSize(110, 25)

        save_btn  =  QtWidgets.QPushButton("Save", self)
        save_btn.clicked.connect(self.write_video)
        save_btn.setToolTip('Save the false colored movie ')
        save_btn.setFixedSize(110, 25)

        sld1  =  QtWidgets.QSlider(QtCore.Qt.Horizontal, self)
        sld1.setMaximum(self.nuclei_tracked.shape[0])
        sld1.valueChanged.connect(self.sld1_val)

        sld2  =  QtWidgets.QSlider(QtCore.Qt.Horizontal, self)
        sld2.setMaximum(self.nuclei_tracked.shape[0])
        sld2.valueChanged.connect(self.sld2_val)

        time_ref1  =  QtWidgets.QLabel("time1  " + '0', self)
        time_ref1.setFixedSize(110, 25)

        time_ref2  =  QtWidgets.QLabel("time2  " + '0', self)
        time_ref2.setFixedSize(110, 25)

        frame_sldrs  =  QtWidgets.QVBoxLayout()
        frame_sldrs.addWidget(framepp1)
        frame_sldrs.addWidget(sld1)
        frame_sldrs.addWidget(sld2)

        keys  =  QtWidgets.QVBoxLayout()
        keys.addWidget(run_btn)
        keys.addWidget(save_btn)
        keys.addStretch()
        keys.addWidget(time_ref1)
        keys.addWidget(time_ref2)

        layout  =  QtWidgets.QHBoxLayout()
        layout.addLayout(frame_sldrs)
        layout.addLayout(keys)

        self.sld1       =  sld1
        self.sld2       =  sld2
        self.time_ref1  =  time_ref1
        self.time_ref2  =  time_ref2
        self.framepp1   =  framepp1

        self.setLayout(layout)
        self.setGeometry(300, 300, 600, 400)
        self.setWindowTitle("False Multicoloured Movie")

    def sld1_val(self):
        self.time_ref1.setText("time1  " + time.strftime("%M:%S", time.gmtime(self.sld1.value() * self.time_step_value)))
        self.sld1_value  =  self.sld1.value()

    def sld2_val(self):
        self.time_ref2.setText("time2  " + time.strftime("%M:%S", time.gmtime(self.sld2.value() * self.time_step_value)))
        self.sld2_value  =  self.sld2.value()

    def run(self):
        # self.false3Ch  =  FalseColored3Ch.FalseColored3Ch(self.nuclei_tracked, self.spots_tracked, self.nuclei_active, self.sld1_value, self.sld2_value).false3ch
        self.false3Ch  =  FalseColored3Ch.FalseColored3Ch(self.nuclei_tracked, self.spots_tracked, self.sld1_value, self.sld2_value).false3ch
        self.framepp1.setImage(self.false3Ch)

    def write_video(self):
        foldername  =  str(QtWidgets.QFileDialog.getExistingDirectory(None, 'Select a folder'))
        tifffile.imwrite(foldername + '/false_colored_time.tif', self.false3Ch.astype("uint8"))

        FalseColored3Ch.FalseColoredTimeWrite(foldername, self.nuclei_tracked, self.false3Ch, self.sld1_value, self.sld2_value, self.time_step_value, self.fnames)


class ActivationDynamicStudy(QtWidgets.QWidget):
    def __init__(self):
        QtWidgets.QWidget.__init__(self)

        fnames_edt  =  QtWidgets.QLineEdit(self)
        fnames_edt.setToolTip('Names of the loaded xls files')

        add_btn  =  QtWidgets.QPushButton("Add Files", self)
        add_btn.clicked.connect(self.add_journal)
        add_btn.setToolTip('Add excels journal files to study')
        add_btn.setFixedSize(110, 25)

        run_btn  =  QtWidgets.QPushButton("Run", self)
        run_btn.clicked.connect(self.run_fitting)
        run_btn.setToolTip('Run the fitting of the averaged activation rates')
        run_btn.setFixedSize(110, 25)

        write_btn  =  QtWidgets.QPushButton("Write", self)
        write_btn.clicked.connect(self.write_results)
        write_btn.setToolTip('Write the results of the fitting')
        write_btn.setFixedSize(110, 25)

        framepp1  =  pg.PlotWidget(self)

        keys  =  QtWidgets.QHBoxLayout()
        keys.addStretch()
        keys.addWidget(add_btn)
        keys.addWidget(run_btn)
        keys.addWidget(write_btn)

        layout  =  QtWidgets.QVBoxLayout()
        layout.addWidget(fnames_edt)
        layout.addWidget(framepp1)
        layout.addLayout(keys)

        journal_list  =  []

        self.framepp1      =  framepp1
        self.fnames_edt    =  fnames_edt
        self.journal_list  =  journal_list
        self.name2write    =  ""

        self.setLayout(layout)
        self.setGeometry(300, 300, 600, 400)
        self.setWindowTitle("Spatial Analysis")

    def add_journal(self):
        journal_name  =  str(QtWidgets.QFileDialog.getOpenFileName(None, "Select a journal file from a previous analysis", filter="*.xls"))
        self.journal_list.append(journal_name)
        self.name2write  +=  journal_name + " --- "
        self.fnames_edt.setText(self.name2write)

    def run_fitting(self):
        self.zz  =  FromJournal2Fitting.FromJournal2Fitting(self.journal_list)
        self.framepp1.plot(self.zz.half_active.sum(1))
        self.framepp1.plot(self.zz.y_vv, pen='r')

    def write_results(self):
        name_file2write  =  str(QtWidgets.QFileDialog.getSaveFileName(None, "Save file", "", ".xls"))
        if name_file2write[-4:] != '.xls':
            name_file2write  +=  '.xls'
        FromJournal2Fitting.WriteResults(name_file2write, self.journal_list, self.zz.half_active, self.zz.half_active, self.zz.popt)


class SpotNcuDistanceThr(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(SpotNcuDistanceThr, self).__init__(parent)

        numb_pixels_lbl  =  QtWidgets.QLabel("Numb of pixels", self)
        numb_pixels_lbl.setFixedSize(110, 25)

        numb_pixels_edt = QtWidgets.QLineEdit(self)
        numb_pixels_edt.setToolTip("Max distance in pixels")
        numb_pixels_edt.setFixedSize(30, 22)
        numb_pixels_edt.textChanged[str].connect(self.numb_pixels_var)

        input_close_btn  =  QtWidgets.QPushButton("Ok", self)
        input_close_btn.clicked.connect(self.input_close)
        input_close_btn.setToolTip('Input values')
        input_close_btn.setFixedSize(50, 25)

        numb_pixels_lbl_edit_box  =  QtWidgets.QHBoxLayout()
        numb_pixels_lbl_edit_box.addWidget(numb_pixels_lbl)
        numb_pixels_lbl_edit_box.addWidget(numb_pixels_edt)

        input_close_box  =  QtWidgets.QHBoxLayout()
        input_close_box.addStretch()
        input_close_box.addWidget(input_close_btn)

        layout  =  QtWidgets.QVBoxLayout()
        layout.addLayout(numb_pixels_lbl_edit_box)
        layout.addLayout(input_close_box)

        self.setWindowModality(Qt.ApplicationModal)
        self.setLayout(layout)
        self.setGeometry(300, 300, 70, 50)
        self.setWindowTitle("Spot Nuc Max Distance")

    def numb_pixels_var(self, text):
        self.numb_pixels_value  =  int(text)

    def input_close(self):
        self.close()

    def numb_pixels(self):
        return self.numb_pixels_value

    @staticmethod
    def getNumb(parent=None):
        dialog       =  SpotNcuDistanceThr(parent)
        result       =  dialog.exec_()
        numb_pixels  =  dialog.numb_pixels()
        return numb_pixels


class TestSpotsDetectionSetting(QtWidgets.QWidget):
    procStart  =  QtCore.pyqtSignal()

    def __init__(self, imarray_red, imarray_green, green4D, time_step_value):
        QtWidgets.QWidget.__init__(self)

        raw_mip_data  =  np.zeros(np.append(imarray_green.shape, 3))
        raw_mip_data[:, :, :, 1]  =  imarray_green
        raw_mip_data[:, :, :, 0]  =  imarray_red / 2.0

        frame1  =  pg.ImageView(self, name='Frame1')
        frame1.setImage(raw_mip_data)
        frame1.ui.roiBtn.hide()
        frame1.ui.menuBtn.hide()
        frame1.timeLine.sigPositionChanged.connect(self.update_time_and_frame)

        frame2  =  pg.ImageView(self, name='Frame2')
        frame2.ui.roiBtn.hide()
        frame2.ui.menuBtn.hide()
        frame2.view.setXLink('Frame1')
        frame2.view.setYLink('Frame1')

        frame3  =  pg.ImageView(self, name='Frame3')
        frame3.ui.roiBtn.hide()
        frame3.ui.menuBtn.hide()
        frame3.timeLine.sigPositionChanged.connect(self.update_frame4)
        frame3.view.setXLink('Frame1')
        frame3.view.setYLink('Frame1')

        frame4  =  pg.ImageView(self, name='Frame4')
        frame4.ui.roiBtn.hide()
        frame4.ui.menuBtn.hide()
        frame4.timeLine.sigPositionChanged.connect(self.update_frame3)
        frame4.view.setXLink('Frame1')
        frame4.view.setYLink('Frame1')

        choose_frame_tggl  =  QtWidgets.QCheckBox("Set Frame", self)
        choose_frame_tggl.stateChanged.connect(self.choose_frame)
        choose_frame_tggl.setToolTip('Choose this frame to work on')
        choose_frame_tggl.setFixedSize(110, 25)

        spots_thr_lbl  =  QtWidgets.QLabel('Spots Thr', self)
        spots_thr_lbl.setFixedSize(60, 25)

        spots_thr_edt  =  QtWidgets.QLineEdit(self)
        spots_thr_edt.textChanged[str].connect(self.spots_thr_var)
        spots_thr_edt.setToolTip('Intensity threshold to segment spots: it is expressed in terms of standard deviation, suggested value 7')
        spots_thr_edt.setFixedSize(35, 25)

        volume_thr_lbl  =  QtWidgets.QLabel('Vol Thr', self)
        volume_thr_lbl.setFixedSize(60, 25)

        volume_thr_edt  =  QtWidgets.QLineEdit(self)
        volume_thr_edt.textChanged[str].connect(self.volume_thr_var)
        volume_thr_edt.setToolTip('Threshold volume on spot detection: suggested value 5')
        volume_thr_edt.setFixedSize(35, 25)

        spots_det_btn  =  QtWidgets.QPushButton("Detect Spots", self)
        spots_det_btn.clicked.connect(self.spots_det)
        spots_det_btn.setToolTip('Detect spots in the choosen frame')
        spots_det_btn.setFixedSize(110, 25)

        close_insert_btn  =  QtWidgets.QPushButton("Close && Insert", self)
        close_insert_btn.clicked.connect(self.close_insert)
        close_insert_btn.setToolTip('Close test tool and insert the values')
        close_insert_btn.setFixedSize(110, 25)

        frame1_box  =  QtWidgets.QHBoxLayout()
        frame1_box.addWidget(frame1)

        frame2_box  =  QtWidgets.QHBoxLayout()
        frame2_box.addWidget(frame2)

        frame3_box  =  QtWidgets.QHBoxLayout()
        frame3_box.addWidget(frame3)

        frame4_box  =  QtWidgets.QHBoxLayout()
        frame4_box.addWidget(frame4)

        tabs_a  =  QtWidgets.QTabWidget()
        tab1_a  =  QtWidgets.QWidget()
        tab2_a  =  QtWidgets.QWidget()

        tab1_a.setLayout(frame1_box)
        tab2_a.setLayout(frame2_box)

        tabs_a.addTab(tab1_a, "Raw Spots")
        tabs_a.addTab(tab2_a, "Detected")

        tabs_b  =  QtWidgets.QTabWidget()
        tab1_b  =  QtWidgets.QWidget()
        tab2_b  =  QtWidgets.QWidget()

        tab1_b.setLayout(frame3_box)
        tab2_b.setLayout(frame4_box)

        tabs_b.addTab(tab1_b, "Raw Spots")
        tabs_b.addTab(tab2_b, "Detected")

        spots_thr_hor  =  QtWidgets.QHBoxLayout()
        spots_thr_hor.addWidget(spots_thr_lbl)
        spots_thr_hor.addWidget(spots_thr_edt)

        volume_thr_hor  =  QtWidgets.QHBoxLayout()
        volume_thr_hor.addWidget(volume_thr_lbl)
        volume_thr_hor.addWidget(volume_thr_edt)

        time_lbl = QtWidgets.QLabel("time     " + '0', self)
        time_lbl.setFixedSize(110, 13)

        frame_numb_lbl = QtWidgets.QLabel("frame  " + '0', self)
        frame_numb_lbl.setFixedSize(110, 13)

        commands  =  QtWidgets.QVBoxLayout()
        commands.addWidget(choose_frame_tggl)
        commands.addLayout(spots_thr_hor)
        commands.addLayout(volume_thr_hor)
        commands.addWidget(spots_det_btn)
        commands.addStretch()
        commands.addWidget(close_insert_btn)
        commands.addStretch()
        commands.addWidget(time_lbl)
        commands.addWidget(frame_numb_lbl)

        layout  =  QtWidgets.QHBoxLayout()
        layout.addWidget(tabs_a)
        layout.addWidget(tabs_b)
        layout.addLayout(commands)

        self.frame1           =  frame1
        self.frame2           =  frame2
        self.frame3           =  frame3
        self.frame4           =  frame4
        self.time_lbl         =  time_lbl
        self.frame_numb_lbl   =  frame_numb_lbl
        self.time_step_value  =  time_step_value

        self.raw_mip_data  =  raw_mip_data
        self.green4D       =  green4D

        self.setLayout(layout)
        self.setGeometry(300, 300, 1000, 500)
        self.setWindowTitle("Test Spot Detection")

    def spots_thr_var(self, text):
        self.spots_thr_value  =  float(text)

    def volume_thr_var(self, text):
        self.volume_thr_value  =  int(text)

    def update_frame4(self):
        self.frame4.setCurrentIndex(self.frame3.currentIndex)

    def update_time_and_frame(self):
        self.time_lbl.setText("time  " + time.strftime("%M:%S", time.gmtime(self.frame1.currentIndex * self.time_step_value)))
        self.frame_numb_lbl.setText("frame  "  +  str(self.frame1.currentIndex))

    def update_frame3(self):
        self.frame3.setCurrentIndex(self.frame4.currentIndex)

    def choose_frame(self, state):
        if state == QtCore.Qt.Checked:
            self.cif  =  self.frame1.currentIndex
            self.frame1.setImage(self.raw_mip_data[self.frame1.currentIndex, :, :])

        else:
            self.frame1.setImage(self.raw_mip_data)
            self.frame1.setCurrentIndex(self.cif)

    def spots_det(self):
        spots  =  SpotsDetection3D.SpotsDetection3D_Single([self.green4D[self.cif, :, :, :], self.spots_thr_value, self.volume_thr_value])
        self.frame2.setImage(np.sign(spots.spots_ints))
        self.frame3.setImage(self.green4D[self.cif, :, :, :])
        self.frame4.setImage(spots.spots_lbls)

    @QtCore.pyqtSlot()
    def close_insert(self):
        self.procStart.emit()


class TestNucleiDetectionSetting(QtWidgets.QWidget):
    procStart  =  QtCore.pyqtSignal()

    def __init__(self, imarray_red, time_step_value):
        QtWidgets.QWidget.__init__(self)

        frame1  =  pg.ImageView(self, name='Frame1')
        frame1.setImage(imarray_red)
        frame1.ui.roiBtn.hide()
        frame1.ui.menuBtn.hide()
        frame1.timeLine.sigPositionChanged.connect(self.update_frame2)

        frame2  =  pg.ImageView(self, name='Frame2')
        frame2.ui.roiBtn.hide()
        frame2.ui.menuBtn.hide()
        frame2.view.setXLink('Frame1')
        frame2.view.setYLink('Frame1')
        frame2.timeLine.sigPositionChanged.connect(self.update_frame1)

        tabs  =  QtWidgets.QTabWidget()
        tab1  =  QtWidgets.QWidget()
        tab2  =  QtWidgets.QWidget()

        frame1_box  =  QtWidgets.QHBoxLayout()
        frame1_box.addWidget(frame1)

        frame2_box  =  QtWidgets.QHBoxLayout()
        frame2_box.addWidget(frame2)

        tab1.setLayout(frame1_box)
        tab2.setLayout(frame2_box)

        tabs.addTab(tab1, "Raw Nuclei")
        tabs.addTab(tab2, "Detected Nuclei")

        nuc_detect_btn  =  QtWidgets.QPushButton("N-Detect", self)
        nuc_detect_btn.clicked.connect(self.nuclei_detection)
        nuc_detect_btn.setToolTip('Nuclei detection')
        nuc_detect_btn.setFixedSize(110, 25)

        param_detect_lbl  =  QtWidgets.QLabel("Gauss Size", self)
        param_detect_lbl.setFixedSize(60, 25)

        param_detect_edt  =  QtWidgets.QLineEdit(self)
        param_detect_edt.textChanged[str].connect(self.param_detect_var)
        param_detect_edt.setToolTip('Sets the size of the Gaussian Filter for the pre-smoothing (1.5) or for the logharitmic detection (0.995 - 1.005)')
        param_detect_edt.setFixedSize(35, 25)

        param_detect_box  =  QtWidgets.QHBoxLayout()
        param_detect_box.addWidget(param_detect_lbl)
        param_detect_box.addWidget(param_detect_edt)

        gaus_log_detect_combo = QtWidgets.QComboBox(self)
        gaus_log_detect_combo.addItem("Gauss Flt")
        gaus_log_detect_combo.addItem("Log Flt")
        gaus_log_detect_combo.setCurrentIndex(0)
        gaus_log_detect_combo.setToolTip('Switch between a linear nuclei detection and a logaritmic nuclei detection')
        gaus_log_detect_combo.activated[str].connect(self.gaus_log_detect)

        nuc_segment_btn  =  QtWidgets.QPushButton("N-Segment", self)
        nuc_segment_btn.clicked.connect(self.nuclei_segmentation)
        nuc_segment_btn.setToolTip('Nuclei Segmentation')
        nuc_segment_btn.setFixedSize(110, 25)

        gfilt_water_lbl  =  QtWidgets.QLabel('W-Shed', self)
        gfilt_water_lbl.setFixedSize(60, 25)

        gfilt_water_edt  =  QtWidgets.QLineEdit(self)
        gfilt_water_edt.textChanged[str].connect(self.gfilt_water_var)
        gfilt_water_edt.setToolTip('Sets the size parameter for the Water Shed algorithm, suggested value 7')
        gfilt_water_edt.setFixedSize(35, 25)

        circ_thr_lbl  =  QtWidgets.QLabel('Circ Thr', self)
        circ_thr_lbl.setFixedSize(60, 25)

        circ_thr_edt  =  QtWidgets.QLineEdit(self)
        circ_thr_edt.textChanged[str].connect(self.circ_thr_var)
        circ_thr_edt.setToolTip('Circularity Threshold of the detected nuclei, suggested value is 0.65')
        circ_thr_edt.setFixedSize(35, 25)

        gfilt_water_box  =  QtWidgets.QHBoxLayout()
        gfilt_water_box.addWidget(gfilt_water_lbl)
        gfilt_water_box.addWidget(gfilt_water_edt)

        circ_thr_box  =  QtWidgets.QHBoxLayout()
        circ_thr_box.addWidget(circ_thr_lbl)
        circ_thr_box.addWidget(circ_thr_edt)

        all_frames_btn  =  QtWidgets.QPushButton("All", self)
        all_frames_btn.clicked.connect(self.all_frames)
        all_frames_btn.setToolTip('Look all the frames')
        all_frames_btn.setFixedSize(110, 25)

        first_frams_btn  =  QtWidgets.QPushButton("Start", self)
        first_frams_btn.clicked.connect(self.first_frame)
        first_frams_btn.setToolTip('First Frame to Study')
        first_frams_btn.setFixedSize(50, 25)

        last_frame_btn  =  QtWidgets.QPushButton("End", self)
        last_frame_btn.clicked.connect(self.last_frame)
        last_frame_btn.setToolTip('Last Frame to Study')
        last_frame_btn.setFixedSize(50, 25)

        close_insert_btn  =  QtWidgets.QPushButton("Close && Insert", self)
        close_insert_btn.clicked.connect(self.close_insert)
        close_insert_btn.setToolTip('Close test tool and insert the values')
        close_insert_btn.setFixedSize(110, 25)

        first_last_box  =  QtWidgets.QHBoxLayout()
        first_last_box.addWidget(first_frams_btn)
        first_last_box.addWidget(last_frame_btn)

        time_lbl = QtWidgets.QLabel("time     " + '0', self)
        time_lbl.setFixedSize(110, 13)

        frame_numb_lbl = QtWidgets.QLabel("frame  " + '0', self)
        frame_numb_lbl.setFixedSize(110, 13)

        keys  =  QtWidgets.QVBoxLayout()
        keys.addLayout(first_last_box)
        keys.addWidget(all_frames_btn)
        keys.addStretch()
        keys.addLayout(param_detect_box)
        keys.addWidget(gaus_log_detect_combo)
        keys.addWidget(nuc_detect_btn)
        keys.addStretch()
        keys.addLayout(gfilt_water_box)
        keys.addLayout(circ_thr_box)
        keys.addWidget(nuc_segment_btn)
        keys.addStretch()
        keys.addWidget(close_insert_btn)
        keys.addStretch()
        keys.addWidget(time_lbl)
        keys.addWidget(frame_numb_lbl)

        layout  =  QtWidgets.QHBoxLayout()
        layout.addWidget(tabs)
        layout.addLayout(keys)

        mycmap  =  np.fromfile("mycmap.bin", "uint16").reshape((10000, 3)) # / 255.0
        self.colors4map  =  []
        for k in range(mycmap.shape[0]):
            self.colors4map.append(mycmap[k, :])
        self.colors4map[0]  =  np.array([0, 0, 0])

        self.imarray_red            =  imarray_red
        self.imarray_red_tot        =  np.copy(imarray_red)
        self.param_detect_edt       =  param_detect_edt
        self.param_detect_lbl       =  param_detect_lbl
        # self.log_flag               =  0
        self.frame1                 =  frame1
        self.frame2                 =  frame2
        self.labbs_flag             =  0
        self.gaus_log_detect_value  =  "Gauss Flt"
        self.frame_numb_lbl         =  frame_numb_lbl
        self.time_lbl               =  time_lbl
        self.time_step_value        =  time_step_value

        self.setLayout(layout)
        self.setGeometry(300, 300, 1000, 500)
        self.setWindowTitle("Test Nuclei Detection")

    def param_detect_var(self, text):
        self.param_detect_value  =  float(text)

    def gfilt_water_var(self, text):
        self.gfilt_water_value  =  float(text)

    def circ_thr_var(self, text):
        self.circ_thr_value  =  float(text)

    def gaus_log_detect(self, text):
        self.gaus_log_detect_value  =  text
        if text == "Gauss Flt":
            self.param_detect_lbl.setText("Gauss Size")
        else:
            self.param_detect_lbl.setText("Thr Coeff")

    def update_frame2(self):
        if self.labbs_flag == 1:
            self.frame2.setCurrentIndex(self.frame1.currentIndex)
        self.time_lbl.setText("time  " + time.strftime("%M:%S", time.gmtime(self.frame1.currentIndex * self.time_step_value)))
        self.frame_numb_lbl.setText("frame  "  +  str(self.frame1.currentIndex))

    def update_frame1(self):
        self.frame1.setCurrentIndex(self.frame2.currentIndex)
        self.time_lbl.setText("time  " + time.strftime("%M:%S", time.gmtime(self.frame2.currentIndex * self.time_step_value)))
        self.frame_numb_lbl.setText("frame  "  +  str(self.frame2.currentIndex))

    def all_frames(self):
        self.imarray_red  =  np.copy(self.imarray_red_tot)
        cid               =  self.frame1.currentIndex
        self.frame1.setImage(self.imarray_red)
        self.frame1.setCurrentIndex(cid)
        self.labbs_flag  =  2

    def first_frame(self):
        cid               =  self.frame1.currentIndex
        self.imarray_red  =  self.imarray_red[cid:, :, :]
        self.frame1.setImage(self.imarray_red)

    def last_frame(self):
        cid               =  self.frame1.currentIndex
        self.imarray_red  =  self.imarray_red[:cid, :, :]
        self.frame1.setImage(self.imarray_red)

    def nuclei_detection(self):
        if self.gaus_log_detect_value == "Log Flt":
            self.labbs  =  NucleiDetectLog.NucleiDetectLog(self.imarray_red, self.param_detect_value).labbs
        else:
            self.labbs  =  NucleiDetect.NucleiDetect(self.imarray_red, self.param_detect_value).labbs

        self.labbs_flag  =  1
        self.frame2.setImage(np.sign(self.labbs))

    def nuclei_segmentation(self):
        self.nucs_segm  =  NucleiSegmentStackMultiCore.NucleiSegmentStackCoordinator([self.labbs, self.circ_thr_value, self.gfilt_water_value]).nuclei_labels
        self.frame2.setImage(self.nucs_segm)
        self.mycmap  =  pg.ColorMap(np.linspace(0, 1, self.nucs_segm.max()), color=self.colors4map)
        self.frame2.setColorMap(self.mycmap)

    @QtCore.pyqtSlot()
    def close_insert(self):
        self.procStart.emit()


class SingleNucleusInspections(QtWidgets.QWidget):
    """Pop up tool to follow in time the activation if a single nucleus"""
    def __init__(self, foldername):
        QtWidgets.QWidget.__init__(self)

        self.nucs     =  np.load(foldername + '/nuclei_tracked.npy')
        self.spts     =  np.load(foldername + '/spots_tracked.npy')
        self.raw_red  =  AnalysisLoader.RawData(foldername).imarray_red
        self.raw_sp   =  np.load(foldername + "/spots_3D_ints.npy")

        frame1  =  pg.ImageView(self)
        frame1.ui.roiBtn.hide()
        frame1.ui.menuBtn.hide()
        frame1.timeLine.sigPositionChanged.connect(self.update_frame)
        frame1.setFixedSize(400, 400)

        frame2  =  pg.PlotWidget(self)

        nuc_tag_lbl  =  QtWidgets.QLabel("Nuc Tag", self)
        nuc_tag_lbl.setFixedSize(110, 25)

        nuc_tag_edt  =  QtWidgets.QLineEdit(self)
        nuc_tag_edt.textChanged[str].connect(self.nuc_tag_var)
        nuc_tag_edt.setToolTip('Sets the size of the Gaussian Filter for the pre-smoothing; suggested value 1.5')
        nuc_tag_edt.setFixedSize(35, 25)

        visualize_nuc_btn  =  QtWidgets.QPushButton("Visualize", self)
        visualize_nuc_btn.clicked.connect(self.visualize_nuc)
        visualize_nuc_btn.setToolTip('Reload selected files')
        visualize_nuc_btn.setFixedSize(110, 25)

        frames  =  QtWidgets.QHBoxLayout()
        frames.addWidget(frame1)
        frames.addWidget(frame2)

        commands  =  QtWidgets.QHBoxLayout()
        commands.addStretch()
        commands.addWidget(nuc_tag_lbl)
        commands.addWidget(nuc_tag_edt)
        commands.addWidget(visualize_nuc_btn)

        layout  =  QtWidgets.QVBoxLayout()
        layout.addLayout(frames)
        layout.addLayout(commands)

        self.int_normalization  =  VisualNucSpot_v2.IntensityMaximum(self.spts, self.raw_sp).int_max

        self.frame1   =  frame1
        self.frame2   =  frame2
        self.t_steps  =  self.nucs.shape[0]
        self.frame1.setImage(self.nucs)

        self.mypen  =  pg.mkPen('w', width=3)

        self.setGeometry(300, 300, 1000, 400)
        self.setWindowTitle('Activation Inspector')
        self.setLayout(layout)
        self.show()

    def nuc_tag_var(self, text):
        self.nuc_tag_value  =  int(text)

    def update_frame(self):
        self.frame2.clear()
        self.frame2.setXRange(0, self.t_steps)
        self.frame2.setYRange(0, self.int_normalization)
        self.frame2.plot(np.arange(self.frame1.currentIndex + 1), self.vis_nuc_spt.spts_prof[:self.frame1.currentIndex + 1], pen=self.mypen, symbol='x', symbolBrush='r')

    def visualize_nuc(self):
        self.frame2.clear()
        self.vis_nuc_spt  =  VisualNucSpot_v2.VisualNucSpot(self.nucs, self.raw_red, self.spts, self.raw_sp, self.nuc_tag_value)
        self.frame2.setXRange(0, self.t_steps)
        self.frame2.setYRange(0, self.int_normalization)
        self.frame1.setImage(self.vis_nuc_spt.mtx3c)


class SpatialComprehensive(QtWidgets.QWidget):
    """Popup tool for spatial analysis of the analysis results"""
    def __init__(self, foldername, nucs_spts_ch):
        QtWidgets.QWidget.__init__(self)

        nucs  =  np.load(foldername + '/nuclei_tracked.npy')

        self.nucs_spts_ch  =  np.fromfile(foldername + "/nucs_spts_ch.bin", "uint16")

        mycmap  =  np.fromfile("mycmap.bin", "uint16").reshape((10000, 3))        # / 255.0
        self.colors5map  =  []
        for k in range(mycmap.shape[0]):
            self.colors5map.append(mycmap[k, :])
        self.colors5map[0]  =  np.array([0, 0, 0])
        mycmap  =  pg.ColorMap(np.linspace(0, 1, nucs.max()), color=self.colors5map)

        wb              =  load_workbook(foldername + '/journal.xlsx')
        pix_size_value  =  wb.active["B16"].value

        fname_edt  =  QtWidgets.QLineEdit(self)
        fname_edt.setToolTip('Names and paths of the loaded files')
        fname_edt.setText(foldername)

        frame_img  =  pg.ImageView(self)
        frame_img.ui.roiBtn.hide()
        frame_img.ui.menuBtn.hide()
        frame_img.setImage(nucs)
        frame_img.setColorMap(mycmap)

        frame_fedt  =  QtWidgets.QVBoxLayout()
        frame_fedt.addWidget(fname_edt)
        frame_fedt.addWidget(frame_img)

        gastru_finder_btn  =  QtWidgets.QPushButton("Gastr Finder", self)
        gastru_finder_btn.setToolTip('Find the position of the gastrulation line')
        gastru_finder_btn.setFixedSize(120, 25)
        gastru_finder_btn.clicked.connect(self.gastru_finder)

        ap_coord_finder_btn  =  QtWidgets.QPushButton("AP Coord", self)
        ap_coord_finder_btn.setToolTip('Find the AP coordinate of the image')
        ap_coord_finder_btn.setFixedSize(120, 25)
        ap_coord_finder_btn.clicked.connect(self.ap_coord_finder)

        show_coord_btn  =  QtWidgets.QPushButton("Show Coord", self)
        show_coord_btn.setToolTip('Show AP gastrulation coordinate')
        show_coord_btn.setFixedSize(120, 25)
        show_coord_btn.clicked.connect(self.show_coord)

        pix_size_lbl  =  QtWidgets.QLabel("Pix Size", self)
        pix_size_lbl.setFixedSize(50, 25)

        pix_size_edt  =  QtWidgets.QLineEdit(self)
        pix_size_edt.textChanged[str].connect(self.pix_size_var)
        pix_size_edt.setToolTip("Pixel size in " + u"µ" + "m")
        pix_size_edt.setFixedSize(60, 25)
        pix_size_edt.setText(str(pix_size_value))

        pix_size_box  =  QtWidgets.QHBoxLayout()
        pix_size_box.addWidget(pix_size_lbl)
        pix_size_box.addWidget(pix_size_edt)

        add_rect_btn  =  QtWidgets.QPushButton("Manual ROI", self)
        add_rect_btn.setToolTip('Add rectangular ROI for the spatial study')
        add_rect_btn.setFixedSize(120, 25)
        add_rect_btn.clicked.connect(self.add_rect)

        set_rect_btn  =  QtWidgets.QPushButton("Set ROI", self)
        set_rect_btn.setToolTip('Set coordinates of rectangular ROI for the spatial study')
        set_rect_btn.setFixedSize(120, 25)
        set_rect_btn.clicked.connect(self.set_rect)

        remove_roi_btn  =  QtWidgets.QPushButton("Remove ROI", self)
        remove_roi_btn.setToolTip('Save the bursting statistics for the nuclei in the selected region')
        remove_roi_btn.setFixedSize(120, 25)
        remove_roi_btn.clicked.connect(self.remove_roi)

        save_analysis_btn  =  QtWidgets.QPushButton("Save", self)
        save_analysis_btn.setToolTip('Save the bursting statistics for the nuclei in the selected region')
        save_analysis_btn.setFixedSize(120, 25)
        save_analysis_btn.clicked.connect(self.save_analysis)

        hor_line_one  =  QtWidgets.QFrame()
        hor_line_one.setFrameStyle(QtWidgets.QFrame.HLine)

        hor_line_two  =  QtWidgets.QFrame()
        hor_line_two.setFrameStyle(QtWidgets.QFrame.HLine)

        empty_lbl  =  QtWidgets.QLabel("  ", self)
        empty_lbl.setFixedSize(10, 16)

        min_lbl  =  QtWidgets.QLabel("min", self)
        min_lbl.setFixedSize(42, 16)

        max_lbl  =  QtWidgets.QLabel("max", self)
        max_lbl.setFixedSize(42, 16)

        x_lbl  =  QtWidgets.QLabel("x(" + u"µ" + "m", self)
        x_lbl.setFixedSize(20, 16)

        y_lbl  =  QtWidgets.QLabel("y('%'AP)", self)
        y_lbl.setFixedSize(20, 16)

        xmin_lbl  =  QtWidgets.QLabel("  0", self)
        xmin_lbl.setFixedSize(42, 16)

        xmax_lbl  =  QtWidgets.QLabel("  0", self)
        xmax_lbl.setFixedSize(42, 16)

        ymin_lbl  =  QtWidgets.QLabel("  0", self)
        ymin_lbl.setFixedSize(42, 16)

        ymax_lbl  =  QtWidgets.QLabel("  0", self)
        ymax_lbl.setFixedSize(42, 16)

        coord_lbl_box_1  =  QtWidgets.QHBoxLayout()
        coord_lbl_box_1.addWidget(empty_lbl)
        coord_lbl_box_1.addWidget(min_lbl)
        coord_lbl_box_1.addWidget(max_lbl)

        coord_lbl_box_2  =  QtWidgets.QHBoxLayout()
        coord_lbl_box_2.addWidget(x_lbl)
        coord_lbl_box_2.addWidget(xmin_lbl)
        coord_lbl_box_2.addWidget(xmax_lbl)

        coord_lbl_box_3  =  QtWidgets.QHBoxLayout()
        coord_lbl_box_3.addWidget(y_lbl)
        coord_lbl_box_3.addWidget(ymin_lbl)
        coord_lbl_box_3.addWidget(ymax_lbl)

        coord_lbl_box  =  QtWidgets.QVBoxLayout()
        coord_lbl_box.addLayout(coord_lbl_box_1)
        coord_lbl_box.addLayout(coord_lbl_box_2)
        coord_lbl_box.addLayout(coord_lbl_box_3)

        keys  =  QtWidgets.QVBoxLayout()
        keys.addWidget(gastru_finder_btn)
        keys.addWidget(ap_coord_finder_btn)
        keys.addLayout(pix_size_box)
        keys.addWidget(show_coord_btn)
        keys.addStretch()
        keys.addWidget(add_rect_btn)
        keys.addWidget(hor_line_one)
        keys.addLayout(coord_lbl_box)
        keys.addWidget(hor_line_two)
        keys.addWidget(set_rect_btn)
        keys.addWidget(remove_roi_btn)
        keys.addStretch()
        keys.addStretch()
        keys.addWidget(save_analysis_btn)

        layout  =  QtWidgets.QHBoxLayout()
        layout.addLayout(frame_fedt)
        layout.addLayout(keys)

        self.foldername     =  foldername
        self.nucs_trk       =  nucs
        self.x_size         =  nucs.shape[1]
        self.y_size         =  nucs.shape[2]
        self.x_tile_coord   =  np.array([0, self.y_size])

        self.frame_img     =  frame_img
        self.xmin_lbl      =  xmin_lbl
        self.xmax_lbl      =  xmax_lbl
        self.ymin_lbl      =  ymin_lbl
        self.ymax_lbl      =  ymax_lbl
        self.red_pen       =  pg.mkPen(color='r', width=2)
        self.blue_pen      =  pg.mkPen(color='b', width=2)
        self.nucs_spts_ch  =  nucs_spts_ch
        self.pix_size_edt  =  pix_size_edt

        self.setLayout(layout)
        self.setGeometry(300, 300, 800, 600)
        self.setWindowTitle("Spatial Analysis")

    def gastru_finder(self):
        """Pick and load and show the file with gastrulation"""
        gastru_fname  =  QtWidgets.QFileDialog.getOpenFileNames(None, "Select lsm file (latest of the movie) to detect gastrulation", filter="*.lsm *.czi *.tif *.lif")[0]
        # filedata      =  MultiLoadLsmOrTif5D.MultiLoadLsmOrTif5D(gastru_fname, self.nucs_spts_ch)

        if str(gastru_fname[0])[-3:] == 'lsm' or str(gastru_fname[0])[-3:] == 'tif':
            filedata  =  MultiLoadLsmOrTif5D.MultiLoadLsmOrTif5D(gastru_fname, self.nucs_spts_ch)
        if str(gastru_fname[0])[-3:] == 'czi':
            filedata  =  MultiLoadCzi5D.MultiProcLoadCzi5D(gastru_fname, self.nucs_spts_ch)
        if str(gastru_fname[0])[-3:] == 'lif':
            filedata  =  MultiLoadLif5D.MultiLoadLif5D(gastru_fname, self.nucs_spts_ch)

        w              =  pg.image(filedata.imarray_red)
        self.roi_line  =  pg.LinearRegionItem(orientation=True, brush=[0, 0, 0, 0])
        w.addItem(self.roi_line)

    def ap_coord_finder(self):
        """Popup tool to eventually correct the orientation of the tile scan"""
        tile_fname  =  QtWidgets.QFileDialog.getOpenFileName(None, "Select lsm or czi tile data files (maximum intensity projected)", filter="*.tif *.lsm *.czi *.lif")[0]
        self.mpp18  =  FlipTile(tile_fname)
        self.mpp18.show()
        self.mpp18.procStart.connect(self.fliped_tile)

    def fliped_tile(self):
        """Function to flip the tile scan"""
        self.x_tile_coord  =  FromTile2GlobCoordinate.FromTile2GlobCoordinate(self.mpp18.tile_img).x_coord_cntr_img
        self.mpp18.close()

    def pix_size_var(self, text):
        """Set the pixel value"""
        self.pix_size_value  =  float(text)

    def show_coord(self):
        """Shox the coordinate on the ImageView"""
        try:
            for i in range(len(self.l_x)):
                self.frame_img.removeItem(self.l_x[i])
            for i in range(len(self.l_y)):
                self.frame_img.removeItem(self.l_y[i])
        except AttributeError:
            pass

        try:
            self.frame_img.removeItem(self.zero_item)
        except AttributeError:
            pass

        y_coord_vec       =  np.round(np.linspace(0, self.x_size, 10) - int(self.roi_line.getRegion()[1]))
        self.l_y          =  list(range(y_coord_vec.size))
        y_coord_step      =  y_coord_vec[1] - y_coord_vec[0]
        y_coord_vec2show  =  np.round(y_coord_vec * self.pix_size_value, decimals=2)

        for i in range(len(self.l_y)):
            self.l_y[i]  =  pg.TextItem(str(y_coord_vec2show[i]), 'g')

        for i in range(y_coord_vec.size):
            self.frame_img.addItem(self.l_y[i])
            self.l_y[i].setPos(-30, i * y_coord_step - 5)

        j  =  np.argmin(np.abs(y_coord_vec))
        self.frame_img.removeItem(self.l_y[j])
        self.zero_item  =  pg.TextItem(str(0), 'g')
        self.frame_img.addItem(self.zero_item)
        self.zero_item.setPos(-30, int(self.roi_line.getRegion()[1]))

        x_coord_vec   =  np.round(np.linspace(self.x_tile_coord[0], self.x_tile_coord[-1], 10))
        x_coord_step  =  self.x_size / 10.0
        self.l_x      =  list(range(x_coord_vec.size))
        for i in range(len(self.l_x)):
            self.l_x[i]  =  pg.TextItem(str(x_coord_vec[i]), 'g')

        for i in range(x_coord_vec.size):
            self.frame_img.addItem(self.l_x[i])
            self.l_x[i].setPos(i * x_coord_step - 5, -30)

    def add_rect(self):
        """Add a roi rectangle to drag and resize and eventually remove the previous one"""
        try:
            self.frame_img.removeItem(self.roi_rect)
        except AttributeError:
            pass

        self.roi_rect  =  pg.RectROI([20, 20], [20, 20], pen=self.red_pen)
        self.frame_img.addItem(self.roi_rect)
        self.roi_rect.sigRegionChanged.connect(self.update_xy_lbls)

    def update_xy_lbls(self):
        """Update the coordinate label of the gui"""
        pts  =  self.roi_rect.parentBounds()
        x0   =  np.round((pts.x() / self.x_size) * (self.x_tile_coord[-1] - self.x_tile_coord[0]) + self.x_tile_coord[0], decimals=2)
        y0   =  np.round(pts.y() - self.mpp18.roi_line.getRegion()[1], decimals=2)
        x1   =  np.round(((pts.x() + pts.width()) / self.x_size) * (self.x_tile_coord[-1] - self.x_tile_coord[0]) + self.x_tile_coord[0], decimals=2)
        y1   =  np.round(pts.y() + pts.height() - self.mpp18.roi_line.getRegion()[1], decimals=2)

        self.xmin_lbl.setText(str(x0))
        self.xmax_lbl.setText(str(x1))
        self.ymin_lbl.setText(str(np.round(y0 * self.pix_size_value, decimals=2)))
        self.ymax_lbl.setText(str(np.round(y1 * self.pix_size_value, decimals=2)))

        self.coords2write  =   np.array([x0, x1, y0 * self.pix_size_value, y1 * self.pix_size_value])

    def set_rect(self):
        """Add a roi rectangle with input corners coordinates"""
        coords             =   RectRoiCoordinates.getCoordinates()
        self.coords2write  =   np.array([coords[0], coords[1], coords[2], coords[3]])

        self.xmin_lbl.setText(str(coords[0]))
        self.xmax_lbl.setText(str(coords[1]))
        self.ymin_lbl.setText(str(coords[2]))
        self.ymax_lbl.setText(str(coords[3]))

        coords[2:]  =  coords[2:] / self.pix_size_value + int(self.roi_line.getRegion()[1])
        coords[0]   =  self.x_size * (coords[0] - self.x_tile_coord[0]) / (self.x_tile_coord[-1] - self.x_tile_coord[0])
        coords[1]   =  self.x_size * (coords[1] - self.x_tile_coord[0]) / (self.x_tile_coord[-1] - self.x_tile_coord[0])

        try:
            self.frame_img.removeItem(self.roi_rect)
        except AttributeError:
            pass

        self.roi_rect  =  pg.RectROI([coords[0], coords[2]], [coords[1] - coords[0], coords[3] - coords[2]], pen=self.blue_pen)
        self.frame_img.addItem(self.roi_rect)
        self.roi_rect.sigRegionChanged.connect(self.update_xy_lbls)

    def remove_roi(self):
        """Remove the roi rectangle from the ImageView"""
        self.frame_img.removeItem(self.roi_rect)

    def set_filter(self):
        """Dialog to set filter parameters"""
        self.filter_set  =  FilterSettings.getSetting()

    def save_analysis(self):
        """Save the analysis with the spatial constrain"""
        book             =  load_workbook(self.foldername + '/journal.xlsx')
        time_step_value  =  book.worksheets[0]["B12"].value
        time_zero        =  book.worksheets[0]["B15"].value

        reload(ComprehensiveBurstAnalysisWriter)
        reload(ComprehensiveActivationWriter)
        self.set_filter()
        pts                          =  self.roi_rect.parentBounds()
        x_pos, y_pos, width, height  =  int(np.round(pts.x())), int(np.round(pts.y())), int(np.round(pts.width())), int(np.round(pts.height()))
        burst_wrt                    =  ComprehensiveBurstAnalysisWriter.ComprehensiveBurstAnalysisWriter(self.foldername, self.nucs_trk, x_pos, y_pos, width, height, self.filter_set, self.roi_line.getRegion()[1], self.x_size, self.x_tile_coord, self.pix_size_value, self.coords2write)

        ComprehensiveActivationWriter.ComprehensiveActivationWriter(self.foldername, self.nucs_trk, burst_wrt.spots_trk, burst_wrt.idxs_in, self.filter_set, self.pix_size_value, self.coords2write, burst_wrt.numb_nucs_area, time_step_value, time_zero)

        reload(SpatialDividedByBkg)
        SpatialDividedByBkg.SpatialDividedByBkg(self.foldername, time_zero, time_step_value)
        GalleryDividedByBackground.GalleryDividedByBackground(self.foldername)


class RectRoiCoordinates(QtWidgets.QDialog):
    """Dialog to input the coordinate values of the corners of a roi rectangle"""
    def __init__(self, parent=None):
        super(RectRoiCoordinates, self).__init__(parent)

        empty_lbl  =  QtWidgets.QLabel("  ", self)
        empty_lbl.setFixedSize(10, 22)

        min_lbl  =  QtWidgets.QLabel("min", self)
        min_lbl.setFixedSize(55, 22)

        max_lbl  =  QtWidgets.QLabel("max", self)
        max_lbl.setFixedSize(55, 22)

        x_lbl  =  QtWidgets.QLabel("x", self)
        x_lbl.setFixedSize(10, 22)

        y_lbl  =  QtWidgets.QLabel("y", self)
        y_lbl.setFixedSize(10, 22)

        xmin_edt  =  QtWidgets.QLineEdit(self)
        xmin_edt.setFixedSize(55, 22)
        xmin_edt.textChanged[str].connect(self.xmin_var)

        xmax_edt  =  QtWidgets.QLineEdit(self)
        xmax_edt.setFixedSize(55, 22)
        xmax_edt.textChanged[str].connect(self.xmax_var)

        ymin_edt  =  QtWidgets.QLineEdit(self)
        ymin_edt.setFixedSize(55, 22)
        ymin_edt.textChanged[str].connect(self.ymin_var)
        ymin_edt.setText("-25.0")

        ymax_edt  =  QtWidgets.QLineEdit(self)
        ymax_edt.setFixedSize(55, 22)
        ymax_edt.textChanged[str].connect(self.ymax_var)
        ymax_edt.setText("25.0")

        enter_values_btn  =  QtWidgets.QPushButton("OK", self)
        enter_values_btn.setToolTip('Set these coordinates value')
        enter_values_btn.setFixedSize(100, 25)
        enter_values_btn.clicked.connect(self.enter_values)

        coord_lbl_box_1  =  QtWidgets.QHBoxLayout()
        coord_lbl_box_1.addWidget(empty_lbl)
        coord_lbl_box_1.addWidget(min_lbl)
        coord_lbl_box_1.addWidget(max_lbl)

        coord_lbl_box_2  =  QtWidgets.QHBoxLayout()
        coord_lbl_box_2.addWidget(x_lbl)
        coord_lbl_box_2.addWidget(xmin_edt)
        coord_lbl_box_2.addWidget(xmax_edt)

        coord_lbl_box_3  =  QtWidgets.QHBoxLayout()
        coord_lbl_box_3.addWidget(y_lbl)
        coord_lbl_box_3.addWidget(ymin_edt)
        coord_lbl_box_3.addWidget(ymax_edt)

        coord_ok_box  =  QtWidgets.QHBoxLayout()
        coord_ok_box.addStretch()
        coord_ok_box.addWidget(enter_values_btn)

        layout  =  QtWidgets.QVBoxLayout()
        layout.addLayout(coord_lbl_box_1)
        layout.addLayout(coord_lbl_box_2)
        layout.addLayout(coord_lbl_box_3)
        layout.addLayout(coord_ok_box)

        self.setWindowModality(Qt.ApplicationModal)
        self.setLayout(layout)
        self.setGeometry(300, 300, 200, 100)
        self.setWindowTitle("Set Coordinates")

    def xmin_var(self, text):
        """Set minimum x value"""
        self.xmin_value  =  float(text)

    def xmax_var(self, text):
        """Set maximum x value"""
        self.xmax_value  =  float(text)

    def ymin_var(self, text):
        """Set minimum y value"""
        self.ymin_value  =  float(text)

    def ymax_var(self, text):
        """Set maximum y value"""
        self.ymax_value  =  float(text)

    def enter_values(self):
        """Organize output as an array"""
        self.coord_vect  =  np.array([self.xmin_value, self.xmax_value, self.ymin_value, self.ymax_value])
        self.close()

    def coords(self):
        """Send output"""
        return self.coord_vect

    @staticmethod
    def getCoordinates(parent=None):
        dialog  =  RectRoiCoordinates(parent)
        result  =  dialog.exec_()
        flag    =  dialog.coords()
        return flag


class FilterSettings(QtWidgets.QDialog):
    """Dialog to set the filter values"""
    def __init__(self, parent=None):
        super(FilterSettings, self).__init__(parent)

        ones_numb_lbl  =  QtWidgets.QLabel("Ones", self)
        ones_numb_lbl.setFixedSize(55, 22)

        zeros_numb_lbl  =  QtWidgets.QLabel("Zeros", self)
        zeros_numb_lbl.setFixedSize(55, 22)

        zeros_numb_edt  =  QtWidgets.QLineEdit(self)
        zeros_numb_edt.setFixedSize(55, 22)
        zeros_numb_edt.setToolTip("Number of consecutive 0 surrounded by 1 that will be considered as misdetection")
        zeros_numb_edt.textChanged[str].connect(self.zeros_numb_var)
        zeros_numb_edt.setText("2")

        ones_numb_edt  =  QtWidgets.QLineEdit(self)
        ones_numb_edt.setFixedSize(55, 22)
        ones_numb_edt.setToolTip("Number of consecutive 1 surrounded by 0 that will be considered as fake detection")
        ones_numb_edt.textChanged[str].connect(self.ones_numb_var)
        ones_numb_edt.setText("30")

        enter_values_btn  =  QtWidgets.QPushButton("OK", self)
        enter_values_btn.setToolTip('Set filter parameters')
        enter_values_btn.setFixedSize(60, 25)
        enter_values_btn.clicked.connect(self.enter_values)

        ones_box  =  QtWidgets.QHBoxLayout()
        ones_box.addWidget(ones_numb_lbl)
        ones_box.addWidget(ones_numb_edt)

        zeros_box  =  QtWidgets.QHBoxLayout()
        zeros_box.addWidget(zeros_numb_lbl)
        zeros_box.addWidget(zeros_numb_edt)

        enter_box  =  QtWidgets.QHBoxLayout()
        enter_box.addStretch()
        enter_box.addWidget(enter_values_btn)

        layout  =  QtWidgets.QVBoxLayout()
        layout.addLayout(zeros_box)
        layout.addLayout(ones_box)
        layout.addLayout(enter_box)

        self.zeros_numb_value  =  2
        self.ones_numb_value   =  30

        self.setWindowModality(Qt.ApplicationModal)
        self.setLayout(layout)
        self.setGeometry(300, 300, 200, 100)
        self.setWindowTitle("Set Filter")

    def ones_numb_var(self, text):
        self.ones_numb_value  =  int(text)

    def zeros_numb_var(self, text):
        self.zeros_numb_value  =  int(text)

    def params(self):
        return self.filters_params

    def enter_values(self):
        self.filters_params  =  np.array([self.zeros_numb_value, self.ones_numb_value])
        self.close()

    @staticmethod
    def getSetting(parent=None):
        dialog  =  FilterSettings(parent)
        result  =  dialog.exec_()
        flag    =  dialog.params()
        return flag


class MultiPlotShowing(QtWidgets.QWidget):
    """Popup Tool to generate pmpts gallery of all the nuclei"""
    def __init__(self):
        QtWidgets.QWidget.__init__(self)

        fname_edt = QtWidgets.QLineEdit(self)
        fname_edt.setToolTip('Name of the file you are working on')

        frame  =  pg.PlotWidget(self)

        sld_tag  =  QtWidgets.QScrollBar(QtCore.Qt.Horizontal, self)
        sld_tag.valueChanged.connect(self.sld_tag_update)

        remove_current_index_btn  =  QtWidgets.QPushButton("Rm Current", self)
        remove_current_index_btn.setToolTip('Remove Current Spot Trace')
        remove_current_index_btn.setFixedSize(100, 25)
        remove_current_index_btn.clicked.connect(self.remove_current_index)
        # remove_current_index_btn.setEnabled(False)

        load_folder_btn  =  QtWidgets.QPushButton("Load Folder", self)
        load_folder_btn.setToolTip('Load Analysis Folder')
        load_folder_btn.setFixedSize(100, 25)
        load_folder_btn.clicked.connect(self.load_folder)

        restore_indexes_btn  =  QtWidgets.QPushButton("Restore Idxs", self)
        restore_indexes_btn.setToolTip('Restore Initial Indexes list')
        restore_indexes_btn.setFixedSize(100, 25)
        restore_indexes_btn.clicked.connect(self.restore_indexes)
        # restore_indexes_btn.setEnabled(False)

        add_xls_file_btn  =  QtWidgets.QPushButton("Add Xls", self)
        add_xls_file_btn.setToolTip('Add xls Files to Merge ')
        add_xls_file_btn.setFixedSize(100, 25)
        add_xls_file_btn.clicked.connect(self.add_xls_file)
        # add_xls_file_btn.setEnabled(False)

        merge_xls_file_btn  =  QtWidgets.QPushButton("Merge Xls", self)
        merge_xls_file_btn.setToolTip('Merge Selected xls files')
        merge_xls_file_btn.setFixedSize(100, 25)
        merge_xls_file_btn.clicked.connect(self.merge_xls_file)
        # merge_xls_file_btn.setEnabled(False)

        min_length_lbl  =  QtWidgets.QLabel('Min Length', self)
        min_length_lbl.setFixedSize(60, 25)
        # min_length_lbl.setEnabled(False)

        min_length_edt  =  QtWidgets.QLineEdit(self)
        min_length_edt.textChanged[str].connect(self.min_length_var)
        min_length_edt.setToolTip('Set Minimum Lenght of a Transcription Event')
        min_length_edt.setFixedSize(35, 25)
        # min_length_edt.setEnabled(False)

        min_length_box  =  QtWidgets.QHBoxLayout()
        min_length_box.addWidget(min_length_lbl)
        min_length_box.addWidget(min_length_edt)

        min_length_btn  =  QtWidgets.QPushButton("Rm Short", self)
        min_length_btn.setToolTip('Restore Initial Indexes list')
        min_length_btn.setFixedSize(100, 25)
        min_length_btn.clicked.connect(self.min_length)
        # min_length_btn.setEnabled(False)

        view_all_btn  =  QtWidgets.QPushButton("View All", self)
        view_all_btn.setToolTip('View All the Plots in One Composed Frame')
        view_all_btn.setFixedSize(100, 25)
        view_all_btn.clicked.connect(self.view_all)

        write_data_btn  =  QtWidgets.QPushButton("Write Data", self)
        write_data_btn.setToolTip('Write Data About Selected Spots')
        write_data_btn.setFixedSize(100, 25)
        write_data_btn.clicked.connect(self.write_data)
        # write_data_btn.setEnabled(False)

        frame_sld_box  = QtWidgets.QVBoxLayout()
        frame_sld_box.addWidget(fname_edt)
        frame_sld_box.addWidget(frame)
        frame_sld_box.addWidget(sld_tag)

        keys  =  QtWidgets.QVBoxLayout()
        keys.addWidget(load_folder_btn)
        keys.addStretch()
        keys.addWidget(remove_current_index_btn)
        keys.addLayout(min_length_box)
        keys.addWidget(min_length_btn)
        keys.addWidget(view_all_btn)
        keys.addStretch()
        keys.addWidget(write_data_btn)
        keys.addWidget(add_xls_file_btn)
        keys.addWidget(merge_xls_file_btn)
        keys.addStretch()
        keys.addWidget(restore_indexes_btn)

        layout  =  QtWidgets.QHBoxLayout()
        layout.addLayout(frame_sld_box)
        layout.addLayout(keys)

        self.frame      =  frame
        self.sld_tag    =  sld_tag
        self.fname_edt  =  fname_edt
        self.xls2merge  =  []

        self.setGeometry(300, 300, 1000, 400)
        self.setWindowTitle('Multi Plot Show')
        self.setLayout(layout)
        self.show()

    def load_folder(self):
        self.foldername  =  str(QtWidgets.QFileDialog.getExistingDirectory(None, 'Select a folder'))
        self.multi_trace  =  MultiTracePlotting_v2.MultiTracePlotting(self.foldername)
        self.fname_edt.setText(self.foldername)
        self.sld_tag.setMaximum(self.multi_trace.tags_list.size - 1)
        self.good_tags         =  np.copy(self.multi_trace.tags_list)
        self.spots_tracked_3d  =  np.copy(self.multi_trace.spots_tracked_3d)
        tag_text  =  pg.TextItem('tag = ' + str(self.multi_trace.tags_list[self.sld_tag.value()]), color='g')
        y         =  (self.multi_trace.raw_sp * (self.multi_trace.spots_tracked_3d == self.multi_trace.tags_list[0])).sum(2).sum(1)
        self.frame.setYRange(0, self.multi_trace.y_sup)
        self.frame.plot(y, pen='r', symbol='o', symbolSize=4)
        tag_text.setPos(1, self.multi_trace.y_sup)
        self.frame.addItem(tag_text)

    def min_length_var(self, text):
        self.min_length_value   =  int(text)

    def sld_tag_update(self):
        self.frame.clear()
        tag_text  =  pg.TextItem('tag = ' + str(self.good_tags[self.sld_tag.value()]), color='g')
        y         =  (self.multi_trace.raw_sp * (self.spots_tracked_3d == self.good_tags[self.sld_tag.value()])).sum(2).sum(1)
        self.frame.setYRange(0, self.multi_trace.y_sup)
        self.frame.plot(y, pen='r', symbol='o', symbolSize=4)
        tag_text.setPos(1, self.multi_trace.y_sup)
        self.frame.addItem(tag_text)

    def remove_current_index(self):
        self.sld_tag.setMaximum(self.sld_tag.maximum() - 1)
        self.frame.clear()
        self.good_tags  =  np.delete(self.good_tags, self.sld_tag.value(), axis=0)
        tag_text        =  pg.TextItem('tag = ' + str(self.good_tags[self.sld_tag.value()]), color='g')
        y               =  (self.multi_trace.raw_sp * (self.spots_tracked_3d == self.good_tags[self.sld_tag.value()])).sum(2).sum(1)
        self.frame.setYRange(0, self.multi_trace.y_sup)
        self.frame.plot(y, pen='r', symbol='o', symbolSize=4)
        tag_text.setPos(1, self.multi_trace.y_sup)
        self.frame.addItem(tag_text)

    def restore_indexes(self):
        self.frame.clear()
        self.good_tags         =  np.copy(self.multi_trace.tags_list)
        self.spots_tracked_3d  =  np.copy(self.multi_trace.spots_tracked_3d)
        self.sld_tag.setMaximum(self.good_tags.size - 1)
        tag_text  =  pg.TextItem('tag = ' + str(self.good_tags[self.sld_tag.value()]), color='g')
        y         =  (self.multi_trace.raw_sp * (self.spots_tracked_3d == self.good_tags[self.sld_tag.value()])).sum(2).sum(1)
        self.frame.setYRange(0, self.multi_trace.y_sup)
        self.frame.plot(y, pen='r', symbol='o', symbolSize=4)
        tag_text.setPos(1, self.multi_trace.y_sup)
        self.frame.addItem(tag_text)

    def min_length(self):
        self.frame.clear()
        reload(RemoveShortTraces2)
        bffr_data              =  RemoveShortTraces2.RemoveShortTraces2(self.good_tags, self.min_length_value, self.spots_tracked_3d)
        self.good_tags         =  bffr_data.new_tags
        self.spots_tracked_3d  =  bffr_data.spots_tracked_3d
        self.sld_tag.setMaximum(self.good_tags.size - 1)
        tag_text        =  pg.TextItem('tag = ' + str(self.good_tags[self.sld_tag.value()]), color='g')
        y               =  (self.multi_trace.raw_sp * (self.spots_tracked_3d == self.good_tags[self.sld_tag.value()])).sum(2).sum(1)
        self.frame.setYRange(0, self.multi_trace.y_sup)
        self.frame.plot(y, pen='r', symbol='o', symbolSize=4)
        tag_text.setPos(1, self.multi_trace.y_sup)
        self.frame.addItem(tag_text)

    def view_all(self):
        n_rows   =  6
        n_cols   =  7
        num_win  =  self.good_tags.size // (n_cols * n_rows) + 1
        for win_idxs in range(num_win):
            time.sleep(3)
            str_win1  =  "win" + str(win_idxs) + "  =  pg.GraphicsWindow()"
            str_win2  =  "win" + str(win_idxs) + ".setWindowTitle('Transcriptional Traces " + str(win_idxs + 1) + "')"
            exec(str_win1)
            exec(str_win2)

            idx_name = 0
            for idx_r in range(n_rows):
                for idx_c in range(n_cols):
                    str_addplot  =  "p" + str(idx_name) +  "= win" + str(win_idxs) + ".addPlot(row=" + str(idx_r) + ", col=" + str(idx_c) + ")"
                    exec(str_addplot)
                    idx_name  +=  1

            for k in range(n_cols * n_rows):
                if k + win_idxs * n_cols * n_rows <= self.good_tags.size - 1:
                    str_cmnd1  =  "p" + str(k) + ".plot((self.multi_trace.raw_sp * (self.spots_tracked_3d == self.good_tags[k + win_idxs * n_cols * n_rows])).sum(2).sum(1), pen='r', symbol='o', symbolSize=2)"
                    str_cmnd2  =  "p" + str(k) + ".setYRange(0, self.multi_trace.y_sup)"
                    str_cmnd3  =  "tag_text" + str(k) + " = pg.TextItem('tag = ' + str(self.good_tags[k + win_idxs * n_cols * n_rows]), color='g')"
                    str_cmnd4  =  "tag_text" + str(k) + ".setPos(1, self.multi_trace.y_sup)"
                    str_cmnd5  =  "p" + str(k) + ".addItem(tag_text" + str(k) + ")"
                    exec(str_cmnd1)
                    exec(str_cmnd2)
                    exec(str_cmnd3)
                    exec(str_cmnd4)
                    exec(str_cmnd5)
                else:
                    break
        print(StrangePatch)

    def write_data(self):
        fname2write  =  QtWidgets.QFileDialog.getSaveFileName(self, 'Select Path and Name for Results File', self.foldername + '/SteadySpotsFeatures.xls')
        WriteSteadySpotsBursting.WriteSteadySpotsBursting(self.foldername, fname2write, self.multi_trace, self.good_tags, self.min_length_value)

    def add_xls_file(self):
        xls_fname  =  str(QtWidgets.QFileDialog.getOpenFileName(None, 'Select an xls file', filter='*.xls'))
        self.xls2merge.append(xls_fname)
        self.mpp12  =  ListOfXlsSelected(self.xls2merge)

    def merge_xls_file(self):
        fname2write  =  QtWidgets.QFileDialog.getSaveFileName(self, 'Select Path and Name for Merged xls File')
        MergeXlsFiles.MergeXlsFiles(fname2write, self.xls2merge)


class ListOfXlsSelected(QtWidgets.QWidget):
    def __init__(self, xls2merge):
        QtWidgets.QWidget.__init__(self)

        layout  =  QtWidgets.QVBoxLayout()
        for j in range(len(xls2merge)):
            fname_lbl  =  QtWidgets.QLabel(xls2merge[j], self)
            fname_lbl.setFixedHeight(25)
            layout.addWidget(fname_lbl)

        self.setGeometry(30, 30, 1000, 100)
        self.setWindowTitle('Selected xls to Merge')
        self.setLayout(layout)
        self.show()


class SetColorChannel(QtWidgets.QWidget):
    """Set the color channels of the raw data to put in the gui"""
    procStart  =  QtCore.pyqtSignal()

    def __init__(self, nucs_spts_ch):
        QtWidgets.QWidget.__init__(self)
        print(nucs_spts_ch)
        nuclei_channel_lbl  =  QtWidgets.QLabel("Nuclei Channel", self)
        nuclei_channel_lbl.setFixedSize(100, 22)

        spots_channel_lbl  =  QtWidgets.QLabel("Spots Channel", self)
        spots_channel_lbl.setFixedSize(100, 22)

        nuclei_channel_combo  =  QtWidgets.QComboBox(self)
        nuclei_channel_combo.addItem("1")
        nuclei_channel_combo.addItem("2")
        nuclei_channel_combo.addItem("3")
        nuclei_channel_combo.activated[str].connect(self.nuclei_channel_switch)
        nuclei_channel_combo.setCurrentIndex(nucs_spts_ch[0])
        nuclei_channel_combo.setFixedSize(45, 25)

        spots_channel_combo  =  QtWidgets.QComboBox(self)
        spots_channel_combo.addItem("1")
        spots_channel_combo.addItem("2")
        spots_channel_combo.addItem("3")
        spots_channel_combo.activated[str].connect(self.spots_channel_switch)
        spots_channel_combo.setCurrentIndex(nucs_spts_ch[1])
        spots_channel_combo.setFixedSize(45, 25)

        enter_values_btn  =  QtWidgets.QPushButton("OK", self)
        enter_values_btn.setToolTip('Set Channels Number')
        enter_values_btn.setFixedSize(60, 25)
        enter_values_btn.clicked.connect(self.enter_values)

        nuclei_box  =  QtWidgets.QHBoxLayout()
        nuclei_box.addWidget(nuclei_channel_lbl)
        nuclei_box.addWidget(nuclei_channel_combo)

        spots_box  =  QtWidgets.QHBoxLayout()
        spots_box.addWidget(spots_channel_lbl)
        spots_box.addWidget(spots_channel_combo)

        enter_box  =  QtWidgets.QHBoxLayout()
        enter_box.addStretch()
        enter_box.addWidget(enter_values_btn)

        layout  =  QtWidgets.QVBoxLayout()
        layout.addLayout(nuclei_box)
        layout.addLayout(spots_box)
        layout.addLayout(enter_box)

        self.nuclei_channel  =  nucs_spts_ch[0]
        self.spots_channel   =  nucs_spts_ch[1]

        self.setWindowModality(Qt.ApplicationModal)
        self.setLayout(layout)
        self.setGeometry(300, 300, 200, 100)
        self.setWindowTitle("Set Channels")

    def nuclei_channel_switch(self, text):
        """Set the nuclei channel"""
        self.nuclei_channel  =  int(text)

    def spots_channel_switch(self, text):
        """Set the spots channel"""
        self.spots_channel  =  int(text)

    @QtCore.pyqtSlot()
    def enter_values(self):
        """Enter the channel values"""
        self.channels_values  =  np.array([self.nuclei_channel, self.spots_channel])
        self.procStart.emit()


class MultiColorIntensity(QtWidgets.QWidget):
    """Generate a multicolor false video to give nuclei a color depending on the intensity of the spots"""
    def __init__(self):
        QtWidgets.QWidget.__init__(self)

        list_widget  =  QtWidgets.QListWidget()

        add_folders_btn  =  QtWidgets.QPushButton("Add Folder", self)
        add_folders_btn.setToolTip("Create a list of analysis folder to merge info")
        add_folders_btn.setFixedSize(120, 25)
        add_folders_btn.clicked.connect(self.add_folder)

        ok_btn  =  QtWidgets.QPushButton("Run", self)
        ok_btn.setToolTip("Analyze and show ")
        ok_btn.setFixedSize(120, 25)
        ok_btn.clicked.connect(self.ok_run)

        write_video_btn  =  QtWidgets.QPushButton("Write Video", self)
        write_video_btn.setToolTip("Write video in a .avi video file")
        write_video_btn.setFixedSize(120, 25)
        write_video_btn.clicked.connect(self.write_video)

        commands  =  QtWidgets.QVBoxLayout()
        commands.addWidget(add_folders_btn)
        commands.addWidget(ok_btn)
        commands.addStretch()
        commands.addWidget(write_video_btn)

        layout  =  QtWidgets.QHBoxLayout()
        layout.addWidget(list_widget)
        layout.addLayout(commands)

        self.list_widget  =  list_widget
        self.all_folders  =  []

        self.setGeometry(30, 30, 1000, 100)
        self.setWindowTitle('Multi Color Intensity')
        self.setLayout(layout)
        self.show()

    def add_folder(self):
        foldername  =  str(QtWidgets.QFileDialog.getExistingDirectory(None, "Select the folder with the analyzed data"))
        self.list_widget.addItem(str(foldername))
        self.all_folders.append(foldername)

    def ok_run(self):
        self.multi_clrs_images  =  MultiColorIntensityGenerator.MultiColorIntensityGenerator(self.all_folders)

    def write_video(self):
        fwritename  =  QtWidgets.QFileDialog.getExistingDirectory(None, "Select the folder to write the video in")
        for ll in range(len(self.multi_clrs_images.nucs_multiclrs_list)):
            fname  =  self.all_folders[ll][len(self.all_folders[ll]) - self.all_folders[ll][::-1].find("/"):]
            tifffile.imwrite(str(fwritename) + "/false_MultiColors_" + fname + ".tif", np.rot90(self.multi_clrs_images.nucs_multiclrs_list[ll][:, :, ::-1, :], axes=(1, 2)).astype("uint8"))


class SpotsTimeAverage(QtWidgets.QWidget):
    """Popup tool to average spots intensity over a small time window and write results"""
    def __init__(self, foldername, selected_spots, t_step, time_zero):

        QtWidgets.QWidget.__init__(self)

        activ_profile  =  np.sign(selected_spots).sum(1)
        int_prof2show  =  np.zeros(selected_spots.shape[0])
        print(t_step)
        print(time_zero)
        for k in range(selected_spots.shape[0]):
            bffr_data         =  selected_spots[k, :]
            int_prof2show[k]  =  np.nan_to_num(bffr_data[bffr_data != 0].mean())

        activ_profile  =  np.append(np.zeros(time_zero), activ_profile)
        int_prof2show  =  np.append(np.zeros(time_zero), int_prof2show)

        fname_edt  =  QtWidgets.QLineEdit(self)
        fname_edt.setToolTip('Names and paths of the loaded files')
        fname_edt.setText(foldername)

        frame_plot  =  pg.PlotWidget(self)
        frame_plot.addLegend()
        frame_plot.plot(activ_profile, symbol='x', pen='r', name='Act Prof')
        frame_plot.plot(int_prof2show * activ_profile.max() / int_prof2show.max(), pen='y', name='Int Prof')

        ok_btn  =  QtWidgets.QPushButton("Ok", self)
        ok_btn.setToolTip('Write the .xls file with extracted information')
        ok_btn.setFixedSize(100, 25)
        ok_btn.clicked.connect(self.ok_function)

        t1_lbl  =  QtWidgets.QLabel('t1', self)
        t1_lbl.setFixedSize(50, 25)

        t1_edt  =  QtWidgets.QLineEdit(self)
        t1_edt.textChanged[str].connect(self.t1_var)
        t1_edt.returnPressed.connect(self.set_t1)
        t1_edt.setToolTip('First Step to Consider')
        t1_edt.setFixedSize(55, 25)

        t1_box  =  QtWidgets.QHBoxLayout()
        t1_box.addWidget(t1_lbl)
        t1_box.addWidget(t1_edt)

        t2_lbl  =  QtWidgets.QLabel('t2', self)
        t2_lbl.setFixedSize(50, 25)

        t2_edt  =  QtWidgets.QLineEdit(self)
        t2_edt.textChanged[str].connect(self.t2_var)
        t2_edt.returnPressed.connect(self.set_t2)
        t2_edt.setToolTip('Last Step to Consider')
        t2_edt.setFixedSize(55, 25)

        t2_box  =  QtWidgets.QHBoxLayout()
        t2_box.addWidget(t2_lbl)
        t2_box.addWidget(t2_edt)

        keys  =  QtWidgets.QVBoxLayout()
        keys.addLayout(t1_box)
        keys.addLayout(t2_box)
        keys.addWidget(ok_btn)
        keys.addStretch()

        keys_frame_box  =  QtWidgets.QHBoxLayout()
        keys_frame_box.addWidget(frame_plot)
        keys_frame_box.addLayout(keys)

        layout  =  QtWidgets.QVBoxLayout()
        layout.addWidget(fname_edt)
        layout.addLayout(keys_frame_box)

        roi  =  pg.LinearRegionItem(orientation=False)
        roi.sigRegionChanged.connect(self.update_t1_t2)
        frame_plot.addItem(roi)

        self.t_step          =  t_step
        self.t1_edt          =  t1_edt
        self.t2_edt          =  t2_edt
        self.roi             =  roi
        self.time_zero       =  time_zero
        self.frame_plot      =  frame_plot
        self.activ_profile   =  activ_profile
        self.foldername      =  foldername
        self.int_prof2show   =  int_prof2show
        self.selected_spots  =  selected_spots

        self.set_time_scale()

        self.setLayout(layout)
        self.setGeometry(300, 300, 800, 600)
        self.setWindowTitle("Spatial Analysis")
        self.show()

    def update_t1_t2(self):
        self.t1_edt.setText(time.strftime("%M:%S", time.gmtime(np.round(self.roi.getRegion()[0], decimals=2))))
        self.t2_edt.setText(time.strftime("%M:%S", time.gmtime(np.round(self.roi.getRegion()[1], decimals=2))))

    def t1_var(self, text):
        seconds        =  reduce(lambda x, y: x * 60 + y, [float(i) for i in (text.replace(':', ',')).split(',')])
        self.t1_value  =  float(seconds)
        self.t1_text   =  text

    def t2_var(self, text):
        seconds        =  reduce(lambda x, y: x * 60 + y, [float(i) for i in (text.replace(':', ',')).split(',')])
        self.t2_value  =  float(seconds)
        self.t2_text   =  text

    def set_t1(self):
        self.roi.setRegion([self.t1_value, self.roi.getRegion()[1]])

    def set_t2(self):
        self.roi.setRegion([self.roi.getRegion()[0], self.t2_value])

    def set_time_scale(self):
        self.frame_plot.clear()
        self.frame_plot.plot(np.arange(self.activ_profile.size) * self.t_step, self.activ_profile, symbol='x', pen='r')
        self.frame_plot.plot(np.arange(self.activ_profile.size) * self.t_step, self.int_prof2show * self.activ_profile.max() / self.int_prof2show.max(), pen='y')
        self.frame_plot.addItem(self.roi)

    def ok_function(self):
        filename  =  QtWidgets.QFileDialog.getSaveFileName(None, "Select xls file to modify or define a new one...")[0]
        AvSpotsTime.AvSpotsTime(filename, self.foldername, self.selected_spots, int(self.t1_value / self.t_step) - self.time_zero - 1, int(self.t2_value / self.t_step) - self.time_zero - 1, self.t1_text, self.t2_text)


class SaturationInfo(QtWidgets.QWidget):
    """Tool to check eventual data saturation"""
    def __init__(self, green4D):
        QtWidgets.QWidget.__init__(self)

        max_value_lbl  =  QtWidgets.QLabel('Max value = ' + str(green4D.max()), self)
        max_value_lbl.setFixedSize(300, 40)

        num_saturated_lbl  =  QtWidgets.QLabel('Number of saturated pixels = ' + str(np.sum(green4D == 65535)), self)
        num_saturated_lbl.setFixedSize(300, 40)

        layout  =  QtWidgets.QVBoxLayout()
        layout.addWidget(max_value_lbl)
        layout.addWidget(num_saturated_lbl)

        self.setLayout(layout)
        self.setGeometry(300, 340, 250, 80)
        self.setWindowTitle("Saturation Info")


class FlipTile(QtWidgets.QWidget):
    """Popup tool to flip tile scan image"""
    procStart  =  QtCore.pyqtSignal()

    def __init__(self, tile_fname):
        QtWidgets.QWidget.__init__(self)

        tile_img  =  FromTile2GlobCoordinate.FromTile2GlobCoordinateLoader(tile_fname).tile_img

        framepp1  =  pg.ImageView(self)
        framepp1.ui.roiBtn.hide()
        framepp1.ui.menuBtn.hide()
        framepp1.setImage(tile_img)

        flip_tile_btn  =  QtWidgets.QPushButton("Horizontal Flip", self)
        flip_tile_btn.setFixedSize(120, 25)
        flip_tile_btn.clicked.connect(self.flip_tile)
        flip_tile_btn.setToolTip('Horizontal flip of the tile image')

        done_btn  =  QtWidgets.QPushButton("Ok", self)
        done_btn.setFixedSize(120, 25)
        done_btn.clicked.connect(self.done)
        done_btn.setToolTip('Done')

        keys  =  QtWidgets.QHBoxLayout()
        keys.addStretch()
        keys.addWidget(flip_tile_btn)
        keys.addWidget(done_btn)

        layout  =  QtWidgets.QVBoxLayout()
        layout.addWidget(framepp1)
        layout.addLayout(keys)

        self.tile_img  =  tile_img
        self.framepp1  =  framepp1

        self.setLayout(layout)
        self.setGeometry(300, 340, 450, 450)
        self.setWindowTitle("Flip Tile")

    def flip_tile(self):
        self.tile_img  =  self.tile_img[::-1, :]
        self.framepp1.setImage(self.tile_img)

    @QtCore.pyqtSlot()
    def done(self):
        self.procStart.emit()


class CheckSelectedRawData(QtWidgets.QWidget):
    procStart  =  QtCore.pyqtSignal()

    def __init__(self, frst_fr, nucs_spts_ch):
        QtWidgets.QWidget.__init__(self)

        # foldername      =  str(QtWidgets.QFileDialog.getExistingDirectory(None, "Select the folder with the analyzed data"))
        self.raw_fname  =  QtWidgets.QFileDialog.getOpenFileNames(None, "Select raw data files with mitosis", filter="*.lsm *.czi *.tif *.lif")[0]

        if str(self.raw_fname[0])[-3:] == 'lsm' or str(self.raw_fname[0])[-3:] == 'tif':
            self.raw_data_nucs  =  MultiLoadLsmOrTif5D.MultiLoadLsmOrTif5D(self.raw_fname, nucs_spts_ch).imarray_red

        if str(self.raw_fname[0])[-3:] == 'czi':
            self.raw_data_nucs  =  MultiLoadCzi5D.MultiProcLoadCzi5D(self.raw_fname, nucs_spts_ch).imarray_red

        if str(self.raw_fname[0])[-3:] == 'lif':
            self.raw_data_nucs  =  MultiLoadLif5D.MultiLoadLif5D(self.raw_fname, nucs_spts_ch).imarray_red

        prof     =  np.abs(frst_fr - self.raw_data_nucs).sum(2).sum(1)
        self.pz  =  np.where(prof == 0)[0]

        frame_img  =  pg.ImageView(self)
        frame_img.ui.roiBtn.hide()
        frame_img.ui.menuBtn.hide()
        frame_img.setImage(self.raw_data_nucs)
        frame_img.timeLine.sigPositionChanged.connect(self.upadateframe)

        frst_frame_lbl  =  QtWidgets.QLabel(' ', self)
        frst_frame_lbl.setFixedSize(200, 25)

        frame_numb_lbl  =  QtWidgets.QLabel("frame 0", self)
        frame_numb_lbl.setFixedSize(200, 25)

        set_mtss_btn  =  QtWidgets.QPushButton("Set Mtss", self)
        set_mtss_btn.clicked.connect(self.set_mtss)
        set_mtss_btn.setToolTip('Set Mitosis Frame')
        set_mtss_btn.setFixedSize(60, 25)

        no_btn  =  QtWidgets.QPushButton("No", self)
        no_btn.clicked.connect(self.no_func)
        no_btn.setToolTip('Missing Info')
        no_btn.setFixedSize(50, 25)

        yes_no_box  =  QtWidgets.QHBoxLayout()
        yes_no_box.addStretch()
        yes_no_box.addWidget(frst_frame_lbl)
        yes_no_box.addStretch()
        yes_no_box.addWidget(frame_numb_lbl)
        yes_no_box.addWidget(no_btn)
        yes_no_box.addWidget(set_mtss_btn)

        layout  =  QtWidgets.QVBoxLayout()
        layout.addWidget(frame_img)
        layout.addLayout(yes_no_box)

        if self.pz.size == 0:
            frst_frame_lbl.setText("First Analysis Frame not Included")
            frst_frame_lbl.setStyleSheet('color: red')
            set_mtss_btn.setEnabled(False)
        else:
            frst_frame_lbl.setText("First Analysis Frame is " + str(self.pz[0]))
            set_mtss_btn.setEnabled(True)

        self.frst_frame_lbl  =  frst_frame_lbl
        self.frame_numb_lbl  =  frame_numb_lbl
        self.frame_img       =  frame_img

        self.setWindowModality(Qt.ApplicationModal)
        self.setLayout(layout)
        self.setGeometry(300, 300, 600, 600)
        self.setWindowTitle("Check Raw Data")

    def upadateframe(self):
        self.frame_numb_lbl.setText("frame " + str(self.frame_img.currentIndex))

    def set_mtss(self):
        self.flag_yes_no  =  "yes"
        self.mtss_frame  =  self.frame_img.currentIndex
        print(self.mtss_frame)
        self.done()

    def no_func(self):
        self.flag_yes_no  =  "no"
        self.done()

    @QtCore.pyqtSlot()
    def done(self):
        self.procStart.emit()


class CalibrationFactor(QtWidgets.QDialog):
    """Popup tool to rescale the spot intensity with a factor input by the user"""
    def __init__(self, parent=None):
        super(CalibrationFactor, self).__init__(parent)

        calib_factor_lbl  =  QtWidgets.QLabel("Calib Factor", self)
        calib_factor_lbl.setFixedSize(110, 25)

        calib_factor_edt = QtWidgets.QLineEdit(self)
        calib_factor_edt.setToolTip("Calibration Factor Value")
        calib_factor_edt.setFixedSize(50, 22)
        calib_factor_edt.textChanged[str].connect(self.calib_factor_var)

        input_close_btn  =  QtWidgets.QPushButton("Ok", self)
        input_close_btn.clicked.connect(self.input_close)
        input_close_btn.setToolTip('Input values')
        input_close_btn.setFixedSize(50, 25)

        calib_factor_lbl_edit_box  =  QtWidgets.QHBoxLayout()
        calib_factor_lbl_edit_box.addWidget(calib_factor_lbl)
        calib_factor_lbl_edit_box.addWidget(calib_factor_edt)

        input_close_box  =  QtWidgets.QHBoxLayout()
        input_close_box.addStretch()
        input_close_box.addWidget(input_close_btn)

        layout  =  QtWidgets.QVBoxLayout()
        layout.addLayout(calib_factor_lbl_edit_box)
        layout.addLayout(input_close_box)

        self.setWindowModality(Qt.ApplicationModal)
        self.setLayout(layout)
        self.setGeometry(300, 300, 100, 50)
        self.setWindowTitle("Set border pixels")

    def calib_factor_var(self, text):
        self.calib_factor_value  =  float(text)

    def input_close(self):
        self.close()

    def calib_factor(self):
        return self.calib_factor_value

    @staticmethod
    def getCalibFact(parent=None):
        dialog        =  CalibrationFactor(parent)
        result        =  dialog.exec_()
        calib_factor  =  dialog.calib_factor()
        return calib_factor


class RemoveMitoticalSpots(QtWidgets.QWidget):
    """Popup tool to track and remove mitotical TS"""
    procStart  =  QtCore.pyqtSignal()

    def __init__(self, imarray_red, imarray_green, green4D, fnames, nucs_spts_ch):
        QtWidgets.QWidget.__init__(self)

        self.imarray_red          =  imarray_red
        self.imarray_green        =  imarray_green
        self.green4D              =  green4D

        filedata2show         =  MultiLoadCzi5D.MultiProcLoadCzi5D(fnames, nucs_spts_ch)
        raw2show              =  np.zeros(filedata2show.imarray_red.shape + (3, ))
        raw2show[:, :, :, 0]  =  filedata2show.imarray_red / 3
        raw2show[:, :, :, 1]  =  filedata2show.imarray_green
        w                     =  pg.image(raw2show)

        raw2chs              =  np.zeros(self.imarray_red.shape + (3,))
        raw2chs[:, :, :, 0]  =  self.imarray_red / 2
        raw2chs[:, :, :, 1]  =  self.imarray_green

        framepp_3chs  =  pg.ImageView(self, name="Frame_3chs")
        framepp_3chs.ui.roiBtn.hide()
        framepp_3chs.ui.menuBtn.hide()
        framepp_3chs.setImage(raw2chs)
        framepp_3chs.timeLine.sigPositionChanged.connect(self.update_frame_segtrk)

        framepp_segtrk  =  pg.ImageView(self, name="framepp_seg_trk")
        framepp_segtrk.ui.roiBtn.hide()
        framepp_segtrk.ui.menuBtn.hide()
        framepp_segtrk.setImage(np.zeros(self.imarray_green.shape))
        framepp_segtrk.timeLine.sigPositionChanged.connect(self.update_frame_3chs)

        framepp_3chs.view.setXLink("framepp_seg_trk")
        framepp_3chs.view.setYLink("framepp_seg_trk")
        framepp_segtrk.view.setXLink("framepp_3chs")
        framepp_segtrk.view.setYLink("framepp_3chs")

        tabs_left   =  QtWidgets.QTabWidget()
        tab_3chs    =  QtWidgets.QWidget()
        tab_segtrk  =  QtWidgets.QWidget()

        frame_3chs_box  =  QtWidgets.QHBoxLayout()
        frame_3chs_box.addWidget(framepp_3chs)

        frame_segtrk_box  =  QtWidgets.QHBoxLayout()
        frame_segtrk_box.addWidget(framepp_segtrk)

        tab_3chs.setLayout(frame_3chs_box)
        tab_segtrk.setLayout(frame_segtrk_box)

        tabs_left.addTab(tab_3chs, "Raw 3 chs")
        tabs_left.addTab(tab_segtrk, "Segm or Track")

        frame_lbl  =  QtWidgets.QLabel(self)
        frame_lbl.setFixedSize(65, 25)
        frame_lbl.setText("Frame 0")

        frame_chosen_lbl  =  QtWidgets.QLabel(self)
        frame_chosen_lbl.setFixedSize(35, 25)
        frame_chosen_lbl.setText("(0)")
        frame_chosen_lbl.setStyleSheet('color: red')

        frame_left_box  =  QtWidgets.QHBoxLayout()
        frame_left_box.addWidget(frame_lbl)
        frame_left_box.addWidget(frame_chosen_lbl)

        spts_thr_lbl  =  QtWidgets.QLabel(self)
        spts_thr_lbl.setFixedSize(51, 25)
        spts_thr_lbl.setText("Spts Thr")

        spts_thr_edt  =  QtWidgets.QLineEdit(self)
        spts_thr_edt.textChanged[str].connect(self.spts_thr_var)
        spts_thr_edt.setToolTip("Threshold to segment spots (suggested value 7)")
        spts_thr_edt.setFixedSize(20, 25)

        spts_thr_box  =  QtWidgets.QHBoxLayout()
        spts_thr_box.addWidget(spts_thr_lbl)
        spts_thr_box.addWidget(spts_thr_edt)

        vol_thr_lbl  =  QtWidgets.QLabel(self)
        vol_thr_lbl.setFixedSize(51, 25)
        vol_thr_lbl.setText("Min Vol")

        vol_thr_edt  =  QtWidgets.QLineEdit(self)
        vol_thr_edt.textChanged[str].connect(self.vol_thr_var)
        vol_thr_edt.setToolTip("Set the minimum volume for a spot (suggested value 5)")
        vol_thr_edt.setFixedSize(20, 25)

        vol_thr_box  =  QtWidgets.QHBoxLayout()
        vol_thr_box.addWidget(vol_thr_lbl)
        vol_thr_box.addWidget(vol_thr_edt)

        spts_detect_params_box  =  QtWidgets.QHBoxLayout()
        spts_detect_params_box.addLayout(spts_thr_box)
        spts_detect_params_box.addLayout(vol_thr_box)

        t_track_end_btn  =  QtWidgets.QPushButton("T End", self)
        t_track_end_btn.clicked.connect(self.t_track_end)
        t_track_end_btn.setToolTip('Spots appearing up to this frframe_3chs_box will be tracked and removd')
        t_track_end_btn.setFixedSize(90, 25)

        frame_lbl_tstart_box  =  QtWidgets.QVBoxLayout()
        frame_lbl_tstart_box.addLayout(frame_left_box)
        frame_lbl_tstart_box.addWidget(t_track_end_btn)

        segm_btn  =  QtWidgets.QPushButton("Segm", self)
        segm_btn.clicked.connect(self.segm)
        segm_btn.setToolTip('Segment Spots')
        segm_btn.setFixedSize(170, 25)

        spts_detect_box  =  QtWidgets.QVBoxLayout()
        spts_detect_box.addLayout(spts_detect_params_box)
        spts_detect_box.addWidget(segm_btn)

        dist_thr_lbl  =  QtWidgets.QLabel(self)
        dist_thr_lbl.setFixedSize(51, 25)
        dist_thr_lbl.setText("Dist Thr")

        dist_thr_edt  =  QtWidgets.QLineEdit(self)
        dist_thr_edt.textChanged[str].connect(self.dist_thr_var)
        dist_thr_edt.setToolTip("Set the maximum distance between two consecutive positions of a spot (suggested value 15)")
        dist_thr_edt.setFixedSize(45, 25)

        dist_thr_box  =  QtWidgets.QHBoxLayout()
        dist_thr_box.addWidget(dist_thr_lbl)
        dist_thr_box.addWidget(dist_thr_edt)

        track_btn  =  QtWidgets.QPushButton("Track", self)
        track_btn.clicked.connect(self.track)
        track_btn.setToolTip('Track Spots')
        track_btn.setFixedSize(110, 25)

        track_box  =  QtWidgets.QVBoxLayout()
        track_box.addLayout(dist_thr_box)
        track_box.addWidget(track_btn)

        vert_line_one  =  QtWidgets.QFrame()
        vert_line_one.setFrameStyle(QtWidgets.QFrame.VLine)

        vert_line_two  =  QtWidgets.QFrame()
        vert_line_two.setFrameStyle(QtWidgets.QFrame.VLine)

        vert_line_three  =  QtWidgets.QFrame()
        vert_line_three.setFrameStyle(QtWidgets.QFrame.VLine)

        send_btn  =  QtWidgets.QPushButton("Send", self)
        send_btn.clicked.connect(self.send)
        send_btn.setToolTip('Send results to the main GUI')
        send_btn.setFixedSize(110, 25)

        commands_box  =  QtWidgets.QHBoxLayout()
        commands_box.addLayout(spts_detect_box)
        commands_box.addWidget(vert_line_one)
        commands_box.addLayout(frame_lbl_tstart_box)
        commands_box.addWidget(vert_line_two)
        commands_box.addLayout(track_box)
        commands_box.addWidget(vert_line_three)
        commands_box.addStretch()
        commands_box.addWidget(send_btn)

        layout  =  QtWidgets.QVBoxLayout()
        layout.addWidget(tabs_left)
        layout.addLayout(commands_box)

        self.setLayout(layout)
        self.setGeometry(300, 300, 600, 400)
        self.setWindowTitle("Remove Mitotical TS")

        self.framepp_3chs       =  framepp_3chs
        self.framepp_segtrk     =  framepp_segtrk
        self.frame_lbl          =  frame_lbl
        self.frame_chosen_lbl   =  frame_chosen_lbl
        self.imarray_green      =  imarray_green
        self.red_ch_ref         =  filedata2show.imarray_green
        self.t_track_end_value  =  0
        self.w                  =  w

    def update_frame_3chs(self):
        """Update frame 3chs from segmented"""
        self.framepp_3chs.setCurrentIndex(self.framepp_segtrk.currentIndex)
        self.frame_lbl.setText("Frame " + str(self.framepp_3chs.currentIndex))

    def update_frame_segtrk(self):
        """Update frame segmented from 3chs"""
        self.framepp_segtrk.setCurrentIndex(self.framepp_3chs.currentIndex)

    def t_track_end(self):
        diff_frms  =  10000
        t          =  -1
        while diff_frms != 0:
            t          +=  1
            diff_frms   =  np.sum(self.imarray_green[t] - self.red_ch_ref[self.w.currentIndex])

        self.t_track_end_value  =  t
        # self.t_track_end_value  =  int(self.framepp_3chs.currentIndex)
        self.frame_chosen_lbl.setText("(" + str(self.t_track_end_value) + ")")

    def spts_thr_var(self, text):
        """Define spots threshold value"""
        self.spts_thr_value  =  float(text)

    def vol_thr_var(self, text):
        """Define spots volume threshold"""
        self.vol_thr_value  =  float(text)

    def dist_thr_var(self, text):
        """Define spots distance threshold"""
        self.dist_thr_value  =  float(text)

    def segm(self):
        """Segment spots for ttracking"""
        self.spots_3D  =  SpotsDetectionChopper.SpotsDetectionChopper(self.green4D, self.spts_thr_value, self.vol_thr_value)
        self.framepp_segtrk.setImage(np.sign(self.spots_3D.spots_vol))
        self.framepp_segtrk.updateImage()

    def track(self):
        """Track mitotical spots"""
        reload(VisualTracked)
        reload(SpotTracker)
        self.spts_trck_info  =  SpotTracker.SpotTracker(np.copy(self.spots_3D.spots_tzxy), self.dist_thr_value, self.t_track_end_value).spts_trck_info
        bff2show             =  VisualTracked.VisualTracked(self.spots_3D.spots_coords, self.spts_trck_info, self.spots_3D.spots_vol, self.green4D.shape).visual_tracked
        self.framepp_segtrk.setImage(bff2show)
        self.framepp_segtrk.updateImage()

    @QtCore.pyqtSlot()
    def send(self):
        """Send changes to the main GUI"""
        reload(VisualTracked)
        self.spots_3D  =  VisualTracked.RemoveTracked(self.spots_3D.spots_coords, self.spots_3D.spots_tzxy, self.spots_3D.spots_ints, self.spots_3D.spots_vol, self.spts_trck_info, self.green4D)
        self.procStart.emit()


class TracesImageTool(QtWidgets.QWidget):
    """Generate a traces image, even pooling several analyses"""
    def __init__(self):
        QtWidgets.QWidget.__init__(self)

        list_widget  =  QtWidgets.QListWidget()

        add_folders_btn  =  QtWidgets.QPushButton("Add Folder", self)
        add_folders_btn.setToolTip("Create a list of analysis folder to merge info")
        add_folders_btn.setFixedSize(120, 25)
        add_folders_btn.clicked.connect(self.add_folder)

        ok_btn  =  QtWidgets.QPushButton("Run", self)
        ok_btn.setToolTip("Analyze and show ")
        ok_btn.setFixedSize(120, 25)
        ok_btn.clicked.connect(self.ok_run)

        commands  =  QtWidgets.QVBoxLayout()
        commands.addStretch()
        commands.addWidget(add_folders_btn)
        commands.addWidget(ok_btn)

        layout  =  QtWidgets.QHBoxLayout()
        layout.addWidget(list_widget)
        layout.addLayout(commands)

        self.list_widget  =  list_widget
        self.all_folders  =  []

        self.setGeometry(30, 30, 1000, 100)
        self.setWindowTitle('Multi Color Intensity')
        self.setLayout(layout)
        self.show()

    def add_folder(self):
        """Add a folder to the folder list"""
        foldername  =  str(QtWidgets.QFileDialog.getExistingDirectory(None, "Select the folder with the analyzed data"))
        self.list_widget.addItem(str(foldername))
        self.all_folders.append(foldername)

    def ok_run(self):
        """Generaate the trace image"""
        TracesImage.TracesImage(self.all_folders)


class RemoveNucleiDust(QtWidgets.QWidget):
    """Popup tool to track and remove mitotical TS"""
    procStart  =  QtCore.pyqtSignal()

    def __init__(self, imarray_red, nuclei_tracked):
        QtWidgets.QWidget.__init__(self)

        mycmap  =  np.fromfile("mycmap.bin", "uint16").reshape((10000, 3))      # / 255.0
        self.colors4map  =  []
        for k in range(mycmap.shape[0]):
            self.colors4map.append(mycmap[k, :])
        self.colors4map[0]  =  np.array([0, 0, 0])

        framepp_raw  =  pg.ImageView(self, name="Frame_raw")
        framepp_raw.ui.roiBtn.hide()
        framepp_raw.ui.menuBtn.hide()
        framepp_raw.setImage(imarray_red)
        framepp_raw.timeLine.sigPositionChanged.connect(self.update_frame_from_raw)

        framepp_trk  =  pg.ImageView(self, name="framepp_trk")
        framepp_trk.ui.roiBtn.hide()
        framepp_trk.ui.menuBtn.hide()
        framepp_trk.setImage(nuclei_tracked)
        framepp_trk.timeLine.sigPositionChanged.connect(self.update_frame_from_trk)
        self.mycmap  =  pg.ColorMap(np.linspace(0, 1, nuclei_tracked.max()), color=self.colors4map)
        framepp_trk.setColorMap(self.mycmap)

        framepp_thr  =  pg.ImageView(self, name="framepp_thr")
        framepp_thr.ui.roiBtn.hide()
        framepp_thr.ui.menuBtn.hide()
        framepp_thr.setImage(nuclei_tracked)
        framepp_thr.timeLine.sigPositionChanged.connect(self.update_frame_from_thr)
        self.mycmap  =  pg.ColorMap(np.linspace(0, 1, nuclei_tracked.max()), color=self.colors4map)
        framepp_thr.setColorMap(self.mycmap)

        framepp_raw.view.setXLink("framepp_trk")
        framepp_raw.view.setYLink("framepp_trk")
        framepp_trk.view.setXLink("frame_raw")
        framepp_trk.view.setYLink("frame_raw")
        framepp_thr.view.setXLink("frame_raw")
        framepp_thr.view.setYLink("frame_raw")

        tabs     =  QtWidgets.QTabWidget()
        tab_raw  =  QtWidgets.QWidget()
        tab_trk  =  QtWidgets.QWidget()
        tab_thr  =  QtWidgets.QWidget()

        frame_raw_box  =  QtWidgets.QHBoxLayout()
        frame_raw_box.addWidget(framepp_raw)

        frame_trk_box  =  QtWidgets.QHBoxLayout()
        frame_trk_box.addWidget(framepp_trk)

        frame_thr_box  =  QtWidgets.QHBoxLayout()
        frame_thr_box.addWidget(framepp_thr)

        tab_raw.setLayout(frame_raw_box)
        tab_trk.setLayout(frame_trk_box)
        tab_thr.setLayout(frame_thr_box)

        tabs.addTab(tab_raw, "Raw")
        tabs.addTab(tab_trk, "Tracked Nuclei")
        tabs.addTab(tab_thr, "Thresholded Nuclei")

        area_thr_lbl  =  QtWidgets.QLabel(self)
        area_thr_lbl.setFixedSize(140, 25)
        area_thr_lbl.setText("Area Threshold")

        area_thr_edt  =  QtWidgets.QLineEdit(self)
        area_thr_edt.textChanged[str].connect(self.area_thr_var)
        area_thr_edt.returnPressed.connect(self.area_thr)
        area_thr_edt.setToolTip("Area threshold to delete small objects")
        area_thr_edt.setFixedSize(100, 25)

        send_btn  =  QtWidgets.QPushButton("Send", self)
        send_btn.clicked.connect(self.send)
        send_btn.setToolTip('Send results to the main GUI')
        send_btn.setFixedSize(110, 25)

        commands_box  =  QtWidgets.QHBoxLayout()
        commands_box.addWidget(area_thr_lbl)
        commands_box.addWidget(area_thr_edt)
        commands_box.addStretch()
        commands_box.addWidget(send_btn)

        layout  =  QtWidgets.QVBoxLayout()
        layout.addWidget(tabs)
        layout.addLayout(commands_box)

        self.setLayout(layout)
        self.setGeometry(300, 300, 900, 700)
        self.setWindowTitle("Remove Mitotical TS")

        self.framepp_raw  =  framepp_raw
        self.framepp_trk  =  framepp_trk
        self.framepp_thr  =  framepp_thr

        self.imarray_red     =  imarray_red
        self.nuclei_tracked  =  nuclei_tracked

    def update_frame_from_raw(self):
        """Update the trk and thr frame position from raw position"""
        self.framepp_trk.setCurrentIndex(self.framepp_raw.currentIndex)
        self.framepp_thr.setCurrentIndex(self.framepp_raw.currentIndex)

    def update_frame_from_trk(self):
        """Update the raw and thr frame position from trk position"""
        self.framepp_raw.setCurrentIndex(self.framepp_trk.currentIndex)
        self.framepp_thr.setCurrentIndex(self.framepp_trk.currentIndex)

    def update_frame_from_thr(self):
        """Update the raw and trk frame position from thr position"""
        self.framepp_raw.setCurrentIndex(self.framepp_thr.currentIndex)
        self.framepp_trk.setCurrentIndex(self.framepp_thr.currentIndex)

    def area_thr_var(self, text):
        """Set the threshold value"""
        self.area_thr_value  =  int(text)

    def area_thr(self):
        """Perform small objects removal"""
        cif  =  self.framepp_thr.currentIndex
        self.nucs_thr  =  RemoveBadNuclei.RemoveSmallNuclei(self.nuclei_tracked, self.area_thr_value).nuclei_thrd
        self.framepp_thr.setImage(self.nucs_thr)
        self.mycmap  =  pg.ColorMap(np.linspace(0, 1, self.nucs_thr.max()), color=self.colors4map)
        self.framepp_thr.setColorMap(self.mycmap)
        self.framepp_thr.setCurrentIndex(cif)

    @QtCore.pyqtSlot()
    def send(self):
        self.procStart.emit()


def except_hook(cls, exception, traceback):
    sys.__excepthook__(cls, exception, traceback)


if __name__ == "__main__":

    sys.excepthook  =  except_hook

    app     =  QtGui.QApplication(sys.argv)
    window  =  MainWindow()
    window.show()
    sys.exit(app.exec_())
