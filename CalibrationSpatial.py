""""This function calibrates the transcriptional time traces.

Starting from the escel file 'DividedByBackground_SpatialSelected',
the algorithm generates another excel in which all the values of the
first are divided by the calibration value.
"""


import datetime
import xlsxwriter
from openpyxl import load_workbook
from PyQt5 import QtWidgets


class CalibrationSpatial:
    """Only one class, does all the job"""
    def __init__(self, foldername, calib_value):

        exp_name  =  foldername[len(foldername) - foldername[:-1][::-1].find('/') - 1:]

        wb_starting  =  load_workbook(foldername + '/DividedByBackground_SpatialSelected.xlsx')
        s_starting   =  wb_starting[wb_starting.sheetnames[0]]

        book    =  xlsxwriter.Workbook(foldername + '/' + exp_name + '_' + 'CalibratedTraces.xlsx')
        sheet1  =  book.add_worksheet("Sheet 1")

        sheet1.write(1, 0, "date")
        sheet1.write(1, 1, datetime.datetime.now().strftime("%d-%b-%Y"))

        sheet1.write(3, 0, "Calib Value")
        sheet1.write(3, 1, calib_value)

        pbar      =  ProgressBar(total1=s_starting.max_column)
        pbar_idx  =  0
        pbar.show()

        for k in range(1, s_starting.max_column + 1):
            sheet1.write(0, k - 1, s_starting.cell(column=k, row=1).value)

        for kk_row in range(2, s_starting.max_row):
            sheet1.write(kk_row - 1, 2, s_starting.cell(column=3, row=kk_row).value)
            sheet1.write(kk_row - 1, 3, s_starting.cell(column=4, row=kk_row).value)

        for k_col in range(4, s_starting.max_column):
            pbar_idx  += 1
            pbar.update_progressbar(pbar_idx)
            for k_row in range(2, s_starting.max_row):
                sheet1.write(k_row - 1, k_col, s_starting.cell(column=k_col, row=k_row).value / calib_value)

        pbar.close()
        book.close()


class ProgressBar(QtWidgets.QWidget):
    """Simple progress bar widget"""
    def __init__(self, parent=None, total1=20):
        super().__init__(parent)
        self.name_line1  =  QtWidgets.QLineEdit()

        self.progressbar  =  QtWidgets.QProgressBar()
        self.progressbar.setMinimum(1)
        self.progressbar.setMaximum(total1)

        main_layout  =  QtWidgets.QGridLayout()
        main_layout.addWidget(self.progressbar, 0, 0)

        self.setLayout(main_layout)
        self.setWindowTitle("Progress")
        self.setGeometry(500, 300, 300, 50)

    def update_progressbar(self, val1):
        """Progress bar updater"""
        self.progressbar.setValue(val1)
        QtWidgets.qApp.processEvents()

