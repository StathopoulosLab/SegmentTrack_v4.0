"""This function writes bursting information into an excel file.

It takes nuclei matrix plus boundary of the region and filter settings as input
and collects and writes info about bursting in that region.
"""


import datetime
import numpy as np
import xlsxwriter
from openpyxl import load_workbook
from skimage.measure import regionprops
from skimage.morphology import label
from PyQt5 import QtWidgets

import SpotsFilter


class ComprehensiveBurstAnalysisWriter:
    """Only class, does all the job"""
    def __init__(self, foldername, nucs, x_pos, y_pos, width, height, filter_set, y_gastrulation, x_size, x_tile_coord, pix_size, coords2write):

        spots          =  np.load(foldername + '/spots_tracked.npy')
        spots_3D_ints  =  np.load(foldername + '/spots_3D_ints.npy')

        wb_b  =  load_workbook(foldername + '/journal.xlsx')
        s_b   =  wb_b[wb_b.sheetnames[0]]

        wb_b_db  =  load_workbook(foldername + '/SpotsIntensityDividedByBackground.xlsx')
        s_b_db   =  wb_b_db[wb_b_db.sheetnames[0]]

        def spts_db(spt_idx):
            """Function to isolate single nucleus trace"""
            ll  =  1
            # while ll < s_b_db.ncols and s_b_db.col(ll)[0].value != "SptId_" + str(spt_idx):
            while ll < s_b_db.max_column and s_b_db.cell(column=ll, row=1).value != "SptId_" + str(spt_idx):
                ll  +=  1
            spt  =  np.zeros(s_b_db.max_row - 1)
            for rr in range(spt.size):
                # spt[rr]  =  s_b_db.col(ll)[rr + 1].value
                spt[rr]  =  s_b_db.cell(column=ll, row=rr + 2).value
            return spt

        start  =  8
        # end    =  s_b.ncols - 1
        end    =  s_b.max_column - 1
        # while str(s_b.col(end)[0].value)[:5] != 'Spot_':                        # List of the tag of activating nuclei from the journal file
        while s_b.cell(column=end, row=1).value[:5] != 'Spot_':
            end  -=  1

        idxs  =  np.array([])
        for k in range(start, end):
            # idxs  =  np.append(idxs, int(s_b.col(k)[0].value[5:]))
            idxs  =  np.append(idxs, int(s_b.cell(column=k, row=1).value[5:]))

        av_coords  =  []

        pbar  =  ProgressBar(total1=idxs.size * 2)
        pbar.show()

        idxs_in   =  np.array([])
        pbar_idx  =  1
        for k in idxs:
            pbar.update_progressbar(pbar_idx)
            pbar_idx  +=  1
            # print k
            x_coords, y_coords  =  np.zeros((nucs.shape[0])), np.zeros((nucs.shape[0]))
            for t in range(nucs.shape[0]):
                if (nucs[t, :, :] == k).sum() > 0:
                    rgp                       =  regionprops((nucs[t, :, :] == k).astype(np.int))
                    x_coords[t], y_coords[t]  =  rgp[0]['centroid'][0], rgp[0]['centroid'][1]
            if x_pos < np.delete(x_coords, np.where(x_coords == 0)).mean() < x_pos + width and y_pos < np.delete(y_coords, np.where(y_coords == 0)).mean() < y_pos + height:           # Check if the nucleus is in the interested region
                idxs_in  =  np.append(idxs_in, k)
                av_coords.append([np.delete(x_coords, np.where(x_coords == 0)).mean(), np.delete(y_coords, np.where(y_coords == 0)).mean()])

        numb_nucs_area  =  len(idxs_in)

        numb_bursts      =  np.zeros(idxs_in.shape)                             # inizialitazion of matrices bigger than needed: they will be cut after (I don't know how big they should be now)
        numb_off         =  np.zeros(idxs_in.shape)
        bursts_duration  =  np.zeros((idxs_in.size, nucs.shape[0]))
        off_duration     =  np.zeros((idxs_in.size, nucs.shape[0]))
        frst_frame       =  np.zeros((idxs_in.size, nucs.shape[0]))
        bursts_ampl_mx   =  np.zeros((idxs_in.size, nucs.shape[0]))
        bursts_ampl_av   =  np.zeros((idxs_in.size, nucs.shape[0]))
        bursts_ampl_in   =  np.zeros((idxs_in.size, nucs.shape[0]))

        bursts_ampl_mx_db   =  np.zeros((idxs_in.size, nucs.shape[0]))          # db suffix mean 'divided by background'
        bursts_ampl_av_db   =  np.zeros((idxs_in.size, nucs.shape[0]))
        bursts_ampl_in_db   =  np.zeros((idxs_in.size, nucs.shape[0]))

        idxs2rm  =  []

        for j in range(idxs_in.size):
            pbar.update_progressbar(pbar_idx)
            pbar_idx  +=  1

            spt     =  spots == idxs_in[j]
            prof    =  np.sign((spots == idxs_in[j]).sum(2).sum(1))
            prof_f  =  SpotsFilter.SpotsFilter(prof, filter_set).prof_f              # filtering appends here, on the binary series
            # print(idxs_in[j])
            spt_db  =  spts_db(int(idxs_in[j]))

            if prof_f.sum() > 0:
                numb_bursts[j]  =  label(prof_f).max()

                if prof_f[0] == 0:
                    numb_off[j]  =  label(1 - prof_f).max() - 1
                else:
                    numb_off[j]  =  label(1 - prof_f).max()

                prof_f_lbl  =  label(prof_f)
                for k in np.unique(prof_f_lbl)[1:]:
                    bursts_duration[j, k - 1]  =  np.sum(prof_f_lbl == k)
                    frst_frame[j, k - 1]       =  np.where((prof_f_lbl == k) == 1)[0][0]
                    bursts_ampl_in[j, k - 1]   =  np.sum((prof_f_lbl == k) * (spt * spots_3D_ints).sum(2).sum(1))
                    bursts_ampl_mx[j, k - 1]   =  ((prof_f_lbl == k) * (spt * spots_3D_ints).sum(2).sum(1)).max()
                    bff_avg                    =  (prof_f_lbl == k) * (spt * spots_3D_ints).sum(2).sum(1)
                    bursts_ampl_av[j, k - 1]   =  bff_avg[bff_avg != 0].mean()
                    # print(spt_db)
                    # print(np.sum((prof_f_lbl == k) * spt_db))
                    bursts_ampl_in_db[j, k - 1]  =  np.sum((prof_f_lbl == k) * spt_db)
                    bursts_ampl_mx_db[j, k - 1]  =  ((prof_f_lbl == k) * spt_db).max()
                    bff_avg_db                   =  (prof_f_lbl == k) * spt_db
                    bursts_ampl_av_db[j, k - 1]  =  bff_avg_db[bff_avg_db != 0].mean()

                prof_f_neg_lbl  =  label(1 - prof_f)
                if prof_f_neg_lbl[0] == 1:
                    for k in range(2, prof_f_neg_lbl.max() + 1):
                        off_duration[j, k - 2]  =  np.sum(prof_f_neg_lbl == k)

                if prof_f_neg_lbl[0] == 0:
                    for k in range(1, prof_f_neg_lbl.max() + 1):
                        off_duration[j, k - 1]  =  np.sum(prof_f_neg_lbl == k)

            else:
                idxs2rm.append(j)

        idxs2rm  =  idxs2rm[::-1]                                               # I need to remove indexes starting from the last for a numeration question
        for jj in range(len(idxs2rm)):
            idxs_in          =  np.delete(idxs_in, idxs2rm[jj])
            numb_bursts      =  np.delete(numb_bursts, idxs2rm[jj])
            numb_off         =  np.delete(numb_off, idxs2rm[jj])
            bursts_duration  =  np.delete(bursts_duration, idxs2rm[jj], axis=0)
            off_duration     =  np.delete(off_duration, idxs2rm[jj], axis=0)
            frst_frame       =  np.delete(frst_frame, idxs2rm[jj], axis=0)
            bursts_ampl_mx   =  np.delete(bursts_ampl_mx, idxs2rm[jj], axis=0)
            bursts_ampl_av   =  np.delete(bursts_ampl_av, idxs2rm[jj], axis=0)
            bursts_ampl_in   =  np.delete(bursts_ampl_in, idxs2rm[jj], axis=0)

            bursts_ampl_mx_db   =  np.delete(bursts_ampl_mx_db, idxs2rm[jj], axis=0)
            bursts_ampl_av_db   =  np.delete(bursts_ampl_av_db, idxs2rm[jj], axis=0)
            bursts_ampl_in_db   =  np.delete(bursts_ampl_in_db, idxs2rm[jj], axis=0)

        jend  =  np.where(bursts_ampl_mx.sum(0) == 0)[0][0]

        bursts_duration  =  bursts_duration[:, :jend]
        off_duration     =  off_duration[:, :jend]
        frst_frame       =  frst_frame[:, :jend]
        bursts_ampl_mx   =  bursts_ampl_mx[:, :jend]
        bursts_ampl_av   =  bursts_ampl_av[:, :jend]
        bursts_ampl_in   =  bursts_ampl_in[:, :jend]

        numb_comp  =  np.max([bursts_ampl_mx.shape[1], off_duration.shape[1]])

        book      =  xlsxwriter.Workbook(foldername + '/ComprehensiveBurstingData.xlsx')
        sheet1    =  book.add_worksheet("Sheet 1")
        sheet2    =  book.add_worksheet("Divided by Bkg")

        sheet1.write(0, 0, 'FOLDER NAME')
        sheet1.write(0, 1, foldername)

        sheet1.write(2, 0, 'Region')
        sheet1.write(3, 0, 'X1')
        sheet1.write(4, 0, 'X2')
        sheet1.write(5, 0, 'Y1')
        sheet1.write(6, 0, 'Y2')

        sheet1.write(3, 1, coords2write[0])
        sheet1.write(4, 1, coords2write[1])
        sheet1.write(5, 1, coords2write[2])
        sheet1.write(6, 1, coords2write[3])

        sheet1.write(8, 0, 'Filter Parameters')
        sheet1.write(9, 0, 'Zeros')
        sheet1.write(10, 0, 'Ones')
        sheet1.write(9, 1, float(filter_set[0]))
        sheet1.write(10, 1, float(filter_set[1]))
        sheet1.write(11, 0, "Pix Size")
        sheet1.write(11, 1, pix_size)
        sheet1.write(12, 0, "Gastr-Pos")
        sheet1.write(12, 1, y_gastrulation)

        sheet1.write(15, 0, "Nuclei ID")
        sheet1.write(15, 1, "X coord")
        sheet1.write(15, 2, "Y coord")
        sheet1.write(15, 3, "Numb of Bursts")
        sheet1.write(15, 4, "Numb of Off")
        sheet1.write(15, 5, "Bursts durations")
        sheet1.write(15, 5 + numb_comp, "Off durations")
        sheet1.write(15, 5 + numb_comp * 2, "Busts av ampl")
        sheet1.write(15, 5 + numb_comp * 3, "Bursts integ ampl")
        sheet1.write(15, 5 + numb_comp * 4, "Bursts max integ")
        sheet1.write(15, 5 + numb_comp * 5, "First Act")

        sheet2.write(3, 1, coords2write[0])
        sheet2.write(4, 1, coords2write[1])
        sheet2.write(5, 1, coords2write[2])
        sheet2.write(6, 1, coords2write[3])

        sheet2.write(8, 0, 'Filter Parameters')
        sheet2.write(9, 0, 'Zeros')
        sheet2.write(10, 0, 'Ones')
        sheet2.write(9, 1, int(filter_set[0]))
        sheet2.write(10, 1, int(filter_set[1]))
        sheet2.write(11, 0, "Pix Size")
        sheet2.write(11, 1, pix_size)

        sheet2.write(15, 0, "Nuclei ID")
        sheet2.write(15, 1, "X coord")
        sheet2.write(15, 2, "Y coord")
        sheet2.write(15, 3, "Numb of Bursts")
        sheet2.write(15, 4, "Numb of Off")
        sheet2.write(15, 5, "Bursts durations")
        sheet2.write(15, 5 + numb_comp, "Off durations")
        sheet2.write(15, 5 + numb_comp * 2, "Busts av ampl")
        sheet2.write(15, 5 + numb_comp * 3, "Bursts integ ampl")
        sheet2.write(15, 5 + numb_comp * 4, "Bursts max integ")
        sheet2.write(15, 5 + numb_comp * 5, "First Act")

        for r in range(idxs_in.size):
            sheet1.write(r + 16, 0, "Nuc_" + str(int(idxs_in[r])))
            sheet2.write(r + 16, 0, "Nuc_" + str(int(idxs_in[r])))
            # sheet1.write(r + 16, 1, av_coords[r][0])
            sheet1.write(r + 16, 1, (av_coords[r][0] / x_size) * (x_tile_coord[-1] - x_tile_coord[0]) + x_tile_coord[0])
            sheet2.write(r + 16, 1, (av_coords[r][0] / x_size) * (x_tile_coord[-1] - x_tile_coord[0]) + x_tile_coord[0])
            sheet1.write(r + 16, 2, av_coords[r][1] - y_gastrulation)
            sheet2.write(r + 16, 2, av_coords[r][1] - y_gastrulation)
            sheet1.write(r + 16, 3, np.int(numb_bursts[r]))
            sheet2.write(r + 16, 3, np.int(numb_bursts[r]))
            sheet1.write(r + 16, 4, np.int(numb_off[r]))
            sheet2.write(r + 16, 4, np.int(numb_off[r]))

            for j in range(numb_comp):
                if bursts_duration[r, j] != 0:
                    sheet1.write(r + 16, 5 + j, bursts_duration[r, j])
                    sheet2.write(r + 16, 5 + j, bursts_duration[r, j])

            for j in range(numb_comp):
                if off_duration[r, j] != 0:
                    sheet1.write(r + 16, 5 + numb_comp + j, off_duration[r, j])
                    sheet2.write(r + 16, 5 + numb_comp + j, off_duration[r, j])

            for j in range(numb_comp):
                if bursts_ampl_av[r, j] != 0:
                    sheet1.write(r + 16, 5 + numb_comp * 2 + j, bursts_ampl_av[r, j])
                    sheet2.write(r + 16, 5 + numb_comp * 2 + j, bursts_ampl_av_db[r, j])

            for j in range(numb_comp):
                if bursts_ampl_in[r, j] != 0:
                    sheet1.write(r + 16, 5 + numb_comp * 3 + j, bursts_ampl_in[r, j])
                    sheet2.write(r + 16, 5 + numb_comp * 3 + j, bursts_ampl_in_db[r, j])

            for j in range(numb_comp):
                if bursts_ampl_mx[r, j] != 0:
                    sheet1.write(r + 16, 5 + numb_comp * 4 + j, bursts_ampl_mx[r, j])
                    sheet2.write(r + 16, 5 + numb_comp * 4 + j, bursts_ampl_mx_db[r, j])

            for j in range(numb_comp):
                if frst_frame[r, j] != 0:
                    sheet1.write(r + 16, 5 + numb_comp * 5 + j, frst_frame[r, j])
                    sheet2.write(r + 16, 5 + numb_comp * 5 + j, frst_frame[r, j])

        pbar.close()

        sheet1.write(r + 20, 0, 'Date')
        sheet1.write(r + 20, 1, datetime.datetime.now().strftime("%d-%b-%Y"))
        book.close()

        self.numb_nucs_area  =  numb_nucs_area
        self.idxs_in         =  idxs_in
        self.spots_trk       =  spots


class ProgressBar(QtWidgets.QWidget):
    """Simple progressbar widget"""
    def __init__(self, parent=None, total1=20):
        super().__init__(parent)
        self.name_line1  =  QtWidgets.QLineEdit()

        self.progressbar1  =  QtWidgets.QProgressBar()
        self.progressbar1.setMinimum(1)
        self.progressbar1.setMaximum(total1)

        main_layout  =  QtWidgets.QGridLayout()
        main_layout.addWidget(self.progressbar1, 0, 0)

        self.setLayout(main_layout)
        self.setWindowTitle("Progress")
        self.setGeometry(500, 300, 300, 50)

    def update_progressbar(self, val1):
        """Progress bar updater"""
        self.progressbar1.setValue(val1)
        QtWidgets.qApp.processEvents()
