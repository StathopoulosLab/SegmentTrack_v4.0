"""This function generate an excel file with the intensity divided by background in a pattern.

It loads and combines all the information of the excel file of the spatial analysis and
spots intensity divided by background and generate an excel with only the info of the
selected spots.
"""


import datetime
import xlsxwriter
from openpyxl import load_workbook
from PyQt5 import QtWidgets


class SpatialDividedByBkg:
    """Main class, does the all job"""
    def __init__(self, foldername, time_zero, time_step_value):

        wb_bybkg     =  load_workbook(foldername + '/SpotsIntensityDividedByBackground.xlsx')
        s_wb_bybkg1  =  wb_bybkg[wb_bybkg.sheetnames[0]]
        s_wb_bybkg2  =  wb_bybkg[wb_bybkg.sheetnames[1]]
        s_wb_bybkg3  =  wb_bybkg[wb_bybkg.sheetnames[2]]
        s_wb_bybkg4  =  wb_bybkg[wb_bybkg.sheetnames[3]]
        nrows_bybkg  =  s_wb_bybkg1.max_row

        wb_spat    =  load_workbook(foldername + '/ComprehensiveBurstingData.xlsx')
        s_wb_spat  =  wb_spat[wb_spat.sheetnames[0]]

        row_start  =  1
        while str(s_wb_spat.cell(column=1, row=row_start).value) != "Nuclei ID":
            row_start  +=  1

        spt_ids  =  []
        j        =  row_start + 1
        while str(s_wb_spat.cell(column=1, row=j).value)[:4] == "Nuc_":
            spt_ids.append(str(s_wb_spat.cell(column=1, row=j).value)[4:])
            j  +=  1

        col_coords  =  []
        for k in spt_ids:
            for l in range(1, s_wb_bybkg1.max_column):
                if str(s_wb_bybkg1.cell(column=l, row=1).value)[6:] == k:
                    col_coords.append(l)

        pbar  =  ProgressBar(total1=len(col_coords))
        pbar.show()

        book    =  xlsxwriter.Workbook(foldername + '/DividedByBackground_SpatialSelected.xlsx')
        sheet1  =  book.add_worksheet("Spots by Background")
        sheet2  =  book.add_worksheet("Average Background")
        sheet3  =  book.add_worksheet("Background Std")
        sheet4  =  book.add_worksheet("Average Spots by Background")

        sheet1.write(0, 0, str(s_wb_bybkg1.cell(column=1, row=1).value))
        sheet1.write(0, 1, str(s_wb_bybkg1.cell(column=2, row=1).value))
        sheet1.write(1, 0, "date")
        sheet1.write(1, 1, datetime.datetime.now().strftime("%d-%b-%Y"))
        sheet1.write(0, 2, "Frame")
        sheet1.write(0, 3, "Time")

        for t in range(1, nrows_bybkg):
            sheet1.write(t, 2, t - 1)
            sheet2.write(t, 2, t - 1)
            sheet3.write(t, 2, t - 1)
            sheet4.write(t, 2, t - 1)

            sheet1.write(t, 3, (t - 1 + time_zero) * time_step_value)
            sheet2.write(t, 3, (t - 1 + time_zero) * time_step_value)
            sheet3.write(t, 3, (t - 1 + time_zero) * time_step_value)
            sheet4.write(t, 3, (t - 1 + time_zero) * time_step_value)

        idx_wrt   =  4
        pbar_idx  =  1
        for k in col_coords:
            pbar.update_progressbar(pbar_idx)
            sheet1.write(0, idx_wrt, "SptId_" + str(spt_ids[idx_wrt - 4]))
            sheet2.write(0, idx_wrt, "SptId_" + str(spt_ids[idx_wrt - 4]))
            sheet3.write(0, idx_wrt, "SptId_" + str(spt_ids[idx_wrt - 4]))
            sheet4.write(0, idx_wrt, "SptId_" + str(spt_ids[idx_wrt - 4]))

            for o in range(1, nrows_bybkg + 1):
                sheet1.write(o, idx_wrt, s_wb_bybkg1.cell(column=k, row=o + 1).value)
                sheet2.write(o, idx_wrt, s_wb_bybkg2.cell(column=k, row=o + 1).value)
                sheet3.write(o, idx_wrt, s_wb_bybkg3.cell(column=k, row=o + 1).value)
                sheet4.write(o, idx_wrt, s_wb_bybkg4.cell(column=k, row=o + 1).value)

            idx_wrt   +=  1
            pbar_idx  +=  1

        pbar.close()
        book.close()


class ProgressBar(QtWidgets.QWidget):
    """This is a simple progress bar"""
    def __init__(self, parent=None, total1=20):
        super().__init__(parent)
        self.name_line1  =  QtWidgets.QLineEdit()

        self.progressbar1  =  QtWidgets.QProgressBar()
        self.progressbar1.setMinimum(1)
        self.progressbar1.setMaximum(total1)

        main_layout  =  QtWidgets.QGridLayout()
        main_layout.addWidget(self.progressbar1, 0, 0)

        self.setLayout(main_layout)
        self.setWindowTitle("Progress")
        self.setGeometry(500, 300, 300, 50)

    def update_progressbar(self, val1):
        """Method to update the progress bar"""
        self.progressbar1.setValue(val1)
        QtWidgets.qApp.processEvents()

