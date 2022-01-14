"""This function converts analysis run with version 3.3 into analysis done with version 4.0.

It works only on the bulk analysis, no spatial or similar.
"""


import numpy as np
import xlrd
import xlsxwriter
from openpyxl import load_workbook

import AnalysisLoader
import SpotsDetectionChopper
import ParametersExtraction
import SpotsConnection


def converjj_xls2xlsx(foldername):
    """This func converts xls file in xlsx"""
    book      =  xlrd.open_workbook(foldername + '/journal.xls')
    book2wrt  =  xlsxwriter.Workbook(foldername + "/journal.xlsx")                                                                  # write results

    sh_cr      =  book.sheet_by_index(0)
    sheet1_wr  =  book2wrt.add_worksheet()
    for k_vr in range(sh_cr.ncols):
        for k_ho in range(sh_cr.nrows):
            sheet1_wr.write(k_ho, k_vr, sh_cr.cell(k_ho, k_vr).value)
    sheet1_wr.write(22, 1, "SegmentTrackSingleCycleGUI_v4.0")
    sheet1_wr.write(16, 1, 0.5)
    sheet1_wr.write(16, 0, "Pixel Size Z")
    sheet1_wr.write(17, 0, "Frames Pre-Track")
    sheet1_wr.write(17, 1, sh_cr.cell(16, 1).value)

    book2wrt.close()


def converspts_dvd_bkg_xls2xlsx(foldername):
    """This func converts xls file in xlsx"""
    book      =  xlrd.open_workbook(foldername + '/SpotsIntensityDividedByBackground.xls')
    book2wrt  =  xlsxwriter.Workbook(foldername + "/SpotsIntensityDividedByBackground.xlsx")                                                                  # write results

    for nn in book.sheet_names():
        sh_cr      =  book.sheet_by_name(nn)
        sheet1_wr  =  book2wrt.add_worksheet(nn)
        for k_vr in range(sh_cr.ncols):
            for k_ho in range(sh_cr.nrows):
                sheet1_wr.write(k_ho, k_vr, sh_cr.cell(k_ho, k_vr).value)

    book2wrt.close()


def convers_burst_xls2xlsx(foldername):
    """This func converts xls file in xlsx"""
    book      =  xlrd.open_workbook(foldername + '/Bursts_Statistics.xls')
    book2wrt  =  xlsxwriter.Workbook(foldername + "/Bursts_Statistics.xlsx")                                                                  # write results

    for nn in book.sheet_names():
        sh_cr      =  book.sheet_by_name(nn)
        sheet1_wr  =  book2wrt.add_worksheet(nn)
        for k_vr in range(sh_cr.ncols):
            for k_ho in range(sh_cr.nrows):
                sheet1_wr.write(k_ho, k_vr, sh_cr.cell(k_ho, k_vr).value)

    book2wrt.close()


class AnalysisConverter:
    """Only class, does all the job"""
    def __init__(self, analysis_folder):

        nuclei_tracked  =  np.fromfile(analysis_folder + '/nuclei_tracked.bin', 'uint16')
        nuclei_tracked  =  nuclei_tracked[3:].reshape((nuclei_tracked[2], nuclei_tracked[1], nuclei_tracked[0]))
        np.save(analysis_folder + '/nuclei_tracked.npy', nuclei_tracked)

        spots_tracked  =  np.fromfile(analysis_folder + '/spots_tracked.bin', 'uint16')
        spots_tracked  =  spots_tracked[3:].reshape((spots_tracked[2], spots_tracked[1], spots_tracked[0]))
        np.save(analysis_folder + '/spots_tracked.npy', spots_tracked)

        converjj_xls2xlsx(analysis_folder)
        converspts_dvd_bkg_xls2xlsx(analysis_folder)
        convers_burst_xls2xlsx(analysis_folder)

        green4D           =  AnalysisLoader.RawData(analysis_folder).green4D
        wb                =  load_workbook(analysis_folder + '/journal.xlsx')
        spots_thr_value   =  wb.active["B7"].value
        volume_thr_value  =  wb.active["B8"].value
        max_dist          =  wb.active["B11"].value
        spots_3D          =  SpotsDetectionChopper.SpotsDetectionChopper(green4D, spots_thr_value, volume_thr_value)
        spots_tracked_3D  =  SpotsConnection.SpotsConnection(nuclei_tracked, np.sign(spots_3D.spots_vol), max_dist).spots_tracked
        features_3D       =  ParametersExtraction.ParametersExtraction(spots_3D.spots_ints, spots_tracked_3D, spots_3D.spots_vol)        # spots_3D.spots_vol * np.sign(self.spots_tracked_3D))

        np.save(analysis_folder + '/spots_3D_tzxy.npy', spots_3D.spots_tzxy.astype("uint16"))
        np.save(analysis_folder + '/spots_3D_vol.npy', spots_3D.spots_vol.astype("uint16"))
        np.save(analysis_folder + '/spots_3D_coords.npy', spots_3D.spots_coords.astype("uint16"))
        np.save(analysis_folder + '/spots_3D_ints.npy', spots_3D.spots_ints.astype("uint16"))
        np.save(analysis_folder + '/spots_features3D.npy', features_3D.statistics_info.astype(float))







# def converjj_xls2xlsx(foldername):
#     """This func converts xls file in xlsx"""
#     book      =  xlrd.open_workbook(foldername + '/journal.xls')
#     book2wrt  =  xlsxwriter.Workbook(foldername + "/journal.xlsx")                                                                  # write results
#
#     for nn in book.sheet_names():
#         sh_cr      =  book.sheet_by_name(nn)
#         sheet1_wr  =  book2wrt.add_worksheet(nn)
#         for k_vr in range(sh_cr.ncols):
#             for k_ho in range(sh_cr.nrows):
#                 if nn == "Sheet 1" and k_ho == 22 and k_vr == 1:
#                     sheet1_wr.write(k_ho, k_vr, "SegmentTrackSingleCycleGUI_v4.0")
#                 if nn == "Sheet 1" and k_ho == 16 and k_vr == 1:
#                     sheet1_wr.write(k_ho, k_vr, 0.5)
#                 if nn == "Sheet 1" and k_ho == 16 and k_vr == 0:
#                     sheet1_wr.write(k_ho, k_vr, "Pixel Size Z")
#                 # if nn == "Sheet 1" and k_ho == 17 and k_vr == 0:
#                 #     sheet1_wr.write(k_ho, k_vr, "Frames Pre-Track")
#                 # if nn == "Sheet 1" and k_ho == 17 and k_vr == 1:
#                 #     sheet1_wr.write(k_ho, k_vr, sh_cr.cell(k_ho - 1, k_vr).value)
#                 else:
#                     sheet1_wr.write(k_ho, k_vr, sh_cr.cell(k_ho, k_vr).value)
#
#     book2wrt.close()

